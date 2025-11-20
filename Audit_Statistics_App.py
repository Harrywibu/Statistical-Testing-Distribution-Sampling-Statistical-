from __future__ import annotations
import os, io, re, json, time, hashlib, contextlib, tempfile, warnings
from datetime import datetime
from typing import Optional, List, Callable, Dict, Any
import numpy as np
import pandas as pd
import streamlit as st
import inspect  # added for inspect.signature

def require_full_data(banner='Chưa có dữ liệu FULL. Hãy dùng **Load full data** trước khi chạy tab này.'):
    df = SS.get('df')
    import pandas as pd
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        st.info(banner); st.stop()
    return df
    
def _first_match(colnames, patterns):
    cols = [c for c in colnames]
    low = {c: str(c).lower() for c in cols}
    for p in patterns:
        p = p.lower()
        for c in cols:
            if p in low[c]:
                return c
    return None

def _decode_bytes_to_str(v):
    if isinstance(v, (bytes, bytearray)):
        for enc in ('utf-8','latin-1','cp1252'):
            try:
                return v.decode(enc, errors='ignore')
            except Exception:
                pass
        try:
            return v.hex()
        except Exception:
            return str(v)
    return v

def sanitize_for_arrow(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame):
        return df
    df = df.copy()
    obj_cols = df.select_dtypes(include=['object']).columns
    for c in obj_cols:
        col = df[c]
        if col.isna().all():
            continue
        # bytes -> str
        if col.map(lambda v: isinstance(v, (bytes, bytearray))).any():
            df[c] = col.map(_decode_bytes_to_str)
            col = df[c]
        try:
            sample = col.dropna().iloc[:1000]
        except Exception:
            sample = col.dropna()
        has_str = any(isinstance(x, str) for x in sample)
        has_num = any(isinstance(x, (int,float,np.integer,np.floating)) for x in sample)
        has_nested = any(isinstance(x, (dict,list,set,tuple)) for x in sample)
        if has_nested or (has_str and has_num):
            df[c] = col.astype(str)
    return df

# ---- Streamlit width compatibility wrappers ----
try:
    _df_params = inspect.signature(st.dataframe).parameters
    _df_supports_width = 'width' in _df_params
except Exception:
    _df_supports_width = False



def st_df(data=None, **kwargs):
    # Normalize width/use_container_width to avoid 'str' passed to Streamlit
    width = kwargs.pop('width', None)
    ucw = kwargs.pop('use_container_width', None)

    # width must be int (pixels); ignore strings like 'stretch'
    if isinstance(width, str):
        try:
            width = int(width)
        except Exception:
            width = None

    if width is not None and _df_supports_width:
        kwargs['width'] = int(width)
        if ucw is not None:
            kwargs['use_container_width'] = bool(ucw)
    else:
        # Fallback to use_container_width
        kwargs['use_container_width'] = True if ucw is None else bool(ucw)

    return st.dataframe(data, **kwargs)  # Không gọi lại st_df
# ====================== PATCH START: Export Capture Proxies ======================
# Đảm bảo SS đã được định nghĩa trước khi sử dụng
SS = st.session_state
# Lưu “bản gốc” của widget Streamlit lần đầu (để tránh wrap lặp gây đệ quy)
if '_orig_plotly_chart' not in SS:
    SS['_orig_plotly_chart'] = st.plotly_chart
if '_orig_dataframe' not in SS:
    SS['_orig_dataframe'] = st.dataframe
if '_orig_table' not in SS:
    SS['_orig_table'] = st.table

# ====================== PATCH END: Export Capture Proxies ======================

from scipy import stats

warnings.filterwarnings('ignore')

# Optional deps
try:
    import plotly.express as px
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    HAS_PLOTLY = True
except Exception:
    HAS_PLOTLY = False
try:
    import plotly.io as pio
    HAS_KALEIDO = True
except Exception:
    HAS_KALEIDO = False
try:
    import docx
    from docx.shared import Inches
    HAS_DOCX = True
except Exception:
    HAS_DOCX = False
try:
    import fitz  # PyMuPDF
    HAS_PDF = True
except Exception:
    HAS_PDF = False
try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    HAS_PYARROW = True
except Exception:
    HAS_PYARROW = False
try:
    from sklearn.model_selection import train_test_split
    from sklearn.linear_model import LinearRegression, LogisticRegression
    from sklearn.metrics import (
        r2_score, mean_squared_error, accuracy_score, roc_auc_score, roc_curve
    )
    HAS_SK = True
except Exception:
    HAS_SK = False

# --------------------------------- App Config ---------------------------------
st.set_page_config(page_title='Audit Statistics', layout='wide', initial_sidebar_state='expanded')
SS = st.session_state

DEFAULTS = {
    'bins': 50,
    'log_scale': False,
    'kde_threshold': 150_000,
    'risk_diff_threshold': 0.05,
    'advanced_visuals': False,
    'use_parquet_cache': False,
    'pv_n': 100,
    'df': None,
 'last_good_df': None,
    'df_preview': None,
 'last_good_preview': None,
    'file_bytes': None,
 'ingest_ready': False,
    'sha12': '',
    'uploaded_name': '',
        'xlsx_sheet': '',
    'header_row': 1,
    'skip_top': 0,
 'col_whitelist': None,
}
for k, v in DEFAULTS.items():
    SS.setdefault(k, v)
SS.setdefault('_export_registry', {})
# ------------------------------- Small Utilities ------------------------------
SS = st.session_state
if not isinstance(SS.get('_plt_seq'), int):
    SS['_plt_seq'] = 0
def file_sha12(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()[:12]

def st_plotly(fig, **kwargs):
    # Đảm bảo bộ đếm luôn là int
    seq = SS.get('_plt_seq')
    if not isinstance(seq, int):
        seq = 0
    seq += 1
    SS['_plt_seq'] = seq

    kwargs.setdefault('use_container_width', True)
    kwargs.setdefault('config', {'displaylogo': False})
    kwargs.setdefault('key', f'plt_{seq}')

    # Nếu plotly sẵn sàng thì vẽ; nếu không, thông báo nhẹ nhàng
    try:
        return st.plotly_chart(fig, **kwargs)
    except Exception as e:
        st.warning(f"Không render được Plotly chart: {e}")
        # (Tuỳ chọn) có thể thêm fallback matplotlib ở đây nếu bạn muốn
        return None

@st.cache_data(ttl=900, show_spinner=False, max_entries=64)
def corr_cached(df: pd.DataFrame, cols: List[str], method: str = 'pearson') -> pd.DataFrame:
    if not cols: return pd.DataFrame()
    sub = df[cols].apply(pd.to_numeric, errors='coerce')
    sub = sub.dropna(axis=1, how='all')
    nunique = sub.nunique(dropna=True)
    keep = [c for c in sub.columns if nunique.get(c, 0) > 1]
    sub = sub[keep]
    if sub.shape[1] < 2: return pd.DataFrame()
    return sub.corr(method=method)

def is_datetime_like(colname: str, s: pd.Series) -> bool:
    return pd.api.types.is_datetime64_any_dtype(s) or bool(re.search(r'(date|time)', str(colname), re.I))

def _downcast_numeric(df: pd.DataFrame) -> pd.DataFrame:
    for c in df.select_dtypes(include=['float64']).columns:
        df[c] = pd.to_numeric(df[c], downcast='float')
    for c in df.select_dtypes(include=['int64']).columns:
        df[c] = pd.to_numeric(df[c], downcast='integer')
    return df

def to_float(x) -> Optional[float]:
    from numbers import Real
    try:
        if isinstance(x, Real): return float(x)
        if x is None: return None
        return float(str(x).strip().replace(',', ''))
    except Exception:
        return None

# ------------------------------- Disk Cache I/O --------------------------------
def _parquet_cache_path(sha: str, key: str) -> str:
    return os.path.join(tempfile.gettempdir(), f'astats_cache_{sha}_{key}.parquet')

@st.cache_data(ttl=6*3600, show_spinner=False, max_entries=24)
def write_parquet_cache(df: pd.DataFrame, sha: str, key: str) -> str:
    if not HAS_PYARROW: return ''
    try:
        df = sanitize_for_arrow(df)
        table = pa.Table.from_pandas(df)
        path = _parquet_cache_path(sha, key)
        pq.write_table(table, path)
        return path
    except Exception:
        return ''

def read_parquet_cache(sha: str, key: str) -> Optional[pd.DataFrame]:
    if not HAS_PYARROW: return None
    path = _parquet_cache_path(sha, key)
    if os.path.exists(path):
        try:
            return pq.read_table(path).to_pandas()
        except Exception:
            return None
    return None

# ------------------------------- Fast Readers ---------------------------------
@st.cache_data(ttl=6*3600, show_spinner=False, max_entries=16)
def list_sheets_xlsx(file_bytes: bytes) -> List[str]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()

@st.cache_data(ttl=6*3600, show_spinner=False, max_entries=16)
def read_csv_fast(file_bytes: bytes, usecols=None) -> pd.DataFrame:
    bio = io.BytesIO(file_bytes)
    # --- FASTEST PATH: Polars ---
    try:
        import polars as pl
        # Đọc Polars và chuyển sang Pandas (zero-copy nếu có Arrow)
        df = pl.read_csv(bio, columns=usecols).to_pandas(use_pyarrow_extension_array=True)
        return _downcast_numeric(df)
    except Exception:
        bio.seek(0)
        # --- FALLBACK PATH: Pandas/PyArrow ---
        try:
            df = pd.read_csv(bio, usecols=usecols, engine='pyarrow')
        except Exception:
            bio.seek(0)
            df = pd.read_csv(bio, usecols=usecols, low_memory=False, memory_map=True)
        return _downcast_numeric(df)

@st.cache_data(ttl=6*3600, show_spinner=False, max_entries=16)
def read_xlsx_fast(file_bytes: bytes, sheet: str, usecols=None,
                   header_row: int = 1, skip_top: int = 0, dtype_map=None) -> pd.DataFrame:
    # --- sanitize input để tránh None gây lỗi so sánh/số học ---
    header_row = 1 if header_row in (None, 0, '', False) else int(header_row)
    skip_top   = 0 if skip_top   in (None, '', False)   else int(skip_top)

    # --- FAST PATH: DuckDB excel extension (rất nhanh nếu sẵn có) ---
    if 'HAS_DUCKDB' in globals() and HAS_DUCKDB:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        try:
            tmp.write(file_bytes); tmp.flush(); tmp.close()
            con = duckdb.connect()
            # excel ext có thể đã có sẵn; nếu không, nhảy qua except và fallback
            try:
                con.execute("INSTALL excel; LOAD excel;")
            except Exception:
                pass

            # Đọc toàn sheet về Arrow/Pandas
            q = f"SELECT * FROM read_excel('{tmp.name}', sheet='{sheet}')"
            try:
                # ưu tiên lấy Arrow table nếu pyarrow sẵn để convert nhanh
                import pyarrow as pa
                table = con.execute(q).arrow()
                pdf = table.to_pandas(types_mapper=pd.ArrowDtype) if hasattr(pd, 'ArrowDtype') else table.to_pandas()
            except Exception:
                pdf = con.execute(q).df()

            # Mô phỏng hành vi header/skip như pandas.read_excel
            header_idx = max(header_row - 1, 0)
            # bảo vệ khi sheet trống hoặc chỉ có vài dòng
            if len(pdf) == 0:
                return _downcast_numeric(pdf)

            pdf.columns = pdf.iloc[header_idx].astype(str).tolist()
            start_data = header_idx + 1 + max(skip_top, 0)
            pdf = pdf.iloc[start_data:].reset_index(drop=True)

            # Chọn cột nếu usecols được truyền
            if usecols:
                cols_sel = [c for c in usecols if c in pdf.columns]
                if cols_sel:
                    pdf = pdf[cols_sel]

            # Áp dtype_map nếu có
            if dtype_map:
                for c, t in dtype_map.items():
                    if c in pdf.columns:
                        with contextlib.suppress(Exception):
                            pdf[c] = pdf[c].astype(t)

            return _downcast_numeric(pdf)
        finally:
            with contextlib.suppress(Exception):
                os.unlink(tmp.name)

    # --- FALLBACK: pandas + openpyxl
    skiprows = list(range(header_row, header_row + skip_top)) if skip_top > 0 else None
    bio = io.BytesIO(file_bytes)
    df = pd.read_excel(
        bio, sheet_name=sheet, usecols=usecols,
        header=header_row - 1, skiprows=skiprows,
        dtype=dtype_map, engine='openpyxl'
    )
    return _downcast_numeric(df)


# ----------------------------- Cached Basic Stats -----------------------------
@st.cache_data(ttl=1800, show_spinner=False, max_entries=64)
def numeric_profile_stats(series: pd.Series):
    s = pd.to_numeric(series, errors='coerce').replace([np.inf, -np.inf], np.nan).dropna()
    desc = s.describe(percentiles=[0.01,0.05,0.10,0.25,0.5,0.75,0.90,0.95,0.99])
    skew = float(stats.skew(s)) if len(s) > 2 else np.nan
    kurt = float(stats.kurtosis(s, fisher=True)) if len(s) > 3 else np.nan
    try:
        p_norm = float(stats.normaltest(s)[1]) if len(s) > 7 else np.nan
    except Exception:
        p_norm = np.nan
    p95 = s.quantile(0.95) if len(s)>1 else np.nan
    p99 = s.quantile(0.99) if len(s)>1 else np.nan
    zero_ratio = float((s==0).mean()) if len(s)>0 else np.nan
    return desc.to_dict(), skew, kurt, p_norm, float(p95), float(p99), zero_ratio

@st.cache_data(ttl=1800, show_spinner=False, max_entries=64)
def cat_freq(series: pd.Series) -> pd.DataFrame:
    s = series.dropna().astype(str)
    vc = s.value_counts(dropna=True)
    out = pd.DataFrame({'category': vc.index, 'count': vc.values})
    out['share'] = out['count']/out['count'].sum()
    return out


# ------------------------------ Benford Helpers -------------------------------
@st.cache_data(ttl=3600, show_spinner=False, max_entries=64)
def _benford_1d(series: pd.Series):
    s = pd.to_numeric(series, errors='coerce').replace([np.inf, -np.inf], np.nan).dropna().abs()
    if s.empty: return None
    def _digits(x):
        xs = ("%.15g" % float(x))
        return re.sub(r"[^0-9]", "", xs).lstrip("0")
    d1 = s.apply(lambda v: int(_digits(v)[0]) if len(_digits(v))>=1 else np.nan).dropna()
    d1 = d1[(d1>=1)&(d1<=9)]
    if d1.empty: return None
    obs = d1.value_counts().sort_index().reindex(range(1,10), fill_value=0).astype(float)
    n=obs.sum(); obs_p=obs/n
    idx=np.arange(1,10); exp_p=np.log10(1+1/idx); exp=exp_p*n
    with np.errstate(divide='ignore', invalid='ignore'):
        chi2=np.nansum((obs-exp)**2/exp)
        pval=1-stats.chi2.cdf(chi2, len(idx)-1)
        mad=float(np.mean(np.abs(obs_p-exp_p)))
        var_tbl=pd.DataFrame({'digit':idx,'expected':exp,'observed':obs.values})
        var_tbl['diff']=var_tbl['observed']-var_tbl['expected']
        var_tbl['diff_pct']=(var_tbl['observed']-var_tbl['expected'])/var_tbl['expected']
        table=pd.DataFrame({'digit':idx,'observed_p':obs_p.values,'expected_p':exp_p})
    return {'table':table, 'variance':var_tbl, 'n':int(n), 'chi2':float(chi2), 'p':float(pval), 'MAD':float(mad)}

@st.cache_data(ttl=3600, show_spinner=False, max_entries=64)
def _benford_2d(series: pd.Series):
    s = pd.to_numeric(series, errors='coerce').replace([np.inf, -np.inf], np.nan).dropna().abs()
    if s.empty: return None
    def _digits(x):
        xs = ("%.15g" % float(x))
        return re.sub(r"[^0-9]", "", xs).lstrip("0")
    def _first2(v):
        ds = _digits(v)
        if len(ds)>=2: return int(ds[:2])
        if len(ds)==1 and ds!="0": return int(ds)
        return np.nan
    d2 = s.apply(_first2).dropna(); d2=d2[(d2>=10)&(d2<=99)]
    if d2.empty: return None
    obs = d2.value_counts().sort_index().reindex(range(10,100), fill_value=0).astype(float)
    n=obs.sum(); obs_p=obs/n
    idx=np.arange(10,100); exp_p=np.log10(1+1/idx); exp=exp_p*n
    with np.errstate(divide='ignore', invalid='ignore'):
        chi2=np.nansum((obs-exp)**2/exp)
        pval=1-stats.chi2.cdf(chi2, len(idx)-1)
    mad=float(np.mean(np.abs(obs_p-exp_p)))
    var_tbl=pd.DataFrame({'digit':idx,'expected':exp,'observed':obs.values})
    var_tbl['diff']=var_tbl['observed']-var_tbl['expected']
    var_tbl['diff_pct']=(var_tbl['observed']-var_tbl['expected'])/var_tbl['expected']
    table=pd.DataFrame({'digit':idx,'observed_p':obs_p.values,'expected_p':exp_p})
    return {'table':table, 'variance':var_tbl, 'n':int(n), 'chi2':float(chi2), 'p':float(pval), 'MAD':float(mad)}

def _benford_ready(series: pd.Series) -> tuple[bool, str]:
    s = pd.to_numeric(series, errors='coerce')
    n_nz = int((s != 0).sum())  # nhận cả số âm, chỉ loại 0
    if n_nz < 1:
        return False, f"Không có giá trị ≠ 0 để chạy Benford (hiện {n_nz}, cần ≥300)."
    s_non = s.dropna()
    if s_non.shape[0] > 0:
        ratio_unique = s_non.nunique()/s_non.shape[0]
        if ratio_unique > 0.95:
            return False, "Tỉ lệ unique quá cao (khả năng ID/Code) — tránh Benford."
    return True, ''

def _plot(fig):
    try:
        st_plotly(fig)
    except Exception:
        st.plotly_chart(fig, use_container_width=True)

def guess_datetime_cols(df, check=3000):
    import numpy as np, pandas as pd
    sample = df.head(check)
    cols = []
    for c in df.columns:
        try:
            if np.issubdtype(df[c].dtype, np.datetime64):
                cols.append(c); continue
            if df[c].dtype == 'object':
                s = pd.to_datetime(sample[c], errors='coerce')
                if s.notna().mean() >= 0.5:
                    cols.append(c)
        except Exception:
            pass
    return cols

# -------------------------- Sidebar: Workflow & perf ---------------------------
up = st.file_uploader('Upload file (.csv, .xlsx)', type=['csv','xlsx'], key='ingest')
if up is not None:
    fb = up.read()  # có thể dùng up.getvalue() cũng được
    new_sha = file_sha12(fb)
    same_file = (SS.get('sha12') == new_sha) and (SS.get('uploaded_name') == up.name)

    # luôn cập nhật metadata/bytes để các bước sau dùng
    SS['file_bytes'] = fb
    SS['uploaded_name'] = up.name
    SS['sha12'] = new_sha

    # 🔒 CHỈ khi đổi file mới reset preview/full
    if not same_file:
        SS['df'] = None
        SS['df_preview'] = None

    st.caption(f"Đã nhận file: {up.name} • SHA12={SS['sha12']}")

    if st.button('Clear file', key='btn_clear_file'):
        base_keys = ['file_bytes','uploaded_name','sha12','df','df_preview','col_whitelist']
        result_keys = [
            'bf1_res','bf2_res','bf1_col','bf2_col','t4_results','last_corr','last_linear',
            'last_logistic','last_numeric_profile','last_gof','fraud_flags','spearman_recommended',
            '_plt_seq','col_filter','dtype_choice','xlsx_sheet','header_row','skip_top',
            'ingest_ready','last_good_df','last_good_preview'
        ]
        # đặt tên biến khác nhau để tránh đè 'k'
        for bk in base_keys:
            SS[bk] = DEFAULTS.get(bk, None)

        for rk in result_keys:
            if rk in SS:
                SS[rk] = None

        st.rerun()
with st.sidebar.expander('3) Cache', expanded=False):
    if not HAS_PYARROW:
        try:
            import duckdb
            HAS_DUCKDB = True
        except Exception:
            HAS_DUCKDB = False
        st.caption('⚠️ PyArrow chưa sẵn sàng — Disk cache (Parquet) sẽ bị tắt.')
        SS['use_parquet_cache'] = False
    SS['use_parquet_cache'] = st.checkbox('Disk cache (Parquet) for faster reloads', value=SS.get('use_parquet_cache', False) and HAS_PYARROW)
    if st.button('🧹 Clear cache'):
        st.cache_data.clear(); st.toast('Cache cleared', icon='🧹')

# ---------------------------------- Main Gate ---------------------------------
st.title('📊 Audit Statistics')
if SS['file_bytes'] is None:
    st.info('Upload a file để bắt đầu.'); st.stop()

fname=SS['uploaded_name']; fb=SS['file_bytes']; sha=SS['sha12']
colL, colR = st.columns([3,2])
with colL:
    st.text_input('File', value=fname or '', disabled=True)
with colR:
    SS['pv_n'] = st.slider('Preview rows', 50, 500, SS.get('pv_n',100), 50)
    do_preview = st.button('🔎 Quick preview', key='btn_prev')

# Ingest flow
if fname.lower().endswith('.csv'):
    if do_preview or SS['df_preview'] is None:
        try:
            SS['df_preview'] = sanitize_for_arrow(read_csv_fast(fb).head(SS['pv_n']))
            SS['last_good_preview'] = SS['df_preview']; SS['ingest_ready']=True
        except Exception as e:
            st.error(f'Lỗi đọc CSV: {e}'); SS['df_preview']=None
    if SS['df_preview'] is not None:
        st_df(SS['df_preview'], use_container_width=True, height=260)
        headers=list(SS['df_preview'].columns)
        selected = st.multiselect('Columns to load', headers, default=headers)
        SS['col_whitelist'] = selected if selected else headers
        if st.button('📥 Load full CSV with selected columns', key='btn_load_csv'):
            sel_key=';'.join(selected) if selected else 'ALL'
            key=f"csv_{hashlib.sha1(sel_key.encode()).hexdigest()[:10]}"
            df_cached = read_parquet_cache(sha, key) if SS['use_parquet_cache'] else None
            if df_cached is None:
                df_full = sanitize_for_arrow(read_csv_fast(fb, usecols=(selected or None)))
                if SS['use_parquet_cache']: write_parquet_cache(df_full, sha, key)
            else:
                df_full = df_cached
            SS['df']=df_full; SS['last_good_df']=df_full; SS['ingest_ready']=True; SS['col_whitelist']=list(df_full.columns)
            st.success(f"Loaded: {len(SS['df']):,} rows × {len(SS['df'].columns)} cols • SHA12={sha}")
else:
    sheets = list_sheets_xlsx(fb)
    with st.expander('📁 Select sheet & header (XLSX)', expanded=True):
        c1,c2,c3 = st.columns([2,1,1])
        idx=0 if sheets else 0
        SS['xlsx_sheet'] = c1.selectbox('Sheet', sheets, index=idx)
        SS['header_row'] = c2.number_input('Header row (1‑based)', 1, 100, SS['header_row'])
        SS['skip_top'] = c3.number_input('Skip N rows after header', 0, 1000, SS['skip_top'])
        SS['dtype_choice'] = st.text_area('dtype mapping (JSON, optional)', SS.get('dtype_choice',''), height=68)
        dtype_map = None
        if (SS.get('dtype_choice') or '').strip():
            try:
                dtype_map = json.loads(SS['dtype_choice'])
            except Exception as e:
                st.warning(f"Dtype mapping JSON không hợp lệ: {e}")
        try:
            prev = sanitize_for_arrow(read_xlsx_fast(fb, SS['xlsx_sheet'], usecols=None, header_row=SS['header_row'], skip_top=SS['skip_top'], dtype_map=dtype_map).head(SS['pv_n']))
            SS['df_preview']=prev; SS['last_good_preview']=prev; SS['ingest_ready']=True
        except Exception as e:
            st.error(f'Lỗi đọc XLSX: {e}'); prev=pd.DataFrame()
        st_df(prev, use_container_width=True, height=260)
        headers=list(prev.columns)
        st.caption(f'Columns: {len(headers)} • SHA12={sha}')
        SS['col_filter'] = st.text_input('🔎 Filter columns', SS.get('col_filter',''))
        filtered = [h for h in headers if SS['col_filter'].lower() in h.lower()] if SS['col_filter'] else headers
        selected = st.multiselect('🧮 Columns to load', filtered if filtered else headers, default=filtered if filtered else headers)
        if st.button('📥 Load full data', key='btn_load_xlsx'):
            key_tuple=(SS['xlsx_sheet'], SS['header_row'], SS['skip_top'], tuple(selected) if selected else ('ALL',))
            key=f"xlsx_{hashlib.sha1(str(key_tuple).encode()).hexdigest()[:10]}"
            df_cached = read_parquet_cache(sha, key) if SS['use_parquet_cache'] else None
            if df_cached is None:
                df_full = sanitize_for_arrow(read_xlsx_fast(fb, SS['xlsx_sheet'], usecols=(selected or None), header_row=SS['header_row'], skip_top=SS['skip_top'], dtype_map=dtype_map))
                if SS['use_parquet_cache']: write_parquet_cache(df_full, sha, key)
            else:
                df_full = df_cached
            SS['df']=df_full; SS['last_good_df']=df_full; SS['ingest_ready']=True; SS['col_whitelist']=list(df_full.columns)
            st.success(f"Loaded: {len(SS['df']):,} rows × {len(SS['df'].columns)} cols • SHA12={sha}")

if SS['df'] is None and SS['df_preview'] is None:
    st.stop()

# Source & typing
DF_FULL = require_full_data('Chưa có dữ liệu. Hãy dùng **Load full data**.')
DF_VIEW = DF_FULL  # alias để không phá code cũ

ALL_COLS = list(DF_FULL.columns)
DT_COLS  = [c for c in ALL_COLS if is_datetime_like(c, DF_FULL[c])]
NUM_COLS = DF_FULL[ALL_COLS].select_dtypes(include=[np.number]).columns.tolist()
CAT_COLS = DF_FULL[ALL_COLS].select_dtypes(include=['object','category','bool']).columns.tolist()
VIEW_COLS = [c for c in DF_FULL.columns if (not SS.get('col_whitelist') or c in SS['col_whitelist'])]


@st.cache_data(ttl=900, show_spinner=False, max_entries=64)
def spearman_flag(df: pd.DataFrame, cols: List[str]) -> bool:
    try:
        if df is None or not isinstance(df, pd.DataFrame):
            return False
    except Exception:
        return False

    for c in (cols or [])[:20]:
        if c not in df.columns:
            continue

        s = pd.to_numeric(df[c], errors='coerce').replace([np.inf, -np.inf], np.nan).dropna()
        if len(s) < 50:
            continue

        sk, ku, tail, p_norm = 0.0, 0.0, 0.0, 1.0  # defaults

        try:
            if len(s) > 2:
                sk = float(stats.skew(s))
        except Exception:
            pass

        try:
            if len(s) > 3:
                ku = float(stats.kurtosis(s, fisher=True))
        except Exception:
            pass

        try:
            p99 = s.quantile(0.99)
            if pd.notna(p99):
                tail = float((s > p99).mean())
        except Exception:
            pass

        try:
            if len(s) > 20:
                p_norm = float(stats.normaltest(s)[1])
        except Exception:
            pass

        if (abs(sk) > 1) or (abs(ku) > 3) or (tail > 0.02) or (p_norm < 0.05):
            return True
    return False
# ------------------------------ Rule Engine Core ------------------------------
class Rule:
    def __init__(self, id: str, name: str, scope: str, severity: str,
                 condition: Callable[[Dict[str,Any]], bool],
                 action: str, rationale: str):
        self.id=id; self.name=name; self.scope=scope; self.severity=severity
        self.condition=condition; self.action=action; self.rationale=rationale

    def eval(self, ctx: Dict[str,Any]) -> Optional[Dict[str,Any]]:
        try:
            if self.condition(ctx):
                return {
                    'id': self.id,
                    'name': self.name,
                    'scope': self.scope,
                    'severity': self.severity,
                    'action': self.action,
                    'rationale': self.rationale,
                }
        except Exception:
            return None
        return None

SEV_ORDER = {'High':3,'Medium':2,'Low':1,'Info':0}

def _get(ctx: Dict[str,Any], *keys, default=None):
    cur = ctx
    for k in keys:
        if cur is None: return default
        cur = cur.get(k) if isinstance(cur, dict) else None
    return cur if cur is not None else default

def build_rule_context() -> Dict[str,Any]:
    ctx = {
        'thr': {
            'benford_diff': SS.get('risk_diff_threshold', 0.05),
            'zero_ratio': 0.30,
            'tail_p99': 0.02,
            'off_hours': 0.15,
            'weekend': 0.25,
            'corr_high': 0.9,
            'gap_p95_hours': 12.0,
            'hhi': 0.2,
        },
        'last_numeric': SS.get('last_numeric_profile'),
        'gof': SS.get('last_gof'),
        'benford': {
            'r1': SS.get('bf1_res'),
            'r2': SS.get('bf2_res')
        },
        't4': SS.get('t4_results'),
        'corr': SS.get('last_corr'),
        'regression': {
            'linear': SS.get('last_linear'),
            'logistic': SS.get('last_logistic'),
        },
        'flags': SS.get('fraud_flags') or [],
    }
    # convenience derivations
    r1 = ctx['benford'].get('r1')
    if r1 and isinstance(r1, dict) and 'variance' in r1:
        try:
            ctx['benford']['r1_maxdiff'] = float(r1['variance']['diff_pct'].abs().max())
        except Exception:
            ctx['benford']['r1_maxdiff'] = None
    r2 = ctx['benford'].get('r2')
    if r2 and isinstance(r2, dict) and 'variance' in r2:
        try:
            ctx['benford']['r2_maxdiff'] = float(r2['variance']['diff_pct'].abs().max())
        except Exception:
            ctx['benford']['r2_maxdiff'] = None
    return ctx


# ----------------------------------- TABS -------------------------------------
TAB0, TAB1, TAB2, TAB3, TAB4, TAB5, TAB6, TAB7 = st.tabs([ '0) Data Quality', '1) Overview (Sales activity)', '2) Profiling/Distribution', '3) Correlation & Trend', '4) Benford', '5) ANOVA & Nonparametric', '6) Regression','7) Pareto (80/20)'])
# ---- TAB 0: Data Quality (FULL) ----

with TAB0:
    st.subheader('🧪 Data Quality')
    if SS.get('df') is None:
        st.info('Hãy **Load full data** để xem Data Quality Tab.')
    else:
        @st.cache_data(ttl=900, show_spinner=False, max_entries=16)
        def data_quality_table(df_in):
            import pandas as pd
            rows = []
            n = len(df_in)
            for c in df_in.columns:
                s = df_in[c]
                is_num = pd.api.types.is_numeric_dtype(s)
                is_dt  = pd.api.types.is_datetime64_any_dtype(s) or is_datetime_like(c, s)
                is_bool= pd.api.types.is_bool_dtype(s)
                is_cat = pd.api.types.is_categorical_dtype(s)
                base_type = 'Numeric' if is_num else ('Datetime' if is_dt else ('Boolean' if is_bool else ('Categorical' if is_cat else 'Text')))
                n_nonnull = int(s.notna().sum())
                n_nan = int(n - n_nonnull)
                n_unique = int(s.nunique(dropna=True))
                mem_mb = float(s.memory_usage(deep=True)) / 1048576.0
                blank = None; blank_pct = None
                zero = None; zero_pct = None
                if base_type in ('Text','Categorical'):
                    s_txt = s[s.notna()].astype(str).str.strip()
                    blank = int((s_txt == '').sum())
                    blank_pct = round(blank / n, 4) if n else None
                if base_type == 'Numeric':
                    s_num = pd.to_numeric(s, errors='coerce')
                    zero = int(s_num.eq(0).sum())
                    zero_pct = round(zero / n, 4) if n else None
                valid = n_nonnull - (blank or 0) if base_type in ('Text','Categorical') else n_nonnull
                rows.append({
                    'column': c,
                    'type': base_type,
                    'rows': n,
                    'valid': int(valid),
                    'valid%': round(valid / n, 4) if n else None,
                    'nan': n_nan,
                    'nan%': round(n_nan / n, 4) if n else None,
                    'blank': blank,
                    'blank%': blank_pct,
                    'zero': zero,
                    'zero%': zero_pct,
                    'unique': n_unique,
                    'memory_MB': round(mem_mb, 3),
                })
            cols_order = ['column','type','rows','valid','valid%','nan','nan%','blank','blank%','zero','zero%','unique','memory_MB']
            dq = pd.DataFrame(rows)
            dq = dq[cols_order]
            return dq.sort_values(['type','column']).reset_index(drop=True)
        try:
            dq = data_quality_table(SS['df'] if SS.get('df') is not None else DF_VIEW)
            st_df(dq, use_container_width=True, height=min(520, 60 + 24*min(len(dq), 18)))
        except Exception as e:
            st.error(f'Lỗi Data Quality: {e}')

# ============================== TAB 1 — OVERVIEW (Sales Activities) ==============================
with TAB1:
    import numpy as np, pandas as pd
    import plotly.express as px
    import plotly.graph_objects as go
    import textwrap # Cần import textwrap

    st.subheader("📈 Overview — Sales Activities")
    # ---------- helpers ----------
    RULE = {"Month":"MS","Quarter":"QS","Year":"YS"}
    P2   = {"MS":"M","QS":"Q","YS":"Y"}
    YOY  = {"MS":12,"QS":4,"YS":1}

    def _clean_time(ts, min_year=1900, max_year=2100):
        t = pd.to_datetime(ts, errors="coerce")
        bad = t.notna() & ((t.dt.year < min_year) | (t.dt.year > max_year))
        return t.mask(bad)

    @st.cache_data(ttl=900, show_spinner="Đang tổng hợp...")
    def _agg_by_period(series_time, values, rule_code):
        t = _clean_time(series_time)
        m = t.notna()
        p = t.dt.to_period({"MS":"M","QS":"Q","YS":"Y"}[rule_code]).dt.start_time
        return (pd.DataFrame({"p": p[m], "v": values[m]})
                .groupby("p", dropna=False)["v"].sum().sort_index())

    def _wrap_label(lbl, width=16):
        s = "" if lbl is None else str(lbl)
        return "<br>".join(textwrap.wrap(s, width=width)) if len(s) > width else s

    def _pie_with_smart_labels(labels, values, colors, height=460):
        labels_wrapped = [_wrap_label(l, 16) for l in labels]
        max_len = max((len(str(l)) for l in labels), default=0)
        side_margin = 80 if max_len <= 14 else 120 if max_len <= 22 else 160
        show_text = len(labels) <= 14
        fig = go.Figure(go.Pie(
            labels=labels_wrapped, values=values, hole=0.35, sort=False,
            textposition="outside",
            texttemplate="%{label}<br>%{percent:.1%}" if show_text else "%{percent:.1%}",
            hovertemplate="%{label}<br>%{percent:.1%} (%{value:,.0f})<extra></extra>"
        ))
        fig.update_traces(marker=dict(colors=colors), automargin=True)
        fig.update_layout(
            margin=dict(l=side_margin, r=side_margin, t=20, b=20),
            showlegend=False, height=height,
            uniformtext_minsize=11, uniformtext_mode="hide"
        )
        return fig

    def _pick(col, label, key, help_=None):
        df_full = SS.get('df') 
        if df_full is None:
            return col.selectbox(label, ["—"], index=0, key=key, help=help_)
        
        v = col.selectbox(label, ["—"] + list(df_full.columns), index=0, key=key, help=help_)
        return None if v == "—" else v

    def _norm_period_value(p_text):
        s = str(p_text).lower() if p_text else "month"
        if s.startswith(("m","tháng")): return "Month"
        if s.startswith(("q","quý")):   return "Quarter"
        if s.startswith(("y","năm")):   return "Year"
        return "Month"

    def _norm_ser(s: pd.Series) -> pd.Series:
        return s.astype(str).str.strip().str.replace(r"\s+", " ", regex=True).str.lower()

    def _norm_list(vals):
        if not vals: return set()
        return set(pd.Series(list(vals)).astype(str).str.strip().str.replace(r"\s+", " ", regex=True).str.lower())

    # === (KHÔI PHỤC) Drill-down per chart (đúng như code gốc) ===
    def _chart_drilldown_mask(prefix: str,
                              dfin: pd.DataFrame,
                              tv: pd.Series, rule_code: str,
                              region_col: str | None,
                              channel_col: str | None,
                              prod_col: str | None,
                              cust_col: str | None,
                              time_col_present: bool = True):
        def _top_values_local(df_local, col, k=200):
            if not col or col not in dfin.columns: return []
            return dfin[col].astype(str).value_counts(dropna=False).head(k).index.tolist()

        with st.expander("🎯 Drill-down filter — Khoanh vùng dữ liệu (biểu đồ này)", expanded=False):
            ckR, ckC, ckP, ckU, ckT = st.columns([1,1,1,1,1])
            useR = ckR.checkbox("Region",  key=f"{prefix}_useR")
            useC = ckC.checkbox("Channel", key=f"{prefix}_useC")
            useP = ckP.checkbox("Product", key=f"{prefix}_useP")
            useU = ckU.checkbox("Customer", key=f"{prefix}_useU")
            useT = ckT.checkbox("Time",    key=f"{prefix}_useT") if time_col_present else False

            m1, m2 = st.columns([1.1, 2.2])
            selR = m1.multiselect("Region",  _top_values_local(dfin, region_col),  key=f"{prefix}_valR") if (useR and region_col and region_col in dfin.columns) else []
            selC = m1.multiselect("Channel", _top_values_local(dfin, channel_col), key=f"{prefix}_valC") if (useC and channel_col and channel_col in dfin.columns) else []
            selP = m2.multiselect("Product",  _top_values_local(dfin, prod_col),   key=f"{prefix}_valP") if (useP and prod_col and prod_col in dfin.columns) else []
            selU = m2.multiselect("Customer", _top_values_local(dfin, cust_col),   key=f"{prefix}_valU") if (useU and cust_col and cust_col in dfin.columns) else []

            if useT and time_col_present and tv is not None and not tv.isna().all():
                per_lbl = {"MS":"Month","QS":"Quarter","YS":"Year"}[rule_code]
                per_str = tv.dt.to_period({"MS":"M","QS":"Q","YS":"Y"}[rule_code]).astype(str)
                uniq_periods = sorted(pd.Series(per_str.loc[dfin.index]).dropna().unique().tolist())
                selT = m2.multiselect(f"Kỳ theo {per_lbl}", uniq_periods, key=f"{prefix}_valT")
            else:
                selT = []

        mask = pd.Series(True, index=dfin.index)
        if useR and region_col and selR and region_col in dfin.columns: mask &= dfin[region_col].astype(str).isin(selR)
        if useC and channel_col and selC and channel_col in dfin.columns: mask &= dfin[channel_col].astype(str).isin(selC)
        if useP and prod_col and selP and prod_col in dfin.columns:    mask &= dfin[prod_col].astype(str).isin(selP)
        if useU and cust_col and selU and cust_col in dfin.columns:    mask &= dfin[cust_col].astype(str).isin(selU)
        if useT and time_col_present and selT and tv is not None:
            per_now = tv.dt.to_period({"MS":"M","QS":"Q","YS":"Y"}[rule_code]).astype(str)
            mask &= per_now.loc[dfin.index].isin(set(selT))
        return mask

    def _sparse_line_labels(y_vals, fmt=lambda v: f"{v:.1f}%", min_dy_ratio=0.08, max_points=22):
        y = np.array([np.nan if v is None else v for v in y_vals], dtype=float)
        if len(y) == 0 or len(y) > max_points:
            return None
        vmin, vmax = np.nanmin(y), np.nanmax(y)
        rng = (vmax - vmin) if np.isfinite(vmax - vmin) and (vmax - vmin)!=0 else 1.0
        out, last = [], None
        for v in y:
            if np.isnan(v):
                out.append("")
                continue
            if (last is None) or (abs(v - last) >= min_dy_ratio * rng):
                out.append(fmt(v)); last = v
            else:
                out.append("")
        return out

    def _bar_text(values, fmt=lambda v: f"{v:,.0f}", max_labels=12):
        if len(values) <= max_labels:
            return [fmt(v) for v in values]
        return None
        
    @st.cache_data(ttl=900, show_spinner="Đang tính toán chiết khấu...")
    def get_discount_analysis(df_source, group_col, revenue_series, discount_series):
        if not group_col or group_col not in df_source.columns:
            return pd.DataFrame(columns=["Group", "Discount_Rate", "Total_Discount", "Total_Revenue"])
        
        g_disc = (pd.DataFrame({
            "Group": df_source[group_col].astype(str).fillna("(NA)"),
            "SalesB": revenue_series,
            "DiscB":  discount_series
        }).groupby("Group").sum(numeric_only=True))

        g_disc = g_disc[g_disc["SalesB"] > 0] 
        if g_disc.empty:
            return pd.DataFrame(columns=["Group", "Discount_Rate", "Total_Discount", "Total_Revenue"])
            
        g_disc["Discount_Rate"] = (-g_disc["DiscB"] / g_disc["SalesB"]) * 100.0
        g_disc = g_disc.sort_values("Discount_Rate", ascending=False)
        
        g_disc.columns = ["Total_Revenue", "Total_Discount", "Discount_Rate"]
        g_disc = g_disc.reset_index() 
        return g_disc[['Group', 'Discount_Rate', 'Total_Discount', 'Total_Revenue']]

    # ---- Data / guard
    df = SS.get("df") 
    if df is None or df.empty:
        st.info("Hãy nạp dữ liệu trước."); st.stop()

    # ====================== 0) Import Input Data — (ĐÃ CẬP NHẬT) ======================
    st.markdown("### ⚙️ Import Input Data — (Required)")
    with st.container(border=True):
        c1, c2, c3, c4, c5 = st.columns(5)
        time_col    = _pick(c1, "🕒 Time", "ov_time", help_="Datetime để resample Month/Quarter/Year.")
        cust_col    = _pick(c2, "👤 Customer", "ov_cust")
        prod_col    = _pick(c3, "📦 Product", "ov_prod")
        region_col  = _pick(c4, "🌍 Region", "ov_region")
        channel_col = _pick(c5, "🛒 Channel", "ov_channel")

        r1, r2, r3, r4 = st.columns(4) 
        rev_col    = _pick(r1, "💰 Revenue", "ov_rev", help_="Doanh thu cho biểu đồ/bảng.")
        weight_vol_col = _pick(r2, "⚖️ Weight (Amount)", "ov_weight_vol", 
                               help_="Dùng cho Avg Price, %Sales(A), và chart Revenue vs Weight.")
        
        map_a = _pick(r3, "🏷️ Mapping A — Transaction", "ov_map_a",
                      help_="Phân Sales (External) vs Transfer (Internal) — theo VOLUME/WEIGHT.")
        map_b = _pick(r4, "🏷️ Mapping B — Value Type", "ov_map_b",
                      help_="Phân Sales(B) vs Discount(B) — theo REVENUE.")
        if map_a and map_b and map_a == map_b:
            st.warning("Mapping A và Mapping B đang dùng **cùng cột**. Hãy chọn cột khác nhau.")

        uniq_a = sorted(df[map_a].astype(str).unique().tolist()) if map_a and map_a in df.columns else []
        uniq_b = sorted(df[map_b].astype(str).unique().tolist()) if map_b and map_b in df.columns else []
        
        with st.expander("Mapping chi tiết", expanded=False):
            a1, a2 = st.columns(2)
            mv_a_sales = a1.multiselect("Sales (External) — A", uniq_a, key="mv_a_sales")
            mv_a_trans = a2.multiselect("Transfer (Internal) — A", uniq_a, key="mv_a_transfer")
            b1, b2 = st.columns(2)
            mv_b_sales = b1.multiselect("Sales (B)", uniq_b, key="mv_b_sales")
            mv_b_disc  = b2.multiselect("Discount (B)", uniq_b, key="mv_b_disc")

    if not rev_col or rev_col not in df.columns:
        st.info("Cần chọn **Revenue** để xem Overview."); 
    else:

        # ====================== 1) Display ======================
        st.markdown("### 🧭 Display")
        d1, d2, d3 = st.columns([1,1,1])
        period_raw = d1.segmented_control("Period", ["Month","Quarter","Year"])
        compare    = d2.segmented_control("Compare", ["Prev","YoY"])
        period     = _norm_period_value(period_raw)
        rule       = RULE[period]

        if time_col and time_col in df.columns:
            all_years = sorted(pd.to_datetime(df[time_col], errors="coerce").dropna().dt.year.unique())
            year_scope = d3.selectbox("Year scope (KPI/Trend)", ["All"]+[str(y) for y in all_years], index=len(all_years))
        else:
            year_scope = "All"
            
        # ====================== 2) Lọc scope năm ======================
        t_all = _clean_time(df[time_col]) if time_col and time_col in df.columns else pd.Series(pd.NaT, index=df.index)
        mask_scope = (t_all.dt.year == int(year_scope)) if (time_col and year_scope!="All") else pd.Series(True, index=df.index)
        
        dfv = df.loc[mask_scope].copy() # dfv là DataFrame chỉ lọc theo năm
        tv  = t_all.loc[mask_scope] if time_col else pd.Series(pd.NaT, index=dfv.index)
        
        if dfv.empty:
            st.info("Không có dữ liệu trong phạm vi đã chọn."); st.stop()
        # =========================================================================

        # series cơ bản
        rev = pd.to_numeric(dfv[rev_col], errors="coerce").fillna(0.0)
        vol_wgt = pd.to_numeric(dfv[weight_vol_col], errors="coerce").fillna(0.0) if weight_vol_col and weight_vol_col in dfv.columns else pd.Series(0.0, index=dfv.index)

        # Mapping A
        if map_a and map_a in dfv.columns:
            A_norm = _norm_ser(dfv[map_a])
            m_salesA    = A_norm.isin(_norm_list(SS.get("mv_a_sales", [])))
            m_transferA = A_norm.isin(_norm_list(SS.get("mv_a_transfer", [])))
        else:
            m_salesA    = pd.Series(False, index=dfv.index)
            m_transferA = pd.Series(False, index=dfv.index)

        salesA_vol    = vol_wgt.where(m_salesA, 0.0)
        transferA_vol = vol_wgt.where(m_transferA, 0.0)
        baseA = float(salesA_vol.abs().sum() + transferA_vol.abs().sum())
        pct_salesA    = (float(salesA_vol.abs().sum())/baseA*100) if baseA>0 else np.nan
        pct_transferA = (float(transferA_vol.abs().sum())/baseA*100) if baseA>0 else np.nan

        # Mapping B
        if map_b and map_b in dfv.columns:
            B_norm = _norm_ser(dfv[map_b])
            is_salesB = B_norm.isin(_norm_list(SS.get("mv_b_sales", [])))
            is_discB  = B_norm.isin(_norm_list(SS.get("mv_b_disc",  [])))
            salesB_rev = rev.where(is_salesB, 0.0)
            discB_rev  = rev.where(is_discB,  0.0)
        else:
            salesB_rev = rev.copy() 
            discB_rev  = pd.Series(0.0, index=dfv.index) 

        # Discount% (Tính toán sớm để KPI dùng)
        disc_avg_month = np.nan; disc_year_pct = np.nan
        mon_sales_gt_0 = pd.DataFrame() 
        
        if time_col and not salesB_rev.empty and not discB_rev.empty:
            mon = (pd.DataFrame({"m": tv.dt.to_period("M").dt.start_time,
                                "SalesB": salesB_rev, "DiscB": discB_rev})
                .groupby("m").sum(numeric_only=True))
            mon_sales_gt_0 = mon[mon["SalesB"] > 0].copy() 
            
            if not mon_sales_gt_0.empty:
                mon_sales_gt_0["Discount%"] = (-mon_sales_gt_0["DiscB"] / mon_sales_gt_0["SalesB"]) * 100.0
                
                y_opts = sorted(mon_sales_gt_0.index.year.unique())
                if y_opts:
                    yr = int(year_scope) if (year_scope!="All" and int(year_scope) in y_opts) else int(y_opts[-1])
                    mon_y = mon_sales_gt_0[mon_sales_gt_0.index.year==yr]
                    
                    if not mon_y.empty:
                        disc_avg_month = float(mon_y["Discount%"].mean())
                        disc_year_pct  = float((-mon_y["DiscB"].sum() / mon_y["SalesB"].sum()) * 100.0)

        revenue_for_charts = salesB_rev if (map_b and map_b in dfv.columns and SS.get("mv_b_sales")) else rev

        # ====================== 3) KPI (ĐÃ CẬP NHẬT) ======================
        orders_total = len(dfv)
        prod_total   = (dfv.loc[revenue_for_charts>0, prod_col].nunique()
                        if (prod_col and prod_col in dfv.columns) else np.nan)
        revenue_total = float(revenue_for_charts.sum())

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Revenue (for charts)", f"{revenue_total:,.0f}")
        k2.metric("Total Transactions", f"{orders_total:,.0f}") 
        k3.metric("Total product", f"{prod_total:,.0f}" if not np.isnan(prod_total) else "—")
        k4.metric("%Sales (A) by Weight/Vol", f"{pct_salesA:.1f}%" if not np.isnan(pct_salesA) else "—") 

        k5, k6, k7, k8 = st.columns(4)
        k5.metric("%Transfer (A) by Weight/Vol", f"{pct_transferA:.1f}%" if not np.isnan(pct_transferA) else "—") 
        k6.metric("Discount% avg monthly (B)", f"{disc_avg_month:.1f}%" if not np.isnan(disc_avg_month) else "—")
        k7.metric("Discount% (YTD, B)", f"{disc_year_pct:.1f}%" if not np.isnan(disc_year_pct) else "—")
        k8.metric("Scope year", year_scope)

        # =============== 4) Trend — Revenue + %Δ ==================
        with st.expander("📊 1. Trend — Revenue & %Δ", expanded=True):
            tmask = _chart_drilldown_mask("ov_trend", dfv, tv, rule, region_col, channel_col, prod_col, cust_col, bool(time_col))
            
            with st.expander("🎨 Màu & nhãn — Trend", expanded=False):
                ccol1, ccol2, ccol3 = st.columns([1,1,1])
                color_bar_trend   = ccol1.color_picker("Màu cột (Revenue)", "#74b9ff", key="clr_tr_bar")
                color_line_trend  = ccol2.color_picker("Màu line (%Δ)",    "#1f77b4", key="clr_tr_line")
                color_text_common = ccol3.color_picker("Màu số liệu (labels)", "#cccccc", key="clr_tr_txt")
                show_all_line_lbl = st.checkbox("Hiện tất cả nhãn line", value=True, key="tr_show_all")

            if time_col:
                g_rev = _agg_by_period(tv.loc[tmask], revenue_for_charts.loc[tmask], rule)
                base  = g_rev.shift(1) if compare=="Prev" else g_rev.shift(YOY[rule])
                pct   = np.where(base!=0, (g_rev/base-1.0)*100.0, np.nan)

                bar_text = _bar_text(g_rev.values)
                line_text = [f"{v:.1f}%" if (v is not None and not np.isnan(v)) else "" for v in pct] if show_all_line_lbl else _sparse_line_labels(pct, fmt=lambda v: f"{v:.1f}%")

                y_pad = max(g_rev.max() * 0.15, 1.0) if not g_rev.empty else 1.0
                fig = go.Figure()
                fig.add_bar(x=g_rev.index, y=g_rev.values, name="Revenue",
                            marker_color=color_bar_trend,
                            text=bar_text, textposition="outside",
                            textfont=dict(color=color_text_common), cliponaxis=False)
                fig.add_scatter(x=g_rev.index, y=pct, yaxis="y2", mode="lines+markers+text", name="%Δ",
                                line=dict(color=color_line_trend),
                                text=line_text, textposition="top center",
                                textfont=dict(color=color_text_common))
                fig.update_layout(
                    xaxis_title=period,
                    yaxis=dict(title="Revenue", range=[0, float(g_rev.max()+y_pad)] if not g_rev.empty else [0,1], title_standoff=8),
                    yaxis2=dict(title="%Δ", overlaying="y", side="right", showgrid=False, title_standoff=14),
                    margin=dict(l=10,r=90,t=10,b=10), showlegend=True, height=430,
                    uniformtext_minsize=10, uniformtext_mode="hide"
                )
                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

                with st.expander("📄 Trend data (table)"):
                    show = pd.DataFrame({
                        period: g_rev.index, "Revenue": g_rev.values,
                        "Base": base.values, "%Δ": pct
                    })
                    # (SỬA LỖI ĐỊNH DẠNG)
                    show_styled = show.copy()
                    show_styled["Revenue"] = show_styled["Revenue"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    show_styled["Base"] = show_styled["Base"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    show_styled["%Δ"] = show_styled["%Δ"].map(lambda x: f"{x:.1f}%" if pd.notna(x) else "—")
                    st.dataframe(show_styled, use_container_width=True, hide_index=True)
            else:
                st.info("Cần chọn **Time** để xem Trend.")

        # =============== (MỚI) TÁCH PHÂN TÍCH CHIẾT KHẤU RA RIÊNG =================
        with st.expander("📉 2. Phân tích Chiết khấu (Discount Analysis)", expanded=False):
            if not time_col or not map_b or not SS.get("mv_b_sales"):
                st.info("Cần chọn **Time** và **Mapping B (Value Type)** (bao gồm Sales và Discount) để xem phân tích Chiết khấu.")
            else:
                dmask = _chart_drilldown_mask("ov_disc", dfv, tv, rule, region_col, channel_col, prod_col, cust_col, bool(time_col))
                
                dfv_disc = dfv.loc[dmask]
                tv_disc = tv.loc[dmask]
                salesB_rev_disc = salesB_rev.loc[dmask]
                discB_rev_disc = discB_rev.loc[dmask]

                st.markdown("#### Tỷ lệ chiết khấu hàng tháng")
                
                mon_disc = (pd.DataFrame({
                    "m": tv_disc.dt.to_period("M").dt.start_time,
                    "SalesB": salesB_rev_disc, 
                    "DiscB":  discB_rev_disc
                }).groupby("m").sum(numeric_only=True))
                mon_disc_gt_0 = mon_disc[mon_disc["SalesB"] > 0].copy() 
                
                if mon_disc_gt_0.empty:
                    st.info("Chưa đủ dữ liệu để tính Monthly Discount (sau khi drill-down).")
                else:
                    mon_disc_gt_0["Discount%"] = (-mon_disc_gt_0["DiscB"] / mon_disc_gt_0["SalesB"]) * 100.0
                    y_opts = sorted(mon_disc_gt_0.index.year.unique())
                    if not y_opts:
                        st.warning("Không tìm thấy năm hợp lệ.")
                    else:
                        default_year_idx = y_opts.index(int(year_scope)) if (year_scope!="All" and int(year_scope) in y_opts) else len(y_opts)-1
                        yr = st.selectbox("Year", y_opts, index=default_year_idx, key="trend_disc_year")

                        show = mon_disc_gt_0[mon_disc_gt_0.index.year == int(yr)].copy()
                        show.index = show.index.strftime("%b %Y")
                        
                        # (SỬA LỖI ĐỊNH DẠNG)
                        styled = show.copy()
                        styled["SalesB"]    = styled["SalesB"].map(lambda x: f"{x:,.0f}")
                        styled["DiscB"]     = styled["DiscB"].map(lambda x: f"{x:,.0f}")
                        styled["Discount%"] = styled["Discount%"].map(lambda x: f"{x:.1f}%")
                        st.dataframe(styled, use_container_width=True, height=340)
                        
                        st.markdown("#### Top Chiết khấu (trong năm đã chọn & đã drill-down)")
                        
                        dfv_year_disc = dfv_disc[tv_disc.dt.year == int(yr)] if time_col else dfv_disc
                        salesB_rev_year_disc = salesB_rev.loc[dfv_year_disc.index]
                        discB_rev_year_disc = discB_rev.loc[dfv_year_disc.index]
                        
                        tabD1, tabD2, tabD3 = st.tabs(["Theo Cửa hàng (Region)", "Theo Sản phẩm", "Theo Khách hàng"])
                        
                        # (SỬA LỖI ĐỊNH DẠNG)
                        def format_disc_table(df_in):
                            df_out = df_in.head(15).copy()
                            df_out["Discount_Rate"] = df_out["Discount_Rate"].map(lambda x: f"{x:.2f}%" if pd.notna(x) else "—")
                            df_out["Total_Discount"] = df_out["Total_Discount"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                            df_out["Total_Revenue"] = df_out["Total_Revenue"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                            return df_out
                        
                        with tabD1:
                            df_disc_region = get_discount_analysis(dfv_year_disc, region_col, salesB_rev_year_disc, discB_rev_year_disc)
                            st.dataframe(format_disc_table(df_disc_region), 
                                         use_container_width=True, hide_index=True)
                        with tabD2:
                            df_disc_prod = get_discount_analysis(dfv_year_disc, prod_col, salesB_rev_year_disc, discB_rev_year_disc)
                            st.dataframe(format_disc_table(df_disc_prod), 
                                         use_container_width=True, hide_index=True)
                        with tabD3:
                            df_disc_cust = get_discount_analysis(dfv_year_disc, cust_col, salesB_rev_year_disc, discB_rev_year_disc)
                            st.dataframe(format_disc_table(df_disc_cust), 
                                         use_container_width=True, hide_index=True)
            
        # ============ 4b) Sales Revenue vs Sales Weight ============
        with st.expander("💹 3. Sales Revenue vs Sales Weight", expanded=False):
            rw_mask = _chart_drilldown_mask("ov_rw", dfv, tv, rule, region_col, channel_col, prod_col, cust_col, bool(time_col))
            
            with st.expander("🎨 Màu & nhãn — Revenue vs Weight", expanded=False):
                c3, c4, c5 = st.columns([1,1,1])
                color_bar_rw   = c3.color_picker("Màu cột (Revenue)", "#74b9ff", key="clr_rw_bar")
                color_line_rw  = c4.color_picker("Màu line (Weight)", "#2ca02c", key="clr_rw_line")
                color_text_rw  = c5.color_picker("Màu số liệu (labels)", "#cccccc", key="clr_rw_txt")
                show_all_line_lbl_rw = st.checkbox("Hiện tất cả nhãn line", value=True, key="rw_show_all")

            if time_col and weight_vol_col and weight_vol_col in df.columns:
                g_rev2 = _agg_by_period(tv.loc[rw_mask], revenue_for_charts.loc[rw_mask], rule)
                g_wgt2 = _agg_by_period(tv.loc[rw_mask], vol_wgt.loc[rw_mask].where(vol_wgt.loc[rw_mask]>0, 0.0), rule)
                idx = g_rev2.index.union(g_wgt2.index)
                g_rev2 = g_rev2.reindex(idx, fill_value=0)
                g_wgt2 = g_wgt2.reindex(idx, fill_value=0)

                bar_text = _bar_text(g_rev2.values)
                line_text = [f"{v:,.0f}" if not pd.isna(v) else "" for v in g_wgt2.values] if show_all_line_lbl_rw else _sparse_line_labels(g_wgt2.values, fmt=lambda v: f"{v:,.0f}")

                y_pad = max(g_rev2.max() * 0.15, 1.0) if not g_rev2.empty else 1.0
                fig2 = go.Figure()
                fig2.add_bar(x=idx, y=g_rev2.values, name="Sales Revenue",
                            marker_color=color_bar_rw,
                            text=bar_text, textposition="outside",
                            textfont=dict(color=color_text_rw),
                            cliponaxis=False)
                fig2.add_scatter(x=idx, y=g_wgt2.values, yaxis="y2", mode="lines+markers+text", name="Sales Weight",
                                line=dict(color=color_line_rw),
                                text=line_text, textposition="top center",
                                textfont=dict(color=color_text_rw))
                fig2.update_layout(
                    xaxis_title=period,
                    yaxis=dict(title="Sales Revenue", range=[0, float(g_rev2.max()+y_pad)] if not g_rev2.empty else [0,1], title_standoff=8),
                    yaxis2=dict(title="Sales Weight", overlaying="y", side="right", showgrid=False, title_standoff=14),
                    margin=dict(l=10,r=90,t=10,b=10), showlegend=True, height=430,
                    uniformtext_minsize=10, uniformtext_mode="hide"
                )
                st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar": False})

                with st.expander("📄 Revenue vs Weight — monthly (table)"):
                    show = pd.DataFrame({period: idx, "Revenue": g_rev2.reindex(idx).values,
                                        "Weight": g_wgt2.reindex(idx).values})
                    # (SỬA LỖI ĐỊNH DẠNG)
                    show_styled = show.copy()
                    show_styled["Revenue"] = show_styled["Revenue"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    show_styled["Weight"] = show_styled["Weight"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    st.dataframe(show_styled, use_container_width=True, hide_index=True)
            else:
                st.info("Cần chọn **Time** và **Weight (Amount)** để xem biểu đồ này.")

        # ====================== 5) Top Contribution  |  Pie ======================
        with st.expander("🧱 4. Top Contribution  |  🥧 Pie", expanded=False):
            tc1, tc2, tc3 = st.columns([2,1,1])
            dim_col = tc1.selectbox("📊 Dimension (X)", ["—"] + list(dfv.columns), index=0, key="ov_dim_topc")
            topN    = tc2.slider("Top-N", 3, 50, 10, key="ov_topn_topc")
            as_share= tc3.checkbox("Chuẩn hóa % (share)", value=False, key="ov_share")

            pal_opts = { "Plotly": px.colors.qualitative.Plotly, "Bold": px.colors.qualitative.Bold, "Pastel": px.colors.qualitative.Pastel, "Set3": px.colors.qualitative.Set3 }
            with st.expander("🎨 Màu & nhãn — Top Contribution", expanded=False):
                cpl, cln = st.columns([1,1])
                pal_name = cpl.selectbox("Bảng màu Bar/Pie", list(pal_opts.keys()), index=0, key="ov_tc_palette")
                color_line_cum = cln.color_picker("Màu line (Cumulative %)", "#636EFA", key="clr_tc_line")
                color_text_tc  = st.color_picker("Màu số liệu (labels)", "#cccccc", key="clr_tc_txt")
                show_all_line_lbl_tc = st.checkbox("Hiện tất cả nhãn line", value=True, key="tc_show_all")

            if (not dim_col) or (dim_col=="—") or (dim_col not in dfv.columns):
                st.info("Chọn Dimension (X) để xem Top Contribution.")
            else:
                tc_mask = _chart_drilldown_mask("ov_tc", dfv, tv, rule, region_col, channel_col, prod_col, cust_col, bool(time_col))
                
                dim_vals = dfv.loc[tc_mask, dim_col].astype(str).fillna("(NA)")
                g = (pd.DataFrame({"d": dim_vals, "v": revenue_for_charts.loc[tc_mask]})
                    .groupby("d", dropna=False)["v"].sum().sort_values(ascending=False))

                total_sel = float(g.sum()) if len(g) else 0.0
                g_top = g.head(topN)
                cum   = (g_top.cumsum()/total_sel*100.0) if total_sel>0 else pd.Series(np.nan, index=g_top.index)
                yvals = (g_top/total_sel*100.0) if (as_share and total_sel>0) else g_top

                palette = pal_opts[pal_name]
                colors_for = {cat: palette[i % len(palette)] for i, cat in enumerate(g_top.index)}

                cL, cR = st.columns([0.7, 0.3])
                with cL:
                    bar_text = _bar_text(yvals.values, fmt=(lambda v: f"{v:.1f}%") if as_share else (lambda v: f"{v:,.0f}"))
                    line_text = [f"{v:.1f}%" if not pd.isna(v) else "" for v in cum.values] if show_all_line_lbl_tc else _sparse_line_labels(cum.values, fmt=lambda v: f"{v:.1f}%")
                    fig_t = go.Figure()
                    fig_t.add_bar(
                        x=g_top.index, y=yvals.values, name="Top-N",
                        marker_color=[colors_for[c] for c in g_top.index],
                        text=bar_text, textposition="outside",
                        textfont=dict(color=color_text_tc), cliponaxis=False
                    )
                    fig_t.add_scatter(
                        x=g_top.index, y=cum.values, yaxis="y2", mode="lines+markers+text", name="Cumulative %",
                        line=dict(color=color_line_cum),
                        text=line_text, textposition="top center",
                        textfont=dict(color=color_text_tc)
                    )
                    fig_t.update_layout(
                        xaxis_title=dim_col,
                        yaxis_title=("Share %" if as_share else "Revenue"),
                        yaxis2=dict(title="Cumulative %", overlaying="y", side="right", showgrid=False, title_standoff=14),
                        margin=dict(l=10,r=90,t=10,b=10), showlegend=True, height=460,
                        uniformtext_minsize=10, uniformtext_mode="hide"
                    )
                    st.plotly_chart(fig_t, use_container_width=True, config={"displayModeBar": False})

                with cR:
                    other_val = max(0.0, total_sel - float(g_top.sum()))
                    labels = list(g_top.index) + (["Other"] if other_val > 0 else [])
                    values = list(g_top.values) + ([other_val] if other_val > 0 else [])
                    colors = [colors_for[c] for c in g_top.index] + (["#BDBDBD"] if other_val > 0 else [])
                    fig_p = _pie_with_smart_labels(labels, values, colors, height=460)
                    st.plotly_chart(fig_p, use_container_width=True, config={"displayModeBar": False})

                with st.expander("📄 Top contribution (table)"):
                    tbl = (pd.DataFrame({"Label": g_top.index, "Value": g_top.values})
                        .assign(Share=lambda d: d["Value"]/d["Value"].sum()*100 if d["Value"].sum()!=0 else np.nan))
                    # (SỬA LỖI ĐỊNH DẠNG)
                    tbl_styled = tbl.copy()
                    tbl_styled["Value"] = tbl_styled["Value"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    tbl_styled["Share"] = tbl_styled["Share"].map(lambda x: f"{x:.1f}%" if pd.notna(x) else "—")
                    st.dataframe(tbl_styled, use_container_width=True, hide_index=True)

        # =============== 6) Avg Price vs Revenue =================
        with st.expander("💹 5. Avg Price vs Revenue", expanded=False):
            pr_mask = _chart_drilldown_mask("ov_avg", dfv, tv, rule, region_col, channel_col, prod_col, cust_col, bool(time_col))
            
            with st.expander("🎨 Màu & nhãn — Avg Price vs Revenue", expanded=False):
                c5, c6, c7 = st.columns([1,1,1])
                color_bar_avg  = c5.color_picker("Màu cột (Revenue)", "#74b9ff", key="clr_avg_bar")
                color_line_avg = c6.color_picker("Màu line (Avg Price)", "#e377c2", key="clr_avg_line")
                color_text_avg = c7.color_picker("Màu số liệu (labels)", "#cccccc", key="clr_avg_txt")
                show_all_line_lbl_avg = st.checkbox("Hiện tất cả nhãn line", value=True, key="avg_show_all")

            if time_col and weight_vol_col and weight_vol_col in df.columns:
                grpM = tv.loc[pr_mask].dt.to_period("M").dt.start_time
                rev_bar = pd.DataFrame({"m": grpM, "v": revenue_for_charts.loc[pr_mask]}).groupby("m")["v"].sum()
                mask_w = vol_wgt.loc[pr_mask] > 0
                num = pd.DataFrame({"m": grpM, "num": revenue_for_charts.loc[pr_mask].where(mask_w, 0.0)}).groupby("m")["num"].sum()
                den = pd.DataFrame({"m": grpM, "den": vol_wgt.loc[pr_mask].where(mask_w, 0.0)}).groupby("m")["den"].sum().replace(0, np.nan)
                avg_price = (num/den).reindex(rev_bar.index)

                bar_text  = _bar_text(rev_bar.values)
                line_text = [f"{v:,.0f}" if not pd.isna(v) else "" for v in avg_price.values] if show_all_line_lbl_avg else _sparse_line_labels(avg_price.values, fmt=lambda v: f"{v:,.0f}")

                y_pad = max(rev_bar.max() * 0.15, 1.0) if not rev_bar.empty else 1.0
                figp = go.Figure()
                figp.add_bar(x=rev_bar.index, y=rev_bar.values, name="Revenue",
                            marker_color=color_bar_avg,
                            text=bar_text, textposition="outside",
                            textfont=dict(color=color_text_avg),
                            cliponaxis=False)
                figp.add_scatter(x=rev_bar.index, y=avg_price.values, yaxis="y2", mode="lines+markers+text", name="Avg Price",
                                line=dict(color=color_line_avg),
                                text=line_text, textposition="top center",
                                textfont=dict(color=color_text_avg))
                figp.update_layout(
                    xaxis_title="Month",
                    yaxis=dict(title="Revenue", range=[0, float(rev_bar.max()+y_pad)] if not rev_bar.empty else [0,1], title_standoff=8),
                    yaxis2=dict(title="Avg Price", overlaying="y", side="right", showgrid=False, title_standoff=14),
                    margin=dict(l=10,r=90,t=10,b=10), showlegend=True, height=430,
                    uniformtext_minsize=10, uniformtext_mode="hide"
                )
                st.plotly_chart(figp, use_container_width=True, config={"displayModeBar": False})

                with st.expander("📄 Avg Price vs Revenue — monthly (table)"):
                    show = pd.DataFrame({
                        "Month": rev_bar.index, "Revenue": rev_bar.values,
                        "Avg Price": avg_price.values
                    })
                    # (SỬA LỖI ĐỊNH DẠNG)
                    show_styled = show.copy()
                    show_styled["Revenue"] = show_styled["Revenue"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    show_styled["Avg Price"] = show_styled["Avg Price"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                    st.dataframe(show_styled, use_container_width=True, hide_index=True)
            else:
                st.info("Cần chọn **Time** và **Weight (Amount)** để xem Avg Price vs Revenue.")

        # =============== 7) Distribution — Region × Channel (stacked) ===============
        with st.expander("🗺️ 6. Distribution — Region × Channel (stacked)", expanded=False):
            ds_mask = _chart_drilldown_mask("ov_dist", dfv, tv, rule, region_col, channel_col, prod_col, cust_col, bool(time_col))

            with st.expander("🎨 Màu — Distribution", expanded=False):
                pal_name2 = st.selectbox("Bảng màu (stacked)", ["Plotly","Bold","Pastel","Set3"], index=0, key="ov_dist_pal")
            pal2 = {"Plotly": px.colors.qualitative.Plotly, "Bold": px.colors.qualitative.Bold, "Pastel": px.colors.qualitative.Pastel, "Set3": px.colors.qualitative.Set3}[pal_name2]

            if region_col and channel_col and region_col in dfv.columns and channel_col in dfv.columns:
                ddf = dfv.loc[ds_mask].copy()
                srev= revenue_for_charts.loc[ds_mask]

                topn_ch = st.slider("Top-N Channel (stacked)", 3, 20, 6, key="ov_dist_topn")
                ch_sum = (pd.DataFrame({"ch": ddf[channel_col].astype(str), "v": srev})
                        .groupby("ch")["v"].sum().sort_values(ascending=False))
                keep = set(ch_sum.head(topn_ch).index)
                ch = ddf[channel_col].astype(str).where(ddf[channel_col].astype(str).isin(keep), other="Other")

                g = (pd.DataFrame({"Region": ddf[region_col].astype(str), "Channel": ch, "v": srev})
                    .groupby(["Region","Channel"])["v"].sum().reset_index())
                piv = g.pivot(index="Region", columns="Channel", values="v").fillna(0.0)
                color_map = {c: pal2[i % len(pal2)] for i, c in enumerate(piv.columns)}

                row_tot = piv.sum(axis=1).replace(0, np.nan)
                share   = piv.div(row_tot, axis=0) * 100.0
                piv     = piv.loc[row_tot.sort_values().index]; share = share.loc[piv.index]

                fig = go.Figure(); thr = 8.0
                for col in piv.columns:
                    vals = piv[col].values
                    pct  = share[col].values
                    text = [f"{v:.1f}%" if pd.notna(v) else "" for v in pct]
                    pos  = ["inside" if (isinstance(p, (int,float)) and p >= thr) else "outside" for p in pct]
                    fig.add_bar(x=piv.index, y=vals, name=str(col),
                                marker_color=color_map[str(col)], text=text, textposition=pos, cliponaxis=False)
                fig.update_layout(
                    barmode="stack", xaxis_title="Region", yaxis_title="Revenue",
                    margin=dict(l=10,r=90,t=10,b=10), showlegend=True, height=460,
                    uniformtext_minsize=10, uniformtext_mode="hide"
                )
                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

                with st.expander("📄 Region × Channel (pivot table)"):
                    # (SỬA LỖI ĐỊNH DẠNG)
                    st.dataframe(piv.style.format("{:,.0f}"), use_container_width=True)
            else:
                st.info("Cần chọn **Region** và **Channel**.")

        # ====================== 8) ✨ Biểu đồ tuỳ chỉnh (pivot-like) ======================
        with st.expander("✨ 7. Biểu đồ tuỳ chỉnh (X/Y/Z như pivot)", expanded=False):
            with st.container(border=True):
                c0, c1, c2, c3 = st.columns([1.2,1,1,1])
                x_col = c0.selectbox("X (Datetime/Categorical)", ["—"] + list(df.columns), index=0, key="pv_x")
                y_col = c1.selectbox("Y (Numeric)", ["—"] + list(df.select_dtypes(include=[np.number]).columns), index=0, key="pv_y")
                z_mode= c2.selectbox("Z (Line)", ["None","% share of Y","Secondary numeric"], index=0, key="pv_zmode")
                chart = c3.selectbox("Chart type", ["Bar","Line","Bar + Line"], index=0, key="pv_chart")

                pv_mask = _chart_drilldown_mask("ov_pv", dfv, tv, rule, region_col, channel_col, prod_col, cust_col, bool(time_col))

                with st.expander("🎨 Màu & nhãn — Custom chart", expanded=False):
                    c7, c8, c9 = st.columns([1,1,1])
                    color_pv_bar  = c7.color_picker("Màu cột (Y)", "#74b9ff", key="clr_pv_bar")
                    color_pv_line = c8.color_picker("Màu line (Z)", "#ff7f0e", key="clr_pv_line")
                    color_pv_txt  = c9.color_picker("Màu số liệu (labels)", "#cccccc", key="clr_pv_txt")
                    show_all_line_lbl_pv = st.checkbox("Hiện tất cả nhãn line", value=True, key="pv_show_all")

                agg = st.radio("Aggregation for Y", ["sum","mean","median","count"], horizontal=True, key="pv_agg")

                if z_mode == "Secondary numeric":
                    z_col = st.selectbox("Z (Numeric for line)", ["—"] + list(df.select_dtypes(include=[np.number]).columns), index=0, key="pv_zcol")
                    z_agg= st.radio("Aggregation for Z", ["sum","mean","median","count"], horizontal=True, key="pv_zagg")
                else:
                    z_col = None; z_agg = None

                if (not x_col) or x_col=="—" or (not y_col) or y_col=="—" or x_col not in df.columns or y_col not in df.columns:
                    st.info("Chọn X và Y để vẽ biểu đồ tuỳ chỉnh.")
                else:
                    base = dfv.loc[pv_mask].copy()

                    x_series = base[x_col]
                    if pd.api.types.is_datetime64_any_dtype(df[x_col]) or 'date' in str(x_col).lower() or 'time' in str(x_col).lower():
                        x_series = pd.to_datetime(x_series, errors="coerce").dt.to_period(P2[rule]).dt.start_time
                        x_title  = f"{x_col} ({period})"
                    else:
                        x_series = x_series.astype(str); x_title = f"{x_col} (category)"

                    y_ser = pd.to_numeric(base[y_col], errors="coerce")
                    agg_map = {"sum":"sum","mean":"mean","median":"median","count":"count"}
                    if agg == "count":
                        dfY = pd.DataFrame({"x": x_series, "y": 1}).groupby("x")["y"].count()
                    else:
                        dfY = pd.DataFrame({"x": x_series, "y": y_ser}).groupby("x")["y"].agg(agg_map[agg])

                    dfZ = None
                    if z_mode == "Secondary numeric" and z_col and z_col != "—" and z_col in df.columns:
                        z_ser = pd.to_numeric(base[z_col], errors="coerce")
                        if z_agg == "count":
                            dfZ = pd.DataFrame({"x": x_series, "z": 1}).groupby("x")["z"].count()
                        else:
                            dfZ = pd.DataFrame({"x": x_series, "z": z_ser}).groupby("x")["z"].agg(agg_map[z_agg])
                        dfZ = dfZ.reindex(dfY.index)
                    elif z_mode == "% share of Y":
                        total_y = float(dfY.sum()) if dfY.notna().any() else 0.0
                        dfZ = (dfY/total_y*100.0) if total_y>0 else dfY*0+np.nan

                    figc = go.Figure()
                    bar_text = _bar_text(dfY.values)
                    y_pad = max(dfY.max() * 0.15, 1.0) if not dfY.empty else 1.0
                    
                    if chart in ("Bar","Bar + Line"):
                        figc.add_bar(x=dfY.index, y=dfY.values, name=f"Y ({agg})",
                                    marker_color=color_pv_bar,
                                    text=bar_text, textposition="outside",
                                    textfont=dict(color=color_pv_txt),
                                    cliponaxis=False)
                    if chart in ("Line","Bar + Line") and dfZ is not None:
                        line_text = [f"{v:.1f}%" if z_mode=="% share of Y" else (f"{v:,.0f}" if not pd.isna(v) else "") for v in dfZ.values] if show_all_line_lbl_pv else _sparse_line_labels(dfZ.values, fmt=(lambda v: f"{v:.1f}%") if z_mode=="% share of Y" else (lambda v: f"{v:,.0f}"))
                        figc.add_scatter(x=dfY.index, y=dfZ.values, yaxis="y2",
                                        mode="lines+markers+text", name="Z (line)",
                                        line=dict(color=color_pv_line),
                                        text=line_text, textposition="top center",
                                        textfont=dict(color=color_pv_txt))
                    figc.update_layout(
                        xaxis_title=x_title,
                        yaxis=dict(title=f"Y = {y_col} [{agg}]", range=[0, float(dfY.max()+y_pad)] if not dfY.empty else [0,1], title_standoff=8),
                        yaxis2=dict(title=("Z = % share" if z_mode=="% share of Y" else f"Z = {z_col} [{z_agg}]"),
                                    overlaying="y", side="right", showgrid=False, title_standoff=14) if (chart!="Bar" and dfZ is not None) else None,
                        margin=dict(l=10,r=90,t=10,b=10), showlegend=True, height=460,
                        uniformtext_minsize=10, uniformtext_mode="hide"
                    )
                    st.plotly_chart(figc, use_container_width=True, config={"displayModeBar": False})

                    with st.expander("📄 Custom chart (table)"):
                        out_tbl = pd.DataFrame({"X": dfY.index, "Y": dfY.values})
                        
                        # (SỬA LỖI ĐỊNH DẠNG)
                        out_tbl_styled = out_tbl.copy()
                        out_tbl_styled["Y"] = out_tbl_styled["Y"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                        
                        if dfZ is not None:
                            out_tbl_styled["Z"] = out_tbl["Z"]
                            if z_mode == "% share of Y":
                                out_tbl_styled["Z"] = out_tbl_styled["Z"].map(lambda x: f"{x:.1f}%" if pd.notna(x) else "—")
                            elif z_agg == "count" or (z_col and z_col in df.columns and ("quantity" in z_col.lower() or "weight" in z_col.lower() or "amount" in z_col.lower())):
                                out_tbl_styled["Z"] = out_tbl_styled["Z"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
                            else:
                                out_tbl_styled["Z"] = out_tbl_styled["Z"].map(lambda x: f"{x:,.2f}" if pd.notna(x) else "—")
                                
                        st.dataframe(out_tbl_styled, use_container_width=True, hide_index=True)

# ============================== TAB 2 — PROFILING / DISTRIBUTION ==============================
with TAB2:
    import numpy as np
    import pandas as pd
    import plotly.graph_objects as go
    import streamlit as st

    # SciPy (nếu có) để kiểm định Normality / skew-kurtosis chuẩn hơn
    try:
        from scipy import stats
        _HAS_SCIPY = True
    except Exception:
        _HAS_SCIPY = False
    MAX_POINTS_PROFILE = 500_000
    st.subheader("📊 Profiling / Distribution")

    df = st.session_state.get("df")
    if df is None or df.empty:
        st.info("Hãy nạp dữ liệu trước.")
        st.stop()

    # ------------------------- Helpers -------------------------
    MAX_TIME_OPTIONS = {"M": 240, "Q": 80, "Y": 40}  # giới hạn số kỳ hiển thị để UI mượt

    def _fmt_safe(x, fmt=".3f", na="—"):
        """Format số an toàn; NaN/None/±inf → na."""
        try:
            xv = float(x)
            if not np.isfinite(xv):
                return na
            return format(xv, fmt)
        except Exception:
            return na

    def _clean_time(ts, min_year=1900, max_year=2100):
        t = pd.to_datetime(ts, errors="coerce")
        bad = t.notna() & ((t.dt.year < min_year) | (t.dt.year > max_year))
        return t.mask(bad)

    def _top_values(df_local, col, k=200):
        if not col or col not in df_local.columns:
            return []
        return df_local[col].astype(str).value_counts(dropna=False).head(k).index.tolist()

    # màu cố định cho các mốc
    MARK_COLORS = {
        "Min":    "#7f8c8d",
        "Q1":     "#8e44ad",
        "Median": "#e84393",
        "Mean":   "#f1c40f",
        "Q3":     "#27ae60",
        "Max":    "#2d3436",
    }

    def _add_vlines_with_legend(
        fig, marks, y_max, dash="dot", annotate=True, label_font_size=11
    ):
        """
        Vẽ vline cho các mốc; label hiển thị ở legend bên phải.
        Đồng thời gắn nhãn trên biểu đồ (so le theo chiều cao để tránh chồng nhau).
        """
        n = len(marks)
        levels = np.linspace(0.92, 0.72, num=n) if n > 1 else [0.90]
        for (lab, xv), frac in zip(marks.items(), levels):
            if xv is None:
                continue
            try:
                xfloat = float(xv)
                if not np.isfinite(xfloat):
                    continue
            except Exception:
                continue

            col = MARK_COLORS.get(lab, "#888")

            # đường dọc + legend
            fig.add_scatter(
                x=[xfloat, xfloat],
                y=[0.0, float(y_max)],
                mode="lines",
                name=str(lab),
                line=dict(color=col, dash=dash, width=1.5),
                hovertemplate=f"{lab}: %{{x:,.4g}}<extra></extra>",
                showlegend=True,
            )
            # nhãn ngay trên chart (so le để không chồng)
            if annotate:
                fig.add_annotation(
                    x=xfloat, y=float(y_max) * float(frac),
                    xref="x", yref="y",
                    text=str(lab),
                    showarrow=False,
                    font=dict(size=label_font_size, color=col),
                    bordercolor="rgba(0,0,0,0)",
                    bgcolor="rgba(0,0,0,0)",
                    xanchor="center", yanchor="bottom",
                    align="center"
                )

    # ---------- Drill-down đúng UI ----------
    def _render_filter_inline(df_in, key_prefix="prof"):
        """Hàm lọc cục bộ, cho phép chọn cột tự do"""
        with st.expander("🔎 Drill-down Filter (Bộ lọc dữ liệu)", expanded=False):
            st.caption("Chọn cột và giá trị để khoanh vùng dữ liệu trước khi phân tích.")
            
            all_cols = ["—"] + list(df_in.columns)
            mask = pd.Series(True, index=df_in.index)

            # Hàng Checkbox
            c1, c2, c3, c4, c5 = st.columns(5)
            use_1 = c1.checkbox("Filter 1 (Region)", key=f"{key_prefix}_chk_1")
            use_2 = c2.checkbox("Filter 2 (Channel)", key=f"{key_prefix}_chk_2")
            use_3 = c3.checkbox("Filter 3 (Product)", key=f"{key_prefix}_chk_3")
            use_4 = c4.checkbox("Filter 4 (Customer)", key=f"{key_prefix}_chk_4")
            use_t = c5.checkbox("Time Filter", key=f"{key_prefix}_chk_t", value=True)

            # Layout chọn cột
            r1, r2 = st.columns([1.5, 2.5])
            
            def _render_sel(label, use_flag, keyword, suffix):
                col_name = None
                if use_flag:
                    with r1:
                        def_idx = 0
                        if keyword: # Tìm cột khớp từ khóa
                            for i, c in enumerate(all_cols):
                                if keyword.lower() in str(c).lower(): def_idx = i; break
                        col_name = st.selectbox(f"Chọn Cột ({label})", all_cols, index=def_idx, key=f"{key_prefix}_col_{suffix}")
                    
                    if col_name and col_name != "—":
                        with r2:
                            # Lấy Top 200 giá trị để hiển thị
                            top_vals = df_in[col_name].astype(str).value_counts().head(200).index.tolist()
                            vals = st.multiselect(f"Giá trị ({col_name})", top_vals, key=f"{key_prefix}_val_{suffix}")
                            return col_name, vals
                return None, []

            # Render 4 filter
            c1_n, v1 = _render_sel("Vị trí", use_1, "region", "1")
            if c1_n and v1: mask &= df_in[c1_n].astype(str).isin(v1)
            
            c2_n, v2 = _render_sel("Kênh", use_2, "channel", "2")
            if c2_n and v2: mask &= df_in[c2_n].astype(str).isin(v2)
            
            c3_n, v3 = _render_sel("Sản phẩm", use_3, "prod", "3")
            if c3_n and v3: mask &= df_in[c3_n].astype(str).isin(v3)
            
            c4_n, v4 = _render_sel("Khách hàng", use_4, "cust", "4")
            if c4_n and v4: mask &= df_in[c4_n].astype(str).isin(v4)

            # Time Filter
            time_col = None
            if use_t:
                with r1:
                    dt_cands = [c for c in df_in.columns if 'date' in str(c).lower() or 'time' in str(c).lower()]
                    dt_opts = ["—"] + dt_cands + [c for c in df_in.columns if c not in dt_cands]
                    time_col = st.selectbox("Cột Thời gian", dt_opts, key=f"{key_prefix}_col_time")
                
                if time_col and time_col != "—":
                    with r2:
                        try:
                            ts = _clean_time(df_in[time_col])
                            # Mặc định lọc theo Tháng
                            periods = sorted(ts.dt.to_period("M").astype(str).dropna().unique())
                            def_sel = periods[-3:] if len(periods) > 3 else periods
                            sel_t = st.multiselect(f"Chọn khoảng thời gian (Tháng)", periods, default=def_sel, key=f"{key_prefix}_val_time")
                            
                            if sel_t: mask &= ts.dt.to_period("M").astype(str).isin(sel_t)
                        except: st.warning("Lỗi định dạng thời gian.")

            n_remain = mask.sum()
            st.caption(f"⚡ Dữ liệu sau lọc: **{n_remain:,}** / {len(df_in):,} dòng.")
            return df_in.loc[mask]

    # --- ÁP DỤNG BỘ LỌC ---
    dfx = _render_filter_inline(df, "prof")

    if dfx.empty:
        st.warning("Không còn dữ liệu sau khi khoanh vùng. Vui lòng nới lỏng bộ lọc.")
        st.stop()

    # ---------- chọn biến numeric ----------
    NUMS = dfx.select_dtypes(include=[np.number]).columns.tolist()
    st.markdown("### 🧮 Chọn biến numeric")
    ncol = st.selectbox("Metric (numeric)", NUMS or ["—"], key="pr_num_sel")
    if (not ncol) or (ncol not in dfx.columns):
        st.info("Chưa chọn biến numeric hợp lệ.")
        st.stop()

    # ---------- làm sạch (log10/>0, bỏ =0, bỏ <0) ----------
    st.markdown("### 🧹 Làm sạch & tuỳ chọn")
    o1, o2, o3, o4 = st.columns([1, 1, 1, 1])
    use_log = o1.checkbox("log10 (chỉ >0)", value=False, key="pr_log")
    drop_eq0 = o2.checkbox("Bỏ = 0", value=False, key="pr_eq0")
    drop_lt0 = o3.checkbox("Bỏ < 0", value=False, key="pr_lt0")
    show_points_ecdf = o4.checkbox("ECDF points", value=False, key="pr_ecdf_pts")

    s_raw = pd.to_numeric(dfx[ncol], errors="coerce").replace([np.inf, -np.inf], np.nan)
    if drop_lt0:
        s_raw = s_raw[s_raw >= 0]
    if drop_eq0:
        s_raw = s_raw[s_raw != 0]
    if use_log:
        s = s_raw[s_raw > 0].copy()
        s = np.log10(s)
        x_title = f"log10({ncol})"
        log_note = " (log10)"
    else:
        s = s_raw.copy()
        x_title = ncol
        log_note = ""

    s = s.dropna()
    if s.empty:
        st.warning("Dữ liệu rỗng sau khi áp điều kiện. Hãy nới bộ lọc.")
        st.stop()
    if len(s) > MAX_POINTS_PROFILE:
        s_sampled_charts = s.sample(MAX_POINTS_PROFILE, random_state=42)
    else:
        s_sampled_charts = s

    MAX_STATS_SAMPLE = 500_000
    if len(s) > MAX_STATS_SAMPLE:
        s_stats_sample = s.sample(MAX_STATS_SAMPLE, random_state=42)
    else:
        s_stats_sample = s

    # ---------- Metric tổng hợp (2 cột; giải thích đưa xuống phần Nhận định) ----------
    desc = s.describe(percentiles=[.05, .25, .5, .75, .95]).to_dict()
    mean_v, median_v = float(s.mean()), float(s.median())
    try:
        mode_v = float(pd.Series(s).mode(dropna=True).iloc[0])
    except Exception:
        mode_v = np.nan
    std_v = float(s.std(ddof=1)) if len(s) > 1 else np.nan
    iqr_v = float(s.quantile(.75) - s.quantile(.25))
    cv_v = float(std_v / mean_v * 100) if (mean_v != 0 and np.isfinite(mean_v) and np.isfinite(std_v)) else np.nan
    if _HAS_SCIPY and len(s) > 2:
        skew_v = float(stats.skew(s_stats_sample))
    else:
        skew_v = float(pd.Series(s).skew()) if len(s) > 2 else np.nan
    if _HAS_SCIPY and len(s) > 3:
        kurt_v = float(stats.kurtosis(s_stats_sample, fisher=True))
    else:
        kurt_v = float(pd.Series(s).kurt()) if len(s) > 3 else np.nan
    if _HAS_SCIPY and len(s) > 7:
        try:
            p_norm = float(stats.normaltest(s_stats_sample)[1])  # D’Agostino K^2
        except Exception:
            p_norm = np.nan
    else:
        p_norm = np.nan

    miss = int(dfx[ncol].isna().sum())
    zero_cnt = int((pd.to_numeric(dfx[ncol], errors="coerce") == 0).sum())
    q1, q3 = float(s.quantile(.25)), float(s.quantile(.75))
    lf, uf = q1 - 1.5 * (q3 - q1), q3 + 1.5 * (q3 - q1)
    out_mask = (s < lf) | (s > uf)
    out_cnt = int(out_mask.sum())
    out_pct = (out_cnt / len(s) * 100.0) if len(s) else 0.0
    range_val = float(s.max() - s.min())

    skew_dir = (
        "lệch phải (đuôi phải)" if (np.isfinite(skew_v) and skew_v > 0.5)
        else ("lệch trái (đuôi trái)" if (np.isfinite(skew_v) and skew_v < -0.5) else "gần đối xứng")
    )
    tail_txt = (
        "đuôi dày hơn chuẩn (leptokurtic)" if (np.isfinite(kurt_v) and kurt_v > 0)
        else ("đuôi mỏng (platykurtic)" if (np.isfinite(kurt_v) and kurt_v < 0) else "gần chuẩn (mesokurtic)")
    )
    if p_norm == p_norm:
        normal_txt = "không bác bỏ giả thuyết **chuẩn**" if p_norm >= 0.05 else "bác bỏ giả thuyết **chuẩn**"
    else:
        normal_txt = "không kiểm định do n nhỏ / thiếu SciPy"
    c_tend = "Mean > Median" if mean_v > median_v else ("Mean < Median" if mean_v < median_v else "Mean ≈ Median")
    spread_g = "phân tán rộng" if (cv_v == cv_v and cv_v > 50) else ("trung bình" if (cv_v == cv_v and cv_v > 20) else "khá chặt")

    st.markdown("### 📋 Metric tổng hợp (Shape • Spread • Central tendency)")
    metric_tbl = pd.DataFrame(
        [
            ("Count", f"{len(s):,}", "Số mẫu sau làm sạch"),
            ("Missing", f"{miss:,}", "Giá trị thiếu (trước log/loại)"),
            ("Zero (=0)", f"{zero_cnt:,}", "Số giá trị bằng 0 (trước log)"),
            ("Min", _fmt_safe(desc.get("min")), "Nhỏ nhất"),
            ("P5", _fmt_safe(desc.get("5%")), "5th percentile"),
            ("Q1", _fmt_safe(q1), "25th percentile"),
            ("Median", _fmt_safe(median_v), "Trung vị (50%)"),
            ("Mean", _fmt_safe(mean_v), "Trung bình"),
            ("Mode", _fmt_safe(mode_v), "Giá trị xuất hiện nhiều nhất"),
            ("Q3", _fmt_safe(q3), "75th percentile"),
            ("P95", _fmt_safe(desc.get("95%")), "95th percentile"),
            ("Max", _fmt_safe(desc.get("max")), "Lớn nhất"),
            ("Std (σ)", _fmt_safe(std_v), "Độ lệch chuẩn"),
            ("IQR", _fmt_safe(iqr_v), "Q3 − Q1 (phần giữa)"),
            ("CV (%)", _fmt_safe(cv_v, ".2f"), "Độ biến thiên tương đối"),
            ("Skewness", _fmt_safe(skew_v), "Độ lệch trái/phải"),
            ("Kurtosis (excess)", _fmt_safe(kurt_v), "Độ nhọn so với chuẩn (0 = chuẩn)"),
            ("Range", _fmt_safe(range_val), "Khoảng trải rộng (max − min)"),
            ("Lower fence", _fmt_safe(lf), "Q1 − 1.5×IQR (mốc outlier)"),
            ("Upper fence", _fmt_safe(uf), "Q3 + 1.5×IQR (mốc outlier)"),
            ("Outliers (count, %)", f"{out_cnt:,} ({_fmt_safe(out_pct, '.2f')}%)", "Số lượng/ tỷ lệ điểm vượt fence"),
            ("Normality p-value", _fmt_safe(p_norm), "p≥0.05 → dữ liệu có thể coi là gần chuẩn"),
        ],
        columns=["Metric", "Value", "Giải thích"],
    )
    st.dataframe(
        metric_tbl, use_container_width=True, hide_index=True,
        height=min(520, 34 * (len(metric_tbl) + 1)),
    )

    # ---------- màu & bins cho phần biểu đồ chính ----------
    with st.expander("🎨 Tùy biến hiển thị (màu/bins)", expanded=False):
        c1, c2, c3, c4 = st.columns([1, 1, 1, 1])
        clr_hist = c1.color_picker("Histogram (bars)", "#74b9ff")
        clr_bell = c2.color_picker("Bell (Normal)", "#e67e22")
        clr_ecdf = c3.color_picker("ECDF", "#1abc9c")
        clr_box  = c4.color_picker("Box/Violin", "#a29bfe")
        bins = st.slider("Số bins (Histogram)", 20, 120, 50, 2, key="pr_bins")

    st.markdown("### 📈 Phân tích Phân phối & Diễn giải Dữ liệu")
    
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
        from scipy import stats

        # --- 1. CHUẨN BỊ DỮ LIỆU & TÍNH TOÁN ---
        # Lấy mẫu nếu dữ liệu quá lớn (>500k)
        if len(s) > 500_000:
            s_chart = s.sample(500_000, random_state=42)
            st.caption("⚠️ Dữ liệu > 500k dòng, biểu đồ & tính toán dùng mẫu 500k để tối ưu.")
        else:
            s_chart = s

        # Các chỉ số xu hướng trung tâm
        mu, sigma = float(s.mean()), float(s.std())
        min_val, max_val = float(s.min()), float(s.max())
        median_val = float(s.median())
        try:
            mode_val = float(s.mode().iloc[0])
        except:
            mode_val = np.nan
            
        # Các chỉ số hình dáng & phân tán
        skewness = float(s.skew())
        kurtosis = float(s.kurt())
        cv_pct = (sigma / mu * 100) if mu != 0 else 0
        range_val = max_val - min_val

        # Outliers (IQR method)
        Q1, Q3 = s.quantile(0.25), s.quantile(0.75)
        IQR = Q3 - Q1
        lower_fence, upper_fence = Q1 - 1.5*IQR, Q3 + 1.5*IQR
        n_outliers = ((s < lower_fence) | (s > upper_fence)).sum()
        pct_outliers = n_outliers / len(s) * 100
        
        # Normality Test (Chỉ chạy nếu đủ mẫu và < 5000 để chính xác, hoặc dùng s_stats_sample đã có từ trước)
        # Giả sử dùng p_norm đã tính ở phần trên (nếu có), hoặc tính nhanh lại
        try:
            if len(s) >= 8:
                stat_norm, p_norm = stats.normaltest(s_chart if len(s_chart)<5000 else s_chart.sample(5000))
            else:
                p_norm = np.nan
        except:
            p_norm = np.nan

        # --- 2. VẼ BIỂU ĐỒ SUBPLOTS ---
        fig = make_subplots(
            rows=1, cols=2, 
            column_widths=[0.7, 0.3], 
            subplot_titles=("Phân phối tần suất (Histogram)", "Mật độ & Outliers (Violin)"),
            horizontal_spacing=0.12 
        )

        # Cột 1: Histogram + Bell Curve + Lines
        n_bins = SS.get('pr_bins', 50)
        counts, bin_edges = np.histogram(s, bins=n_bins)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
        y_max_hist = max(counts)

        fig.add_trace(
            go.Bar(x=bin_centers, y=counts, name="Thực tế", marker_color=clr_hist, opacity=0.6, showlegend=True),
            row=1, col=1
        )

        if sigma > 0:
            x_bell = np.linspace(min_val, max_val, 200)
            pdf = stats.norm.pdf(x_bell, mu, sigma)
            y_bell = pdf * len(s) * (max_val - min_val) / n_bins
            fig.add_trace(
                go.Scatter(x=x_bell, y=y_bell, mode='lines', name='Phân phối chuẩn (Lý thuyết)', 
                           line=dict(color='#57606f', width=2, dash='solid')),
                row=1, col=1
            )
            y_max_hist = max(y_max_hist, max(y_bell))

        # Các đường chỉ báo (Mean, Median, Mode)
        y_range = [0, y_max_hist * 1.1]
        fig.add_trace(go.Scatter(x=[mu, mu], y=y_range, mode='lines', name=f'Mean: {mu:,.2f}',
                       line=dict(color='#e74c3c', width=2.5, dash='dash')), row=1, col=1)
        
        if abs(median_val - mu) > sigma*0.01: # Chỉ vẽ Median nếu lệch Mean một chút
             fig.add_trace(go.Scatter(x=[median_val, median_val], y=y_range, mode='lines', name=f'Median: {median_val:,.2f}',
                           line=dict(color='#2ecc71', width=2.5, dash='dot')), row=1, col=1)

        if not np.isnan(mode_val) and abs(mode_val - mu) > sigma*0.1:
             fig.add_trace(go.Scatter(x=[mode_val, mode_val], y=y_range, mode='lines', name=f'Mode: {mode_val:,.2f}',
                           line=dict(color='#9b59b6', width=2, dash='dashdot')), row=1, col=1)

        # Cột 2: Violin + Box
        fig.add_trace(
            go.Violin(y=s_chart, name=x_title, box_visible=True, meanline_visible=True, points=False,
                line_color=clr_box, fillcolor=clr_box, opacity=0.6, showlegend=False),
            row=1, col=2
        )

        # Layout
        fig.update_layout(
            height=460, hovermode="x unified",
            margin=dict(l=10, r=10, t=40, b=20),
            legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02, bgcolor="rgba(255,255,255,0.8)")
        )
        st.plotly_chart(fig, use_container_width=True)

        # --- 3. LOGIC DIỄN GIẢI TỰ ĐỘNG (MERGED) ---
        
        # A. Đánh giá Hình dáng (Shape) & Chuẩn hóa
        skew_lbl = "đối xứng (Symmetrical)"
        if skewness > 0.5: skew_lbl = "lệch phải (Right Skewed) - Đuôi kéo về phía giá trị lớn"
        elif skewness < -0.5: skew_lbl = "lệch trái (Left Skewed) - Đuôi kéo về phía giá trị nhỏ"
        
        kurt_lbl = "bình thường"
        if kurtosis > 1: kurt_lbl = "nhọn (Leptokurtic) - Dữ liệu tập trung quanh tâm, đuôi dày"
        elif kurtosis < -1: kurt_lbl = "bẹt (Platykurtic) - Dữ liệu phân tán dàn trải"

        norm_status = "Dữ liệu KHÔNG tuân theo phân phối chuẩn (p < 0.05)." if (p_norm < 0.05) else "Chưa đủ bằng chứng bác bỏ phân phối chuẩn (p >= 0.05)."

        # B. Đánh giá Xu hướng trung tâm
        trend_insight = ""
        if abs(mu - median_val) / (abs(mu)+1) < 0.05:
            trend_insight = "Mean ≈ Median: Dữ liệu khá cân bằng, có thể dùng Mean làm đại diện."
        elif mu > median_val:
            trend_insight = "Mean > Median: Giá trị trung bình bị kéo lên bởi các giá trị lớn đột biến."
        else:
            trend_insight = "Mean < Median: Giá trị trung bình bị kéo xuống bởi các giá trị nhỏ đột biến."

        # C. Đánh giá Độ phân tán (Spread)
        spread_eval = "Thấp (Ổn định)"
        if cv_pct > 50: spread_eval = "Rất cao (Biến động mạnh)"
        elif cv_pct > 20: spread_eval = "Trung bình"
        
        # D. Khuyến nghị Audit
        audit_action = "Dữ liệu ổn định, có thể dùng các phương pháp kiểm toán thông thường (Analytical Review)."
        if n_outliers > 0 and pct_outliers > 5:
            audit_action = "⚠️ Rủi ro cao: Tỷ lệ ngoại lai lớn (>5%). Cần trích mẫu kiểm tra các giao dịch vượt ngưỡng (Upper Fence) để phát hiện gian lận/sai sót."
        elif cv_pct > 100:
            audit_action = "⚠️ Rủi ro biến động: Dữ liệu dao động quá mạnh. Mean không còn ý nghĩa đại diện. Cần phân nhóm (Stratification) trước khi phân tích."
        elif skewness > 1:
            audit_action = "ℹ️ Lưu ý: Dữ liệu lệch phải. Nên dùng Median thay vì Mean để đánh giá xu hướng chung."

        # --- 4. HIỂN THỊ BOX DIỄN GIẢI (ST.INFO) ---
        st.info(f"""
        **🧠 Nhận định & Diễn giải Dữ liệu:**
        
        **1. Hình dáng Phân phối (Shape & Normality):**
        - Dữ liệu có xu hướng **{skew_lbl}**.
        - Độ nhọn: **{kurt_lbl}** (Kurtosis={kurtosis:.2f}).
        - Kiểm định chuẩn: {norm_status}
        
        **2. Xu hướng Trung tâm (Central Tendency):**
        - **Mean** ({mu:,.2f}) vs **Median** ({median_val:,.2f}).
        - 👉 **Nhận xét:** {trend_insight}
        
        **3. Độ Phân tán & Biến động (Spread):**
        - Độ biến động (CV) là **{cv_pct:.1f}%** ⇒ Mức độ phân tán **{spread_eval}**.
        - Khoảng trải rộng (Range) từ {min_val:,.2f} đến {max_val:,.2f}.
        
        **4. Ngoại lai (Outliers - Tukey's Fence):**
        - Phát hiện **{n_outliers:,}** điểm ngoại lai (chiếm **{pct_outliers:.2f}%**).
        - Các điểm này nằm ngoài khoảng: [{lower_fence:,.2f} ; {upper_fence:,.2f}].
        
        💡 **Khuyến nghị Audit:** {audit_action}
        """)

    except Exception as e:
        st.error(f"Lỗi hiển thị biểu đồ: {e}")

# ============================== TAB 3 — CORRELATION (BUSINESS-ORIENTED) ==============================
with TAB3:
    import numpy as np
    import pandas as pd
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    import streamlit as st

    # ===== Optional SciPy (p-value, Spearman nhanh, quantile Z) =====
    try:
        from scipy import stats
        _HAS_SCIPY = True
        _Z975 = stats.norm.ppf(0.975)
    except Exception:
        _HAS_SCIPY = False
        _Z975 = 1.96

    # ============================ Helpers ============================
    MAX_TIME_OPTIONS = {"M": 240, "Q": 80, "Y": 40}  # giới hạn #kỳ cho UI mượt

    def _fmt(x, fmt=".3f", na="—"):
        try:
            xv = float(x)
            if not np.isfinite(xv):
                return na
            return format(xv, fmt)
        except Exception:
            return na

    def _clean_time(ts, min_year=1900, max_year=2100):
        t = pd.to_datetime(ts, errors="coerce")
        bad = t.notna() & ((t.dt.year < min_year) | (t.dt.year > max_year))
        return t.mask(bad)

    def _top_values(df_local, col, k=200):
        if not col or col not in df_local.columns:
            return []
        return df_local[col].astype(str).value_counts(dropna=False).head(k).index.tolist()

    def strength_label(r_abs: float) -> str:
        if r_abs < 0.1: return "rất yếu"
        if r_abs < 0.3: return "yếu"
        if r_abs < 0.5: return "trung bình"
        if r_abs < 0.7: return "mạnh"
        if r_abs < 0.9: return "rất mạnh"
        return "gần hoàn hảo"

    def fisher_ci(r: float, n: int):
        """CI 95% cho Pearson r (Fisher z)."""
        try:
            if not np.isfinite(r) or n is None or n <= 3 or abs(r) >= 0.9999:
                return (np.nan, np.nan)
            z = np.arctanh(r)
            se = 1.0 / np.sqrt(n - 3)
            zlo, zhi = z - _Z975*se, z + _Z975*se
            return (np.tanh(zlo), np.tanh(zhi))
        except Exception:
            return (np.nan, np.nan)

    def corr_one(x: pd.Series, y: pd.Series, method="pearson"):
        """Trả về r, p, n (fallback nếu không có SciPy)."""
        s = pd.concat([x, y], axis=1).dropna()
        s.columns = ["x", "y"]
        n = len(s)
        if n < 3:
            return np.nan, np.nan, n
        if _HAS_SCIPY:
            if method.lower() == "pearson":
                r, p = stats.pearsonr(s["x"], s["y"])
            elif method.lower() == "spearman":
                r, p = stats.spearmanr(s["x"], s["y"])
            else:
                r, p = stats.kendalltau(s["x"], s["y"])
        else:
            r = np.corrcoef(s["x"], s["y"])[0, 1]
            p = np.nan
        return r, p, n

    def prepare_xy(xs: pd.Series, ys: pd.Series, drop_lt0, drop_eq0, use_log):
        s = pd.concat([xs, ys], axis=1).rename(columns={xs.name: "x", ys.name: "y"})
        s = s.replace([np.inf, -np.inf], np.nan).dropna()
        if drop_lt0:
            s = s[(s["x"] >= 0) & (s["y"] >= 0)]
        if drop_eq0:
            s = s[(s["x"] != 0) & (s["y"] != 0)]
        if use_log:
            s = s[(s["x"] > 0) & (s["y"] > 0)]
            s["x"] = np.log10(s["x"])
            s["y"] = np.log10(s["y"])
        return s["x"], s["y"]

    def guard(ok: bool, msg: str) -> bool:
        if ok: return True
        st.info(msg); st.write("---")
        return False

    # ============================ Data ============================
    df = st.session_state.get("df")
    if df is None or df.empty:
        st.info("Hãy nạp dữ liệu trước.")
        st.stop()

    st.subheader("🔗 Correlation")

    # ====================== 1. Drill-down filter (INLINE & FLEXIBLE) ======================
    def _render_filter_inline(df_in, key_prefix="corr"):
        """Hàm lọc cục bộ, cho phép chọn cột tự do"""
        with st.expander("🔎 Drill-down Filter (Bộ lọc dữ liệu)", expanded=False):
            st.caption("Chọn cột và giá trị để khoanh vùng dữ liệu trước khi phân tích.")
            
            all_cols = ["—"] + list(df_in.columns)
            mask = pd.Series(True, index=df_in.index)

            # Hàng Checkbox
            c1, c2, c3, c4, c5 = st.columns(5)
            use_1 = c1.checkbox("Filter 1 (Region)", key=f"{key_prefix}_chk_1")
            use_2 = c2.checkbox("Filter 2 (Channel)", key=f"{key_prefix}_chk_2")
            use_3 = c3.checkbox("Filter 3 (Product)", key=f"{key_prefix}_chk_3")
            use_4 = c4.checkbox("Filter 4 (Customer)", key=f"{key_prefix}_chk_4")
            use_t = c5.checkbox("Time Filter", key=f"{key_prefix}_chk_t", value=True)

            # Layout chọn cột
            r1, r2 = st.columns([1.5, 2.5])
            
            def _render_sel(label, use_flag, keyword, suffix):
                col_name = None
                if use_flag:
                    with r1:
                        def_idx = 0
                        if keyword: # Tìm cột khớp từ khóa
                            for i, c in enumerate(all_cols):
                                if keyword.lower() in str(c).lower(): def_idx = i; break
                        col_name = st.selectbox(f"Chọn Cột ({label})", all_cols, index=def_idx, key=f"{key_prefix}_col_{suffix}")
                    
                    if col_name and col_name != "—":
                        with r2:
                            top_vals = df_in[col_name].astype(str).value_counts().head(200).index.tolist()
                            vals = st.multiselect(f"Giá trị ({col_name})", top_vals, key=f"{key_prefix}_val_{suffix}")
                            return col_name, vals
                return None, []

            # Render 4 filter
            c1_n, v1 = _render_sel("Vị trí", use_1, "region", "1")
            if c1_n and v1: mask &= df_in[c1_n].astype(str).isin(v1)
            
            c2_n, v2 = _render_sel("Kênh", use_2, "channel", "2")
            if c2_n and v2: mask &= df_in[c2_n].astype(str).isin(v2)
            
            c3_n, v3 = _render_sel("Sản phẩm", use_3, "prod", "3")
            if c3_n and v3: mask &= df_in[c3_n].astype(str).isin(v3)
            
            c4_n, v4 = _render_sel("Khách hàng", use_4, "cust", "4")
            if c4_n and v4: mask &= df_in[c4_n].astype(str).isin(v4)

            # Time Filter
            time_col, per_rule_out = None, "M"
            if use_t:
                with r1:
                    dt_cands = [c for c in df_in.columns if 'date' in str(c).lower() or 'time' in str(c).lower()]
                    dt_opts = ["—"] + dt_cands + [c for c in df_in.columns if c not in dt_cands]
                    time_col = st.selectbox("Cột Thời gian", dt_opts, key=f"{key_prefix}_col_time")
                
                if time_col and time_col != "—":
                    with r2:
                        c_gra, c_per = st.columns([1, 1.5])
                        per_rule_raw = c_gra.radio("Kỳ", ["Month", "Quarter", "Year"], horizontal=True, key=f"{key_prefix}_rule_time")
                        per_rule_out = {"Month":"M", "Quarter":"Q", "Year":"Y"}.get(per_rule_raw, "M")
                        
                        try:
                            ts = _clean_time(df_in[time_col])
                            periods = sorted(ts.dt.to_period(per_rule_out).astype(str).dropna().unique())
                            def_sel = periods[-3:] if len(periods) > 3 else periods
                            sel_t = c_per.multiselect(f"Chọn khoảng thời gian", periods, default=def_sel, key=f"{key_prefix}_val_time")
                            if sel_t: mask &= ts.dt.to_period(per_rule_out).astype(str).isin(sel_t)
                        except: st.warning("Lỗi định dạng thời gian.")

            n_remain = mask.sum()
            st.caption(f"⚡ Dữ liệu sau lọc: **{n_remain:,}** / {len(df_in):,} dòng.")
            
            # --- FIX QUAN TRỌNG: Trả về None nếu time_col là '—' ---
            safe_time_col = time_col if (time_col and time_col != "—") else None
            return df_in.loc[mask], safe_time_col, per_rule_out

    # GỌI HÀM LỌC (INLINE)
    dfx, time_col, per_rule = _render_filter_inline(df, "corr")

    if dfx.empty:
        st.warning("Không còn dữ liệu sau khi khoanh vùng. Vui lòng nới lỏng bộ lọc.")
        st.stop()

    # ====================== Chọn biến ======================
    st.markdown("### 🎯 Chọn biến (Target Y & Drivers X)")
    NUMS = dfx.select_dtypes(include=[np.number]).columns.tolist()
    if not NUMS:
        st.info("Không có cột numeric để tính tương quan.")
        st.stop()

    c1, c2 = st.columns([1, 2])
    y_col = c1.selectbox("Target (numeric Y)", NUMS, index=0)

    numeric_wo_y = [c for c in NUMS if c != y_col]
    var_rank = dfx[numeric_wo_y].var(numeric_only=True).sort_values(ascending=False)
    x_default = var_rank.head(min(10, len(var_rank))).index.tolist()
    x_cols = c2.multiselect("Drivers X (numeric, multi-select)", numeric_wo_y, default=x_default)

    if not x_cols:
        st.info("Chọn ít nhất 1 biến X để tính tương quan.")
        st.stop()

    # ==================== Làm sạch & tuỳ chọn ====================
    st.markdown("### 🧹 Làm sạch & tuỳ chọn")
    o1, o2, o3, o4 = st.columns([1, 1, 1, 1])
    drop_eq0 = o1.checkbox("Bỏ = 0", value=False)
    drop_lt0 = o2.checkbox("Bỏ < 0", value=False)
    use_log  = o3.checkbox("log10 (áp dụng với biến >0)", value=False)
    method   = o4.radio("Phương pháp", ["Pearson", "Spearman"], horizontal=True)

    # ================== 4. Phân tích Chuỗi thời gian & Tương quan ==================
    with st.expander("⏱ Phân tích Chuỗi thời gian & Tương quan (Y vs X)", expanded=True):
        if guard(time_col is not None and len(x_cols) > 0, 
                "Cần bật **Time** trong Drill-down và chọn **Target/Driver** để xem phân tích Chuỗi thời gian."):
            
            st.markdown("#### ⚙️ Cấu hình Trend & Lag")
            c_driver, c_index, c_win, c_lag = st.columns([1.5, 1.2, 1.2, 1.2])
            
            drv_for_trend = c_driver.selectbox("Chọn 1 driver (X)", x_cols, index=0, key="corr_trend_x")
            
            c_agg = st.columns(2)
            aggY = c_agg[0].radio("Gộp Y theo", ["sum", "mean"], horizontal=True, key="corr_aggY")
            aggX = c_agg[1].radio("Gộp X theo", ["sum", "mean"], horizontal=True, key="corr_aggX")
            
            use_index = c_index.checkbox("Chuẩn hoá Index = 100", value=True, key="corr_use_index")
            win = c_win.slider("Cửa sổ Rolling-corr", 3, 24, 6, key="corr_roll_win")
            
            # --- Tính toán Trend và Lag ---
            tdt = _clean_time(dfx[time_col])
            tmp = (pd.DataFrame({"t": tdt, "Y": dfx[y_col], "X": dfx[drv_for_trend]})
                    .replace([np.inf, -np.inf], np.nan).dropna())
            if drop_lt0: tmp = tmp[(tmp["Y"] >= 0) & (tmp["X"] >= 0)]
            if drop_eq0: tmp = tmp[(tmp["Y"] != 0) & (tmp["X"] != 0)]
            if use_log:
                tmp = tmp[(tmp["Y"] > 0) & (tmp["X"] > 0)]
                tmp["Y"] = np.log10(tmp["Y"]); tmp["X"] = np.log10(tmp["X"])

            def _agg_by(freq_code):
                p = tmp["t"].dt.to_period(freq_code)
                g = tmp.groupby(p).agg(
                    Y=("Y", "sum" if aggY == "sum" else "mean"),
                    X=("X", "sum" if aggX == "sum" else "mean"),
                ).sort_index()
                g.index = g.index.to_timestamp(how="start")
                return g

            try_order = [per_rule] + [f for f in ("Q", "M") if f != per_rule]
            g, used_freq = None, per_rule
            for f in try_order:
                gg = _agg_by(f)
                if len(gg) >= 12: 
                    g, used_freq = gg, f
                    break
            
            # --- Hiển thị kết quả Trend/Lag ---
            if g is None:
                st.info("Chỉ có dưới 12 kỳ sau khi nhóm theo thời gian. Không đủ dữ liệu để phân tích Lagged/Rolling Corr.")
            else:
                g_plot = (g / g.iloc[0] * 100.0) if use_index else g
                y_left_title = "Y (Index=100)" if use_index else y_col
                y_right_title = "X (Index=100)" if use_index else drv_for_trend

                # --- Chart 1: Trend Dual-Axis Chart (figT) ---
                figT = make_subplots(specs=[[{"secondary_y": True}]])
                figT.add_bar(x=g_plot.index, y=g_plot["Y"], name=y_col, marker_color="#74b9ff", opacity=0.9, secondary_y=False)
                figT.add_scatter(x=g_plot.index, y=g_plot["X"], name=drv_for_trend, mode="lines+markers", line=dict(color="#e84393", width=2), marker=dict(size=5), secondary_y=True)
                figT.update_layout(height=380, bargap=0.35, legend=dict(orientation="h", y=1.1, x=0), margin=dict(l=10, r=10, t=10, b=10))
                if used_freq == "M":   figT.update_xaxes(dtick="M1",  tickformat="%b %Y")
                elif used_freq == "Q": figT.update_xaxes(dtick="M3",  tickformat="%b %Y")
                else:                   figT.update_xaxes(dtick="M12", tickformat="%Y")
                figT.update_yaxes(title_text=y_left_title, secondary_y=False)
                figT.update_yaxes(title_text=y_right_title, secondary_y=True, showgrid=False)

                # --- Chart 3: Lagged Correlation (figL) ---
                max_lag = c_lag.slider("Max Lag (Số kỳ trễ)", 1, min(len(g)//2, 24), 6, key="corr_lag_max", disabled=False)
                lags, corrs = [], []
                for lag in range(-max_lag, max_lag + 1):
                    if lag == 0: r, _, _ = corr_one(g["Y"], g["X"], method="pearson")
                    elif lag > 0: r, _, _ = corr_one(g["Y"], g["X"].shift(lag), method="pearson")
                    else: r, _, _ = corr_one(g["Y"].shift(-lag), g["X"], method="pearson")
                    lags.append(lag); corrs.append(r)
                lag_df = pd.DataFrame({"Lag": lags, "Correlation": corrs}).dropna()

                figL = go.Figure(go.Bar(x=lag_df["Lag"], y=lag_df["Correlation"], marker_color=np.where(lag_df["Correlation"] >= 0, "#27ae60", "#e74c3c")))
                n_eff = len(g) - max_lag; ci_bound = 2 / np.sqrt(n_eff) if n_eff > 4 else np.nan
                figL.add_hline(y=0, line=dict(color="#95a5a6", dash="dot"))
                if np.isfinite(ci_bound):
                    figL.add_hline(y=ci_bound, line=dict(color="#3498db", dash="dash"))
                    figL.add_hline(y=-ci_bound, line=dict(color="#3498db", dash="dash"), name="95% CI")
                figL.update_layout(height=380, title=f"Lagged Correlation", 
                                   xaxis_title=f"Lag (Kỳ trễ - {used_freq})", yaxis_title="r",
                                   yaxis=dict(range=[-1, 1]), showlegend=False, margin=dict(l=10, r=10, t=10, b=10))

                # --- Chart 2: Rolling Correlation Chart (figR) ---
                r_roll = g["Y"].rolling(win).corr(g["X"])
                figR = go.Figure()
                figR.add_scatter(x=g.index, y=r_roll, mode="lines+markers", name=f"Pearson-r rolling ({win})", line=dict(color="#2ecc71"), marker=dict(size=5), yaxis='y1')
                figR.add_hline(y=0, line=dict(color="#95a5a6", dash="dot"))
                figR.update_layout(height=300, margin=dict(l=10, r=10, t=10, b=10), hovermode="x unified", yaxis=dict(range=[-1, 1], title="Rolling r"))
                figR.update_xaxes(type="date")

                # --- OUTPUT CHARTS (Gom vào Columns) ---
                st.markdown("---")
                col_trend, col_lag = st.columns(2)
                with col_trend:
                    st.markdown("#### ⏱ Trend Y & X (Dual-Axis)")
                    st.plotly_chart(figT, use_container_width=True, config={"displayModeBar": False})
                with col_lag:
                    st.markdown("#### ⏳ Lagged Correlation (Correlogram)")
                    st.plotly_chart(figL, use_container_width=True, config={"displayModeBar": False})
                
                st.markdown("#### 🔄 Rolling Correlation")
                st.plotly_chart(figR, use_container_width=True, config={"displayModeBar": False})

                # --- SUMMARY TEXT (Gom lại) ---
                st.markdown("---")
                st.markdown("#### 🧠 Nhận định từ Chuỗi thời gian")
                
                last_r_roll = r_roll.dropna().iloc[-1] if r_roll.notna().any() else np.nan
                delta_y = (g.iloc[-1, 0] / g.iloc[0, 0] - 1) * 100 if len(g) >= 2 else np.nan
                delta_x = (g.iloc[-1, 1] / g.iloc[0, 1] - 1) * 100 if len(g) >= 2 else np.nan
                lbl = "tăng cùng chiều" if last_r_roll == last_r_roll and last_r_roll > 0 else ("giảm ngược chiều" if last_r_roll == last_r_roll and last_r_roll < 0 else "không rõ chiều")
                st.markdown(
                    f"- **Diễn biến Trend**: Y `{_fmt(delta_y,'.1f')}%`, X `{_fmt(delta_x,'.1f')}%` từ kỳ đầu → kỳ cuối.  "
                    f"- **Rolling-r (gần nhất)**: r={_fmt(last_r_roll,'.3f')} ⇒ **{lbl}** trong cửa sổ {win} kỳ."
                )
                
                best_lag = lag_df.iloc[lag_df["Correlation"].abs().argmax()]
                lag_dir = "Y (hiện tại) bị ảnh hưởng bởi X (quá khứ)" if best_lag["Lag"] > 0 else ("X (hiện tại) bị ảnh hưởng bởi Y (quá khứ)" if best_lag["Lag"] < 0 else "Tác động tức thì")
                st.info(f"💡 **Tương quan Lagged mạnh nhất:** xảy ra ở **Lag {best_lag['Lag']}** (r={best_lag['Correlation']:.3f}). \n\n*Nhận định: {lag_dir} (với độ trễ {abs(best_lag['Lag'])} kỳ).*")
                
    # ==================== 5. Tính tương quan X~Y (Giữ nguyên) ====================
    # (Đoạn code này tính toán r, p, n cho tất cả các biến X so với Y)
    rows = []
    for col in x_cols:
        xx, yy = prepare_xy(dfx[col], dfx[y_col], drop_lt0, drop_eq0, use_log)
        r, p, n = corr_one(xx, yy, method=method)
        lo, hi = fisher_ci(r, n) if method.lower() == "pearson" else (np.nan, np.nan)
        rows.append({
            "X": col, "N": n, "r": r, "p_value": p, "CI_low": lo, "CI_high": hi,
            "abs_r": abs(r), "direction": "dương (+)" if r == r and r > 0 else ("âm (−)" if r == r and r < 0 else "—"),
            "strength": strength_label(abs(r)) if r == r else "—",
        })
    corr_tbl = pd.DataFrame(rows).sort_values("abs_r", ascending=False).reset_index(drop=True)

    # ==================== Tính tương quan X~Y ====================
    rows = []
    for col in x_cols:
        xx, yy = prepare_xy(dfx[col], dfx[y_col], drop_lt0, drop_eq0, use_log)
        r, p, n = corr_one(xx, yy, method=method)
        lo, hi = fisher_ci(r, n) if method.lower() == "pearson" else (np.nan, np.nan)
        rows.append({
            "X": col, "N": n, "r": r, "p_value": p, "CI_low": lo, "CI_high": hi,
            "abs_r": abs(r), "direction": "dương (+)" if r == r and r > 0 else ("âm (−)" if r == r and r < 0 else "—"),
            "strength": strength_label(abs(r)) if r == r else "—",
        })
    corr_tbl = pd.DataFrame(rows).sort_values("abs_r", ascending=False).reset_index(drop=True)

    # ======================== KPIs nhanh ========================
    cA, cB, cC, cD = st.columns(4)
    cA.metric("Số biến X", f"{len(x_cols)}")
    cB.metric("n tối đa", f"{int(corr_tbl['N'].max() if len(corr_tbl) else 0):,}")
    best = corr_tbl.iloc[0] if len(corr_tbl) else None
    cC.metric("Mạnh nhất (|r|)", f"{_fmt(best['abs_r'],'.3f') if best is not None else '—'}")
    sig_rate = (corr_tbl["p_value"] < 0.05).mean() if _HAS_SCIPY and len(corr_tbl) else np.nan
    cD.metric("% quan hệ p<.05", f"{_fmt(100*sig_rate,'.1f') if sig_rate==sig_rate else '—'}%")

    # ==================== Bar r + CI (Pearson) ====================
    st.markdown("### 📊 Correlation với Target (có khoảng tin cậy)")
    topN = st.slider("Hiển thị Top-N theo |r|", 3, min(30, len(corr_tbl)), min(10, len(corr_tbl)))
    view_df = corr_tbl.head(topN).copy()
    colors = np.where(view_df["r"] >= 0, "#27ae60", "#e74c3c")

    fig_bar = go.Figure()
    fig_bar.add_bar(
        x=view_df["X"], y=view_df["r"], marker_color=colors,
        text=[_fmt(v, ".3f") for v in view_df["r"]], textposition="outside", cliponaxis=False,
        error_y=dict(
            type="data",
            array=(view_df["CI_high"] - view_df["r"]).astype(float),
            arrayminus=(view_df["r"] - view_df["CI_low"]).astype(float),
            visible=True if method.lower() == "pearson" else False
        ),
        customdata=np.stack([
            view_df["CI_low"].fillna(np.nan),
            view_df["CI_high"].fillna(np.nan),
            view_df["N"].fillna(0),
            view_df["p_value"].fillna(np.nan)
        ], axis=1),
        hovertemplate=(
            "X=%{x}<br>r=%{y:.3f}"
            + ("<br>95% CI=[%{customdata[0]:.3f}; %{customdata[1]:.3f}]" if method.lower()=="pearson" else "")
            + "<br>n=%{customdata[2]:,}"
            + ("<br>p=%{customdata[3]:.4f}" if _HAS_SCIPY else "")
            + "<extra></extra>"
        )
    )
    fig_bar.update_layout(
        height=440, xaxis_title="Biến X", yaxis_title=f"r ({method})",
        margin=dict(l=10, r=10, t=10, b=50),
        yaxis=dict(range=[min(-1, float(view_df["r"].min()) - 0.05), max(1, float(view_df["r"].max()) + 0.05)])
    )
    st.plotly_chart(fig_bar, use_container_width=True, config={"displayModeBar": False})

    # ==================== Bảng chi tiết ====================
    st.markdown("### 📋 Bảng chi tiết (r, CI, p, n, dấu, mức độ)")
    show_cols = ["X", "N", "r", "CI_low", "CI_high", "p_value", "direction", "strength"]
    st.dataframe(
        corr_tbl[show_cols].rename(columns={
            "X": "Biến X", "N": "n", "r": "r", "CI_low": "CI thấp", "CI_high": "CI cao",
            "p_value": "p-value", "direction": "Dấu", "strength": "Mức độ"
        }),
        use_container_width=True, hide_index=True,
        height=min(480, 32*(len(corr_tbl)+1))
    )

    # ==================== Heatmap (giới hạn) ====================
    st.markdown("### 🌡️ Heatmap tương quan")
    cmax1, cmax2 = st.columns([2, 1])
    max_h_cols = cmax2.slider("Tối đa số biến", 4, min(20, len(x_cols) + 1), min(12, len(x_cols) + 1))
    top_vars = corr_tbl.head(max_h_cols - 1)["X"].tolist()
    hm_cols = [y_col] + top_vars
    sub = dfx[hm_cols].replace([np.inf, -np.inf], np.nan).dropna()
    if len(sub) >= 3:
        corrM = sub.corr(method="pearson" if method.lower()=="pearson" else "spearman")
        fig_hm = go.Figure(data=go.Heatmap(
            z=corrM.values, x=corrM.columns, y=corrM.index,
            zmin=-1, zmax=1, colorscale="RdBu", reversescale=True,
            colorbar=dict(title="r")
        ))
        fig_hm.update_layout(height=420, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig_hm, use_container_width=True, config={"displayModeBar": False})
    else:
        st.info("Không đủ dữ liệu sạch để vẽ heatmap.")

    # ==================== Scatter (top mối quan hệ) ====================
    st.markdown("### 🔍 Scatter (Top quan hệ)")
    top_scatter = corr_tbl.head(min(4, len(corr_tbl)))["X"].tolist()
    max_points = st.slider("Giới hạn số điểm/biểu đồ (sampling)", 500, 10000, 3000, 500)
    if top_scatter:
        rows = int(np.ceil(len(top_scatter) / 2))
        fig_sc = make_subplots(rows=rows, cols=2, subplot_titles=[f"{x} vs {y_col}" for x in top_scatter])
        r_idx, c_idx = 1, 1
        for x in top_scatter:
            s = pd.concat([dfx[x], dfx[y_col]], axis=1).replace([np.inf, -np.inf], np.nan).dropna()
            if drop_lt0: s = s[(s.iloc[:, 0] >= 0) & (s.iloc[:, 1] >= 0)]
            if drop_eq0: s = s[(s.iloc[:, 0] != 0) & (s.iloc[:, 1] != 0)]
            if use_log:
                s = s[(s.iloc[:, 0] > 0) & (s.iloc[:, 1] > 0)]
                s = np.log10(s)

            if len(s) > max_points:
                s = s.sample(max_points, random_state=42)

            r_sc, p_sc, n_sc = corr_one(s.iloc[:, 0], s.iloc[:, 1], method=method)
            fig_sc.add_trace(
                go.Scattergl(x=s.iloc[:, 0], y=s.iloc[:, 1], mode="markers",
                             marker=dict(size=4, opacity=0.6),
                             hovertemplate=f"{x}=%{{x:,.4g}}<br>{y_col}=%{{y:,.4g}}<extra></extra>",
                             showlegend=False),
                row=r_idx, col=c_idx
            )
            # đường fit tuyến tính sơ bộ
            if len(s) >= 2 and np.ptp(s.iloc[:, 0]) > 0:
                coefs = np.polyfit(s.iloc[:, 0], s.iloc[:, 1], deg=1)
                xs = np.linspace(s.iloc[:, 0].min(), s.iloc[:, 0].max(), 100)
                ys = coefs[0] * xs + coefs[1]
                fig_sc.add_trace(go.Scatter(x=xs, y=ys, mode="lines", line=dict(color="#d35400"),
                                            showlegend=False), row=r_idx, col=c_idx)

            fig_sc.update_xaxes(title_text=x, row=r_idx, col=c_idx)
            fig_sc.update_yaxes(title_text=y_col, row=r_idx, col=c_idx)

            # Annotation r/p/n — FIX xref/yref cho subplot đầu tiên
            axis_num = (r_idx - 1) * 2 + c_idx
            xref = "x domain" if axis_num == 1 else f"x{axis_num} domain"
            yref = "y domain" if axis_num == 1 else f"y{axis_num} domain"
            fig_sc.add_annotation(
                xref=xref, yref=yref, x=1.0, y=1.04, xanchor="right",
                text=f"r={_fmt(r_sc,'.3f')}, n={n_sc}{'' if not _HAS_SCIPY else f', p={_fmt(p_sc,'.3g')}'}",
                showarrow=False, font=dict(size=12)
            )

            c_idx += 1
            if c_idx > 2: c_idx = 1; r_idx += 1

        fig_sc.update_layout(height=300*rows, margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig_sc, use_container_width=True, config={"displayModeBar": False})
    else:
        st.info("Chọn thêm biến X để xem scatter.")

    # ==================== Collinearity giữa các X ====================
    st.markdown("### 🧯 Cảnh báo collinearity giữa các X")
    if len(x_cols) >= 2:
        subx = dfx[x_cols].replace([np.inf, -np.inf], np.nan).dropna()
        if len(subx) >= 5:
            cxx = subx.corr(method="pearson" if method.lower()=="pearson" else "spearman").abs()
            pairs, cols = [], cxx.columns.tolist()
            for i in range(len(cols)):
                for j in range(i+1, len(cols)):
                    pairs.append((cols[i], cols[j], cxx.iloc[i, j]))
            col_warn = pd.DataFrame(pairs, columns=["X1", "X2", "|r|"]).sort_values("|r|", ascending=False)
            st.dataframe(col_warn.head(10), use_container_width=True, hide_index=True)
            if (col_warn["|r|"] > 0.8).any():
                st.warning("Có cặp X tương quan cao (|r|>0.8). Cân nhắc chọn bớt để tránh trùng thông tin.")
        else:
            st.info("Không đủ dữ liệu sạch để kiểm tra collinearity.")
    else:
        st.caption("Cần ≥2 biến X để kiểm tra collinearity.")

    # ==================== Nhận định từ dữ liệu hiện tại ====================
    st.markdown("### 🧠 Nhận định từ dữ liệu hiện tại")
    bullets = []

    pos = corr_tbl[corr_tbl["r"] > 0].head(3)
    neg = corr_tbl[corr_tbl["r"] < 0].head(3)
    top_x_col = corr_tbl.loc[0, 'X'] if len(corr_tbl) else None

    if not pos.empty:
        s_txt = "; ".join(
            [f"{r.X} (r={_fmt(r.r,'.3f')}, {r.strength}, {r.direction}"
             + (f", CI[{_fmt(r.CI_low,'.3f')};{_fmt(r.CI_high,'.3f')}]" if method.lower()=="pearson" else "")
             + (f", p={_fmt(r.p_value,'.3g')}" if _HAS_SCIPY else "")
             + f", n={int(r.N)})"
             for _, r in pos.iterrows()]
        )
        bullets.append(f"**Tăng cùng chiều với {y_col}**: {s_txt}.")

    if not neg.empty:
        s_txt = "; ".join(
            [f"{r.X} (r={_fmt(r.r,'.3f')}, {r.strength}, {r.direction}"
             + (f", CI[{_fmt(r.CI_low,'.3f')};{_fmt(r.CI_high,'.3f')}]" if method.lower()=="pearson" else "")
             + (f", p={_fmt(r.p_value,'.3g')}" if _HAS_SCIPY else "")
             + f", n={int(r.N)})"
             for _, r in neg.iterrows()]
        )
        bullets.append(f"**Giảm ngược chiều với {y_col}**: {s_txt}.")

    if method.lower() == "pearson":
        unstable = corr_tbl[(corr_tbl["CI_low"] < 0) & (corr_tbl["CI_high"] > 0)].head(5)
        if len(unstable):
            bullets.append("**Không chắc chắn (CI cắt 0)**: " + ", ".join([f"{r.X}" for _, r in unstable.iterrows()]))

    if _HAS_SCIPY:
        weak_sig = corr_tbl[(corr_tbl["abs_r"] < 0.3) & (corr_tbl["p_value"] < 0.05)].head(5)
        if len(weak_sig):
            bullets.append("**p<.05 nhưng hiệu ứng nhỏ**: " + ", ".join([f"{r.X}" for _, r in weak_sig.iterrows()]))

    bullets.append("**Gợi ý**: ưu tiên biến **|r|≥0.5** (mạnh). Nếu **Spearman ≫ Pearson** → quan hệ có thể **phi tuyến**; nên xem scatter & cân nhắc biến đổi.")
    st.markdown("\n".join([f"- {b}" for b in bullets]) if bullets else "Chưa đủ thông tin để nhận định.")

# ------------------------------- TAB 3: Benford -------------------------------
with TAB4:
    for k in ['bf1_res','bf2_res','bf1_col','bf2_col']:
        if k not in SS: SS[k]=None
    st.subheader('🔢 Benford Law — 1D & 2D')
    base_df = DF_FULL
    if not NUM_COLS:
        st.info('Không có cột numeric để chạy Benford.')
    else:
        run_on_full = True
        data_for_benford = DF_FULL
        # info removed
        c1,c2 = st.columns(2)
        with c1:
            amt1 = st.selectbox('Amount (1D)', NUM_COLS, key='bf1_col')
            if st.button('Run Benford 1D', key='btn_bf1'):
                ok,msg = _benford_ready(data_for_benford[amt1])
                if not ok: st.warning(msg)
                else: SS['bf1_res']=_benford_1d(data_for_benford[amt1])
        with c2:
            default_idx = 1 if len(NUM_COLS)>1 else 0
            amt2 = st.selectbox('Amount (2D)', NUM_COLS, index=default_idx, key='bf2_col')
            if st.button('Run Benford 2D', key='btn_bf2'):
                ok,msg = _benford_ready(data_for_benford[amt2])
                if not ok: st.warning(msg)
                else: SS['bf2_res']=_benford_2d(data_for_benford[amt2])
        g1,g2 = st.columns(2)
        with g1:
            if SS.get('bf1_res'):
                r=SS['bf1_res']; tb, var, p, MAD = r['table'], r['variance'], r['p'], r['MAD']
                if HAS_PLOTLY:
                    fig1 = go.Figure(); fig1.add_trace(go.Bar(x=tb['digit'], y=tb['observed_p'], name='Observed'))
                    fig1.add_trace(go.Scatter(x=tb['digit'], y=tb['expected_p'], name='Expected', mode='lines', line=dict(color='#F6AE2D')))
                    src_tag = 'FULL' if (SS['df'] is not None and SS.get('bf_use_full')) else 'SAMPLE'
                    fig1.update_layout(title=f'Benford 1D — Obs vs Exp ({SS.get("bf1_col")}, {src_tag})', height=340)
                    st_plotly(fig1)
                # --- Data quality — cột 1D đã chọn ---
                    _raw1 = data_for_benford[SS.get('bf1_col')]
                    _num1 = pd.to_numeric(_raw1, errors='coerce')
                    _total1 = len(_raw1)
                    _none_like1 = _raw1.astype('string').str.strip().str.lower().isin(['none','null']).sum()
                    _n_nan1  = _num1.isna().sum()
                    _n_zero1 = (_num1 == 0).sum()
                    _n_pos1  = (_num1 > 0).sum()
                    _n_neg1  = (_num1 < 0).sum()
                    _used1   = (_num1 != 0).sum()            
                    _base_clean1 = max(_total1 - _n_nan1 - _n_zero1, 0)
                    
                    qdf1 = pd.DataFrame({
                        'type': ['Total rows','NaN (numeric)','None/Null (text)','Zero (==0)',
                                 'Positive (>0)','Negative (<0)','Used for Benford (≠0)'],
                        'count': [int(_total1), int(_n_nan1), int(_none_like1), int(_n_zero1),
                                  int(_n_pos1), int(_n_neg1), int(_used1)]
                    })
                    qdf1['% vs total'] = (qdf1['count'] / _total1 * 100.0).round(2) if _total1>0 else 0.0
                    qdf1['% vs non-missing&non-zero'] = (
                        (qdf1['count'] / _base_clean1 * 100.0).round(2) if _base_clean1>0 else 0.0
                    )
                    st.caption('📋 Data quality — cột 1D đã chọn')
                    st_df(qdf1, use_container_width=True, height=180)
                    # --- Bảng % 1D (expected% / observed%) & diff% = observed% - expected% ---
                    color_thr_pct = 5.0  # drill-down theo chuẩn 5%
                    
                    t1 = pd.DataFrame({
                        'digit': tb['digit'].astype(int),
                        'expected_%': tb['expected_p'] * 100.0,
                        'observed_%': tb['observed_p'] * 100.0,
                    })
                    t1['diff_%'] = t1['observed_%'] - t1['expected_%']
                    
                    def _hl_percent1(v):
                        try:
                            return 'color: #d32f2f' if abs(float(v)) >= color_thr_pct else ''
                        except Exception:
                            return ''
                    
                    sty1 = (
                        t1.round(2)
                          .style
                          .format({'expected_%': '{:.2f}%', 'observed_%': '{:.2f}%', 'diff_%': '{:.2f}%'})
                          .applymap(_hl_percent1, subset=['diff_%'])
                    )
                    st_df(sty1, use_container_width=True, height=220)
                    
                    # --- Drill-down 1D cho những digit lệch ≥5% (tính theo diff_% ở trên) ---
                    bad_digits_1d = t1.loc[t1['diff_%'].abs() >= color_thr_pct, 'digit'].astype(int).tolist()
                    if bad_digits_1d:
                        with st.expander('🔎 Drill-down 1D: các chữ số lệch (|diff%| ≥ 5%)', expanded=False):
                            mode1 = st.radio('Chế độ hiển thị', ['Ngắn gọn','Xổ hết'], index=0,
                                             horizontal=True, key='bf1_drill_mode')
                    
                            import re as _re_local
                            def _digits_str(x):
                                xs = ("%.15g" % float(x))
                                return _re_local.sub(r"[^0-9]", "", xs).lstrip("0")
                            def _first1(v):
                                ds = _digits_str(v)
                                return int(ds[0]) if len(ds) >= 1 else np.nan
                    
                            s1_num = pd.to_numeric(data_for_benford[SS['bf1_col']], errors='coerce') \
                                       .replace([np.inf, -np.inf], np.nan).dropna().abs()
                            d1 = s1_num.apply(_first1).dropna()
                    
                            for dg in bad_digits_1d:
                                idx = d1[d1 == dg].index
                                st.markdown(f'**Digit {dg}** — {len(idx):,} rows')
                                if len(idx) == 0:
                                    continue
                                if mode1 == 'Xổ hết':
                                    st_df(data_for_benford.loc[idx].head(2000), use_container_width=True, height=260)
                                else:
                                    st_df(data_for_benford.loc[idx, [SS.get("bf1_col")]].head(200),
                                          use_container_width=True, height=220)
                    
                    # --- Thông điệp trạng thái dùng ngưỡng slider (so sánh theo tỷ lệ, không phải % point) ---
                    thr = SS.get('risk_diff_threshold', 0.05)               # ví dụ 0.05 = 5%
                    maxdiff_pp = float(t1['diff_%'].abs().max())            # % point
                    maxdiff_ratio = maxdiff_pp / 100.0                      # đổi về tỷ lệ để so với thr
                    
                    msg = '🟢 Green'
                    if maxdiff_ratio >= 2*thr:
                        msg = '🚨 Red'
                    elif maxdiff_ratio >= thr:
                        msg = '🟡 Yellow'
                    
                    sev = '🟢 Green'
                    if (p < 0.01) or (MAD > 0.015): sev = '🚨 Red'
                    elif (p < 0.05) or (MAD > 0.012): sev = '🟡 Yellow'
                    
                    st.info(f"Diff% status: {msg} • p={p:.4f}, MAD={MAD:.4f} ⇒ Benford severity: {sev}")

                    
        with g2:
            if SS.get('bf2_res'):
                r2=SS['bf2_res']; tb2, var2, p2, MAD2 = r2['table'], r2['variance'], r2['p'], r2['MAD']
                if HAS_PLOTLY:
                    fig2 = go.Figure(); fig2.add_trace(go.Bar(x=tb2['digit'], y=tb2['observed_p'], name='Observed'))
                    fig2.add_trace(go.Scatter(x=tb2['digit'], y=tb2['expected_p'], name='Expected', mode='lines', line=dict(color='#F6AE2D')))
                    src_tag = 'FULL' if (SS['df'] is not None and SS.get('bf_use_full')) else 'SAMPLE'
                    fig2.update_layout(title=f'Benford 2D — Obs vs Exp ({SS.get("bf2_col")}, {src_tag})', height=340)
                    st_plotly(fig2)
                # --- Data quality — cột 2D đã chọn ---
                    _raw2 = data_for_benford[SS.get('bf2_col')]
                    _num2 = pd.to_numeric(_raw2, errors='coerce')
                    _total2 = len(_raw2)
                    _none_like2 = _raw2.astype('string').str.strip().str.lower().isin(['none','null']).sum()
                    _n_nan2  = _num2.isna().sum()
                    _n_zero2 = (_num2 == 0).sum()
                    _n_pos2  = (_num2 > 0).sum()
                    _n_neg2  = (_num2 < 0).sum()
                    _used2   = (_num2 != 0).sum()            # Used for Benford: > 0 (giữ đúng logic tab này)
                    _base_clean2 = max(_total2 - _n_nan2 - _n_zero2, 0)
                    
                    qdf2 = pd.DataFrame({
                        'type': ['Total rows','NaN (numeric)','None/Null (text)','Zero (==0)',
                                 'Positive (>0)','Negative (<0)','Used for Benford (≠0)'],
                        'count': [int(_total2), int(_n_nan2), int(_none_like2), int(_n_zero2),
                                  int(_n_pos2), int(_n_neg2), int(_used2)]
                    })
                    qdf2['% vs total'] = (qdf2['count'] / _total2 * 100.0).round(2) if _total2>0 else 0.0
                    qdf2['% vs non-missing&non-zero'] = (
                        (qdf2['count'] / _base_clean2 * 100.0).round(2) if _base_clean2>0 else 0.0
                    )
                    st.caption('📋 Data quality — cột 2D đã chọn')
                    st_df(qdf2, use_container_width=True, height=180)
                    
                   # --- Bảng % 2D (expected% / observed%) & diff% = observed% - expected% ---
                    color_thr_pct = 5.0  # drill-down theo chuẩn 5%
                    
                    t2 = pd.DataFrame({
                        'digit': tb2['digit'].astype(int),
                        'expected_%': tb2['expected_p'] * 100.0,
                        'observed_%': tb2['observed_p'] * 100.0,
                    })
                    t2['diff_%'] = t2['observed_%'] - t2['expected_%']
                    
                    def _hl_percent2(v):
                        try:
                            return 'color: #d32f2f' if abs(float(v)) >= color_thr_pct else ''
                        except Exception:
                            return ''
                    
                    sty2 = (
                        t2.round(2)
                          .style
                          .format({'expected_%': '{:.2f}%', 'observed_%': '{:.2f}%', 'diff_%': '{:.2f}%'})
                          .applymap(_hl_percent2, subset=['diff_%'])
                    )
                    st_df(sty2, use_container_width=True, height=220)
                    
                    # --- Drill-down 2D cho những digit lệch ≥5% (tính theo diff_% ở trên) ---
                    bad_digits_2d = t2.loc[t2['diff_%'].abs() >= color_thr_pct, 'digit'].astype(int).tolist()
                    if bad_digits_2d:
                        with st.expander('🔎 Drill-down 2D: các chữ số lệch (|diff%| ≥ 5%)', expanded=False):
                            mode2 = st.radio('Chế độ hiển thị', ['Ngắn gọn','Xổ hết'], index=0,
                                             horizontal=True, key='bf2_drill_mode')
                    
                            import re as _re_local
                            def _digits_str(x):
                                xs = ("%.15g" % float(x))
                                return _re_local.sub(r"[^0-9]", "", xs).lstrip("0")
                            def _first2(v):
                                ds = _digits_str(v)
                                return int(ds[:2]) if len(ds) >= 2 else (int(ds) if len(ds) == 1 and ds != '0' else np.nan)
                    
                            s2_num = pd.to_numeric(data_for_benford[SS['bf2_col']], errors='coerce') \
                                       .replace([np.inf, -np.inf], np.nan).dropna().abs()
                            d2 = s2_num.apply(_first2).dropna()
                    
                            for dg in bad_digits_2d:
                                idx = d2[d2 == dg].index
                                st.markdown(f'**Digit {dg}** — {len(idx):,} rows')
                                if len(idx) == 0:
                                    continue
                                if mode2 == 'Xổ hết':
                                    st_df(data_for_benford.loc[idx].head(2000), use_container_width=True, height=260)
                                else:
                                    st_df(data_for_benford.loc[idx, [SS.get("bf2_col")]].head(200),
                                          use_container_width=True, height=220)
                    
                    # --- Thông điệp trạng thái dùng ngưỡng slider (so sánh theo tỷ lệ, không phải % point) ---
                    thr = SS.get('risk_diff_threshold', 0.05)
                    maxdiff2_pp = float(t2['diff_%'].abs().max())
                    maxdiff2_ratio = maxdiff2_pp / 100.0
                    
                    msg2 = '🟢 Green'
                    if maxdiff2_ratio >= 2*thr:
                        msg2 = '🚨 Red'
                    elif maxdiff2_ratio >= thr:
                        msg2 = '🟡 Yellow'
                    
                    sev2 = '🟢 Green'
                    if (p2 < 0.01) or (MAD2 > 0.015): sev2 = '🚨 Red'
                    elif (p2 < 0.05) or (MAD2 > 0.012): sev2 = '🟡 Yellow'
                    
                    st.info(f"Diff% status: {msg2} • p={p2:.4f}, MAD={MAD2:.4f} ⇒ Benford severity: {sev2}")

# ------------------------------ TAB ? : Statistics Test (ANOVA & Nonparametric, balanced UI) ------------------------------
# ------------------------------ TAB 5 : Statistics Test (ANOVA & Nonparametric, balanced UI) ------------------------------
with TAB5:
    import numpy as np, pandas as pd, re
    import plotly.express as px
    import plotly.graph_objects as go
    from scipy import stats
    import streamlit as st

    st.subheader("📊 Hypothesis — ANOVA & Nonparametric")

    # ===== Data guard =====
    DF = SS.get('df')
    if DF is None or len(DF) == 0:
        st.info("Hãy nạp dữ liệu trước.")
        st.stop()

    # ===== Type helpers =====
    def is_num(c):
        try: return pd.api.types.is_numeric_dtype(DF[c])
        except: return False
    def is_dt(c):
        if c not in DF.columns: return False
        if pd.api.types.is_datetime64_any_dtype(DF[c]): return True
        return bool(re.search(r'(date|time|ngày|thời gian)', str(c), flags=re.I))
    def is_cat(c):
        return (not is_num(c)) and (not is_dt(c))

    NUM_COLS = [c for c in DF.columns if is_num(c)]
    CAT_COLS = [c for c in DF.columns if is_cat(c)]

    # ===== Small utils =====
    def topn_cat(s: pd.Series, n=10):
        vc = s.astype(str).fillna("NaN").value_counts()
        keep = vc.index[:n].tolist()
        return s.astype(str).where(s.astype(str).isin(keep), "Khác")

    def group_summary(y, g):
        d = pd.DataFrame({"y": y, "g": g}).dropna()
        if d.empty: 
            return pd.DataFrame(columns=["group","n","mean","std","median","se","ci95"])
        agg = d.groupby("g")["y"].agg(n="count", mean="mean", std="std", median="median")
        agg["se"] = agg["std"] / np.sqrt(agg["n"].clip(lower=1))
        agg["ci95"] = 1.96 * agg["se"]
        out = agg.reset_index().rename(columns={"g":"group"})
        return out.replace([np.inf, -np.inf], np.nan).fillna(0.0)

    def holm_bonferroni(pvals, labels):
        p = np.asarray(pvals, dtype=float); m = len(p); order = np.argsort(p)
        adj = np.empty(m, dtype=float); running_max = 0.0
        for r, idx in enumerate(order):
            adj_val = (m - r) * p[idx]
            running_max = max(running_max, adj_val)
            adj[idx] = min(1.0, running_max)
        return pd.DataFrame({"pair": labels, "p_raw": p, "p_adj_holm": adj}).sort_values("p_adj_holm")

    def one_way_anova_fast(y, g):
        d = pd.DataFrame({"y": pd.to_numeric(y, errors="coerce"), "g": g}).dropna()
        if d["g"].nunique() < 2 or len(d) < 3:
            return np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan
        try:
            levene_p = stats.levene(*[grp["y"].values for _, grp in d.groupby("g")], center="median").pvalue
        except Exception:
            levene_p = np.nan
        grp = d.groupby("g")["y"].agg(n="count", mean="mean"); ssq = d.assign(y2=d["y"]**2).groupby("g")["y2"].sum()
        grand_mean = float(d["y"].mean()); ssb = float((grp["n"] * (grp["mean"] - grand_mean) ** 2).sum())
        ssw = float((ssq - grp["n"] * (grp["mean"] ** 2)).sum()); sst = float(((d["y"] - grand_mean) ** 2).sum())
        k = int(grp.shape[0]); n = int(d.shape[0]); df1 = k - 1; df2 = max(n - k, 1)
        msb = ssb / max(df1, 1); msw = ssw / max(df2, 1)
        F = (msb / msw) if msw > 0 else np.inf
        p = 1 - stats.f.cdf(F, df1, df2) if np.isfinite(F) else 0.0
        eta2 = (ssb / sst) if sst > 0 else np.nan
        omega2 = ((ssb - df1 * msw) / (sst + msw)) if (sst + msw) > 0 else np.nan
        return float(F), float(p), float(df1), float(df2), float(eta2), float(omega2), float(levene_p)

    def kruskal_eps2(H, k, n):
        return float((H - (k - 1)) / (n - k)) if (n - k) > 0 else np.nan

    def _dtype_name(col):
        if col is None: return "—"
        try:
            if pd.api.types.is_datetime64_any_dtype(DF[col]): return "datetime"
            if pd.api.types.is_numeric_dtype(DF[col]): return "numeric"
            return "categorical"
        except Exception:
            return "unknown"

    def _type_hint(label, col, expect):
        actual = _dtype_name(col)
        ok = (actual == expect)
        icon = "✅" if ok else "⚠️"
        st.caption(f"{icon} {label}: `{col}` · {actual} (yêu cầu: {expect})")

    # ===== (THAY ĐỔI) Đặt Hướng dẫn Popover gọn gàng BÊN TRÊN Tabs =====
    st.markdown(
        """
        <style>
            /* Thu nhỏ popover trigger button */
            button[data-testid="stPopoverTrigger"] {
                padding: 0px 6px;
                font-size: 0.9rem;
                margin-left: 8px;
            }
        </style>
        """, unsafe_allow_html=True
    )

    col_t1, col_t2 = st.columns(2)
    with col_t1:
        with st.popover("ⓘ ANOVA (Parametric)?"):
            st.markdown("**Khi nào dùng:** Dữ liệu Y (numeric) của bạn *gần* tuân theo phân phối chuẩn.")
            st.markdown("**Kiểm tra (ở TAB 2):**")
            st.markdown("* `Skewness` (độ lệch): Lý tưởng là `gần 0` (ví dụ: trong khoảng -1 đến 1).")
            st.markdown("* `Kurtosis` (độ nhọn): Lý tưởng là `gần 0` (ví dụ: trong khoảng -2 đến 2).")
            st.markdown("* `Normality p-value`: > 0.05.")
            st.markdown("---")
            st.markdown("⚠️ *Nếu vi phạm nặng (p < 0.05, skew/kurtosis lớn), hãy ưu tiên **Nonparametric**.*")
    with col_t2:
        with st.popover("ⓘ Nonparametric?"):
             st.markdown("**Khi nào dùng:**")
             st.markdown("* Dữ liệu Y *lệch nhiều* hoặc *không chuẩn* (vi phạm giả định Parametric).")
             st.markdown("* Dữ liệu có *nhiều outliers* (bị kéo đuôi).")
             st.markdown("* Bạn muốn so sánh *trung vị (median)* thay vì *trung bình (mean)*.")

    tab_a, tab_np = st.tabs(["ANOVA (Parametric)", "Nonparametric"])

# ====================== ANOVA (Parametric) — (UI Tinh giản) ======================
    with tab_a:
        mode_a = st.radio("Testing", ["Independent (between)", "Repeated (within)"], horizontal=True, key="anova_mode")
    
        # ---------- Independent (between) ----------
        if mode_a == "Independent (between)":
            if len(NUM_COLS) == 0 or len(CAT_COLS) == 0:
                st.info("Cần tối thiểu 1 cột numeric (Y) và 1 cột categorical (factor).")
            else:
                box_top = st.container(border=True)
                with box_top:
                    st.markdown("### ANOVA — Independent (between)")
                    with st.popover("ⓘ Hướng dẫn"):
                        st.markdown(
                            "**Independent (between):**\n\n"
                            "So sánh các nhóm *khác biệt, độc lập* với nhau.\n\n"
                            "*Ví dụ: So sánh doanh số giữa Region A, Region B, và Region C.*"
                        )
                    
                    y_col  = st.selectbox("🎯 Dependent (numeric)", NUM_COLS, key="av_y")
                    a_col  = st.selectbox("🏷️ Factor A (categorical)", CAT_COLS, key="av_a")
                    
                    use_two = st.toggle("➕ Two-way ANOVA (Thêm Factor B)", value=False, key="av_two")
                    b_col = None
                    if use_two:
                        # *** (THAY ĐỔI) Lọc thông minh: B không thể là A ***
                        b_choices = [c for c in CAT_COLS if c != a_col]
                        if not b_choices:
                            st.warning("Không đủ cột categorical khác để làm Factor B.")
                            use_two = False
                        else:
                            b_col = st.selectbox("🏷️ Factor B (categorical)", b_choices, key="av_b")
                    
                    _type_hint("Dependent", y_col, "numeric")
                    _type_hint("Factor A", a_col, "categorical")
                    if use_two and b_col:
                        _type_hint("Factor B", b_col, "categorical")
    
                # Controls (UI Tinh giản - Ẩn Cài đặt Nâng cao)
                box_ctl = st.container(border=True)
                with box_ctl:
                    st.checkbox("Hiện 95% CI (Biểu đồ)", value=True, key="av_ci")
                    if not use_two:
                        st.checkbox("Pairwise Post-hoc (So sánh cặp)", value=True, key="av_posthoc")
                    
                    # *** (THAY ĐỔI) Cài đặt nâng cao được ẩn đi (mặc định) ***
                    topN_A = 10
                    topN_B = 8
                    max_fit = 300_000
                    
                    run = st.button("▶️ Run", use_container_width=True, key="av_run")
    
                # Compute & report
                if run:
                    if not use_two:
                        # ----- One-way (TỰ ĐỘNG CHỌN T-TEST / ANOVA) -----
                        sub = DF[[y_col, a_col]].copy()
                        if len(sub) > max_fit:
                            sub = sub.sample(n=max_fit, random_state=42)
                        sub[a_col] = topn_cat(sub[a_col], n=topN_A)
    
                        y = pd.to_numeric(sub[y_col], errors="coerce")
                        g = sub[a_col].astype(str)
                        ok = (~y.isna()) & (~g.isna())
                        y, g = y[ok], g[ok]
    
                        summ = group_summary(y, g).sort_values("mean", ascending=False)
                        st.dataframe(summ, use_container_width=True, hide_index=True)
                        if summ.shape[0] < 2 or len(y) < 3:
                            st.warning("Không đủ nhóm/hàng để chạy.")
                            st.stop()
                        
                        k = summ.shape[0] # Số lượng nhóm

                        if k == 2:
                            # ----- (THAY ĐỔI) CHẠY WELCH T-TEST CHO 2 NHÓM -----
                            st.markdown("#### Kết quả Welch's T-test (cho 2 nhóm)")
                            groups = [y[g == lv].values for lv in summ["group"]]
                            tt = stats.ttest_ind(groups[0], groups[1], equal_var=False) # equal_var=False là Welch
                            
                            m1, m2, m3 = st.columns(3)
                            m1.metric("t-statistic", f"{tt.statistic:.4f}")
                            m2.metric("p-value", f"{tt.pvalue:.4g}")
                            m3.metric("Groups", f"{summ['group'].iloc[0]} vs {summ['group'].iloc[1]}")
                            
                            # *** (THAY ĐỔI) Dùng Violin Plot cho 2 nhóm ***
                            st.markdown("#### Biểu đồ phân phối (Violin)")
                            fig = px.violin(
                                pd.DataFrame({"group": g, "y": y}), 
                                x="group", y="y", 
                                box=True, points=False, 
                                labels={"group": a_col, "y": y_col}
                            )
                            st.plotly_chart(fig, use_container_width=True)
                            
                            if tt.pvalue < 0.05:
                                st.success(f"**Kết luận:** Có sự khác biệt có ý nghĩa thống kê giữa 2 nhóm (p < 0.05).")
                            else:
                                st.info(f"**Kết luận:** Chưa đủ bằng chứng về sự khác biệt giữa 2 nhóm (p >= 0.05).")

                        else:
                            # ----- CHẠY ANOVA (như cũ, cho 3+ nhóm) -----
                            st.markdown(f"#### Kết quả One-Way ANOVA (cho {k} nhóm)")
                            F, p, df1, df2, eta2, omega2, lev_p = one_way_anova_fast(y, g)
        
                            m1, m2, m3, m4 = st.columns(4)
                            m1.metric("F", f"{F:.3f}")
                            m2.metric("p-value", f"{p:.4g}")
                            m3.metric("η²", f"{eta2:.3f}" if not np.isnan(eta2) else "—")
                            m4.metric("ω²", f"{omega2:.3f}" if not np.isnan(omega2) else "—")
                            st.caption(f"Levene (phương sai bằng nhau) p = {lev_p:.4g}")

                            # *** CẢNH BÁO LEVENE (ĐÃ THÊM) ***
                            if lev_p < 0.05:
                                st.warning(
                                    "⚠️ Levene's p < 0.05: Giả định về phương sai bằng nhau đã bị vi phạm. "
                                    "Kết quả F-test của ANOVA có thể không đáng tin cậy. "
                                    "**Gợi ý:** Hãy ưu tiên kết quả Post-hoc 'Welch t-test' (vì nó robust) "
                                    "hoặc sử dụng tab 'Nonparametric' (Kruskal-Wallis)."
                                )
        
                            # chart
                            if not SS.get("av_ci"): # Bỏ check 'fast'
                                fig = px.bar(summ, x="group", y="mean", labels={"group": a_col, "mean": f"Mean {y_col}"})
                            else:
                                fig = go.Figure(go.Bar(x=summ["group"], y=summ["mean"],
                                                       error_y=dict(type="data", array=summ["ci95"], visible=True)))
                                fig.update_layout(yaxis_title=f"{y_col} (mean ± 95% CI)")
                            st.plotly_chart(fig, use_container_width=True)
        
                            # post-hoc (Welch t-test + Holm)
                            if SS.get("av_posthoc") and summ.shape[0] >= 2:
                                groups_dict = {lv: y[g == lv].values for lv in summ["group"]}
                                pvals, labs = [], []
                                group_names = summ["group"].tolist()
                                for i in range(len(group_names)):
                                    gi = group_names[i]; xi = groups_dict[gi]
                                    for j in range(i+1, len(group_names)):
                                        gj = group_names[j]; xj = groups_dict[gj]
                                        if len(xi) >= 2 and len(xj) >= 2:
                                            tt = stats.ttest_ind(xi, xj, equal_var=False)
                                            pvals.append(float(tt.pvalue)); labs.append(f"{gi} vs {gj}")
                                if pvals:
                                    adj = holm_bonferroni(np.array(pvals), np.array(labs))
                                    diffs = []
                                    for pair in adj["pair"]:
                                        gi, gj = str(pair).split(" vs ")
                                        mi = summ.loc[summ["group"] == gi, "mean"].values[0]
                                        mj = summ.loc[summ["group"] == gj, "mean"].values[0]
                                        diffs.append(mi - mj)
                                    adj["mean_diff"] = diffs
                                    st.dataframe(adj.head(50), use_container_width=True, hide_index=True)
                                    st.caption("Pairwise Welch t-test (Holm-adjusted).")
        
                            # *** (GIỮ NGUYÊN) Kết luận Mạnh/Yếu/Vừa ***
                            strength = ("yếu" if (np.isnan(eta2) or eta2 < 0.06) else ("vừa" if eta2 < 0.14 else "mạnh"))
                            best = str(summ.iloc[0]["group"]) if len(summ) else "—"
                            st.success(f"**Kết luận:** Khác biệt giữa các nhóm **{strength}** (η²={eta2:.2f}). Nhóm cao nhất: **{best}**.")
    
                    else:
                        # ----- Two-way ANOVA (OLS + anova_lm) -----
                        try:
                            import statsmodels.api as sm
                            import statsmodels.formula.api as smf
                        except Exception:
                            st.error("Two-way ANOVA cần `statsmodels`. Hãy cài đặt gói này.")
                            st.stop()
    
                        sub = DF[[y_col, a_col, b_col]].dropna().copy()
                        if len(sub) > max_fit:
                            sub = sub.sample(n=max_fit, random_state=42)
                        sub[a_col] = topn_cat(sub[a_col], n=topN_A)
                        sub[b_col] = topn_cat(sub[b_col], n=topN_B)
    
                        d = sub.rename(columns={y_col: "Y", a_col: "A", b_col: "B"})
                        d["Y"] = pd.to_numeric(d["Y"], errors="coerce")
                        d = d.dropna(subset=["Y"])
                        if d["A"].nunique() < 2 or d["B"].nunique() < 2:
                            st.warning("Cần ≥2 mức cho mỗi factor sau khi Top-N.")
                            st.stop()
    
                        model = smf.ols("Y ~ C(A) + C(B) + C(A):C(B)", data=d).fit()
                        an_tbl = sm.stats.anova_lm(model, typ=2)
                        st.dataframe(an_tbl, use_container_width=True)
    
                        if "Residual" in an_tbl.index and "sum_sq" in an_tbl.columns:
                            ss_res = float(an_tbl.loc["Residual", "sum_sq"])
                            def peta(row): 
                                ss = float(row["sum_sq"])
                                return ss / (ss + ss_res) if (ss + ss_res) > 0 else np.nan
                            peta_vals = an_tbl.apply(peta, axis=1)
                            pe = peta_vals.to_dict()
                        else:
                            pe = {}
    
                        def card_val(name, col):
                            if name in an_tbl.index:
                                Fv = an_tbl.loc[name, "F"]; pv = an_tbl.loc[name, "PR(>F)"]
                                ev = pe.get(name, np.nan)
                                col.metric(name.replace("C(","").replace(")",""), f"F={Fv:.2f}", f"p={pv:.3g}")
                                if not np.isnan(ev): col.caption(f"partial η² ≈ {ev:.3f}")
                            else:
                                col.metric(name, "—", "—")
    
                        c1, c2, c3 = st.columns(3)
                        card_val("C(A)", c1); card_val("C(B)", c2); card_val("C(A):C(B)", c3)
    
                        grp = d.groupby(["A","B"])["Y"].agg(n="count", mean="mean").reset_index()
                        fig = px.bar(grp, x="A", y="mean", color="B", barmode="group",
                                     labels={"A": a_col, "B": b_col, "mean": f"Mean {y_col}"})
                        st.plotly_chart(fig, use_container_width=True)
    
                        pA = float(an_tbl.loc["C(A)", "PR(>F)"]) if "C(A)" in an_tbl.index else np.nan
                        pB = float(an_tbl.loc["C(B)", "PR(>F)"]) if "C(B)" in an_tbl.index else np.nan
                        pI = float(an_tbl.loc["C(A):C(B)", "PR(>F)"]) if "C(A):C(B)" in an_tbl.index else np.nan
                        msg = []
                        if not np.isnan(pI) and pI < 0.05:
                            msg.append("**có tương tác A×B** (p<0.05) — nên đọc theo từng lát cắt.")
                        if not np.isnan(pA) and pA < 0.05:
                            msg.append("Factor **A** có ý nghĩa.")
                        if not np.isnan(pB) and pB < 0.05:
                            msg.append("Factor **B** có ý nghĩa.")
                        if not msg: msg = ["Chưa thấy hiệu ứng có ý nghĩa (p≥0.05)."]
                        st.success(" ; ".join(msg))
    
        # ---------- Repeated (within) ----------
        else:
            cand_id = [c for c in DF.columns if is_cat(c)]
            cand_factor = [c for c in CAT_COLS]
            if len(NUM_COLS) == 0 or len(cand_id) == 0 or len(cand_factor) == 0:
                st.info("Cần: 1 numeric (Y), 1 ID (subject), 1 categorical (condition).")
            else:
                box_top_r = st.container(border=True)
                with box_top_r:
                    st.markdown("### ANOVA — Repeated (within)")
                    with st.popover("ⓘ Hướng dẫn"):
                        st.markdown(
                            "**Repeated (within):**\n\n"
                            "So sánh *cùng một đối tượng* qua nhiều điều kiện hoặc thời gian.\n\n"
                            "*Ví dụ: So sánh doanh số của từng Cửa hàng vào Tháng 1, Tháng 2, và Tháng 3.*"
                        )
                    y_col = st.selectbox("🎯 Y (numeric)", NUM_COLS, key="av_rep_y")
                    id_col = st.selectbox("🧑‍🤝‍🧑 ID (subject)", cand_id, key="av_rep_id")
                    
                    # *** (THAY ĐỔI) Lọc thông minh: Condition không thể là ID ***
                    cond_choices = [c for c in cand_factor if c != id_col]
                    if not cond_choices:
                        st.warning("Không đủ cột categorical khác để làm Condition.")
                        st.stop()
                    else:
                        cond_col = st.selectbox("🏷️ Condition (within)", cond_choices, key="av_rep_cond")

                    _type_hint("Y", y_col, "numeric")
                    _type_hint("ID", id_col, "categorical")
                    _type_hint("Condition", cond_col, "categorical")
    
                box_ctl_r = st.container(border=True)
                with box_ctl_r:
                    # *** (THAY ĐỔI) Ẩn cài đặt nâng cao ***
                    max_subj_fit = 5_000
                    plot_subj = 80
                    run = st.button("▶️ Run", use_container_width=True, key="av_rep_run")
    
                if run:
                    try:
                        from statsmodels.stats.anova import AnovaRM
                    except Exception:
                        st.error("RM-ANOVA cần `statsmodels`. Bạn có thể dùng tab **Nonparametric → Friedman** như một thay thế.")
                        st.stop()
    
                    d0 = DF[[y_col, id_col, cond_col]].dropna().copy()
                    cnt = d0.groupby([id_col, cond_col]).size().unstack(cond_col).dropna()
                    keep_ids = cnt.index
                    d = d0[d0[id_col].isin(keep_ids)]
                    uniq_ids = d[id_col].unique()
                    if len(uniq_ids) > max_subj_fit:
                        keep = pd.Index(uniq_ids).sample(max_subj_fit, random_state=42)
                        d = d[d[id_col].isin(keep)]
    
                    if d.empty or d[cond_col].nunique() < 2:
                        st.warning("Không đủ subject/điều kiện để chạy RM-ANOVA.")
                        st.stop()
    
                    model = AnovaRM(d, depvar=y_col, subject=id_col, within=[cond_col])
                    res = model.fit()
                    # *** (THAY ĐỔI) Hiển thị DataFrame sạch sẽ ***
                    st.dataframe(res.anova_summary, use_container_width=True)
    
                    # Means + spaghetti
                    pivot = d.pivot_table(index=id_col, columns=cond_col, values=y_col, aggfunc="mean")
                    levels = list(pivot.columns)
                    means = pivot.mean().reset_index()
                    means.columns = ["cond","mean"]
                    fig = px.line(means, x="cond", y="mean", markers=True)
                    
                    # *** (THAY ĐỔI) Bỏ biểu đồ lặp, chỉ vẽ 1 lần ***
                    if plot_subj > 0 and pivot.shape[0] > 0:
                        samp = pivot.sample(min(plot_subj, pivot.shape[0]), random_state=42)
                        for _, row in samp.iterrows():
                            fig.add_trace(go.Scatter(x=levels, y=row.values, mode="lines", opacity=0.25, showlegend=False))
                    st.plotly_chart(fig, use_container_width=True)
    
                    st.success("**Kết luận:** xem p-value của within-factor trong bảng; p<0.05 ⇒ có khác biệt giữa các điều kiện.")

    # ====================== NONPARAMETRIC (UI Tinh giản) ======================
    with tab_np:
        mode = st.radio("Testing", ["Independent (between)", "Repeated (within)"], horizontal=True, key="np_mode")

        # ---------- Independent (between) ----------
        if mode == "Independent (between)":
            if len(NUM_COLS) == 0 or len(CAT_COLS) == 0:
                st.info("Cần 1 numeric (Y) và 1 categorical (group).")
            else:
                box_top_np = st.container(border=True)
                with box_top_np:
                    st.markdown("### Nonparametric — Independent")
                    with st.popover("ⓘ Hướng dẫn"):
                        st.markdown(
                            "**Independent (between):**\n\n"
                            "So sánh các nhóm *khác biệt, độc lập* với nhau (dùng Median).\n\n"
                            "*Ví dụ: So sánh doanh số giữa Region A, B, C khi dữ liệu bị lệch (skewed).* \n\n"
                            "• 2 nhóm: **Mann–Whitney U**.\n\n"
                            "• 3+ nhóm: **Kruskal–Wallis**."
                        )
                    y_col = st.selectbox("🎯 Y (numeric)", NUM_COLS, key="np_y")
                    g_col = st.selectbox("🏷️ Group (categorical)", CAT_COLS, key="np_g")
                    _type_hint("Y", y_col, "numeric")
                    _type_hint("Group", g_col, "categorical")

                box_ctl_np = st.container(border=True)
                with box_ctl_np:
                    # *** (THAY ĐỔI) Ẩn cài đặt nâng cao ***
                    topN = 10
                    max_fit = 300_000
                    run = st.button("▶️ Run", use_container_width=True, key="np_run")

                if run:
                    sub = DF[[y_col, g_col]].copy()
                    if len(sub) > max_fit:
                        sub = sub.sample(max_fit, random_state=42)
                    sub[g_col] = topn_cat(sub[g_col], n=topN)
                    y = pd.to_numeric(sub[y_col], errors="coerce")
                    g = sub[g_col].astype(str)
                    ok = (~y.isna()) & (~g.isna())
                    y, g = y[ok], g[ok]

                    summ = group_summary(y, g).sort_values("median", ascending=False)
                    st.dataframe(summ, use_container_width=True, hide_index=True)

                    groups = [y[g == lv].values for lv in summ["group"]]
                    k = len(groups)
                    n = int(sum(len(arr) for arr in groups))

                    if k == 2:
                        # Mann–Whitney U
                        ures = stats.mannwhitneyu(groups[0], groups[1], alternative="two-sided")
                        p = float(ures.pvalue); U = float(ures.statistic)
                        z = float(stats.norm.isf(p / 2.0)) if p > 0 else np.inf
                        r_eff = z / np.sqrt(n) if n > 0 and np.isfinite(z) else np.nan
                        st.markdown(f"**Mann–Whitney U**: U = {U:.3f}, p = {p:.4g}, r ≈ {r_eff:.3f} (effect size)")

                        fig = px.violin(pd.DataFrame({g_col: g, y_col: y}), x=g_col, y=y_col,
                                        box=True, points=False)
                        st.plotly_chart(fig, use_container_width=True)

                        hi = str(summ.iloc[0]['group']) if len(summ) else "—"
                        level = ("mạnh" if (not np.isnan(r_eff) and r_eff >= 0.5)
                                 else "vừa" if (not np.isnan(r_eff) and r_eff >= 0.3) else "yếu")
                        st.success(f"**Kết luận:** Khác biệt **{level}** (r≈{r_eff:.2f}). Nhóm median cao nhất: **{hi}**.")
                    
                    elif k > 2:
                        # Kruskal–Wallis
                        H, p = stats.kruskal(*groups)
                        eps2 = kruskal_eps2(H, k, n)
                        st.markdown(f"**Kruskal–Wallis**: H = {H:.3f}, p = {p:.4g}, ε² = {eps2:.3f} (effect size)")

                        # (Biểu đồ bar cho 3+ nhóm là phù hợp)
                        fig = go.Figure(go.Bar(x=summ["group"], y=summ["median"],
                                               error_y=dict(array=summ["ci95"], visible=True)))
                        fig.update_layout(yaxis_title=f"{y_col} (median ± 95% CI≈)")
                        st.plotly_chart(fig, use_container_width=True)

                        # Post-hoc: pairwise Mann–Whitney + Holm
                        pvals, labs = [], []
                        for i in range(k):
                            for j in range(i+1, k):
                                u = stats.mannwhitneyu(groups[i], groups[j], alternative="two-sided")
                                pvals.append(float(u.pvalue))
                                labs.append(f"{summ['group'].iloc[i]} vs {summ['group'].iloc[j]}")
                        if pvals:
                            adj = holm_bonferroni(np.array(pvals), np.array(labs))
                            st.dataframe(adj.head(50), use_container_width=True, hide_index=True)
                            st.caption("Pairwise Mann–Whitney (Holm-adjusted).")

                        strength = ("yếu" if (np.isnan(eps2) or eps2 < 0.06)
                                    else ("vừa" if eps2 < 0.14 else "mạnh"))
                        hi = str(summ.iloc[0]["group"]) if len(summ) else "—"
                        st.success(f"**Kết luận:** Khác biệt **{strength}** (ε²={eps2:.2f}). Nhóm median cao nhất: **{hi}**.")
                    else:
                        st.warning("Cần ít nhất 2 nhóm để so sánh.")

        # ---------- Repeated (within) ----------
        else:
            cand_id = [c for c in DF.columns if is_cat(c)]
            cand_factor = [c for c in CAT_COLS]
            if len(NUM_COLS) == 0 or len(cand_id) == 0 or len(cand_factor) == 0:
                st.info("Cần: 1 numeric (Y), 1 ID (subject), 1 categorical (condition).")
            else:
                box_top_r = st.container(border=True)
                with box_top_r:
                    st.markdown("### Nonparametric — Repeated (within)")
                    with st.popover("ⓘ Hướng dẫn"):
                        st.markdown(
                            "**Repeated (within):**\n\n"
                            "So sánh *cùng một đối tượng* qua nhiều điều kiện (dùng Median).\n\n"
                            "*Ví dụ: So sánh doanh số (bị lệch) của từng Cửa hàng vào Tháng 1, 2, 3.*\n\n"
                            "• 2 điều kiện: **Wilcoxon**.\n\n"
                            "• 3+ điều kiện: **Friedman**."
                        )
                    y_col = st.selectbox("🎯 Y (numeric)", NUM_COLS, key="rep_y")
                    id_col = st.selectbox("🧑‍🤝‍🧑 ID (subject)", cand_id, key="rep_id")
                    
                    # *** (THAY ĐỔI) Lọc thông minh: Condition không thể là ID ***
                    cond_choices = [c for c in cand_factor if c != id_col]
                    if not cond_choices:
                        st.warning("Không đủ cột categorical khác để làm Condition.")
                        st.stop()
                    else:
                        cond_col = st.selectbox("🏷️ Condition (within)", cond_choices, key="rep_cond")

                    _type_hint("Y", y_col, "numeric")
                    _type_hint("ID", id_col, "categorical")
                    _type_hint("Condition", cond_col, "categorical")

                box_ctl_r = st.container(border=True)
                with box_ctl_r:
                    # *** (THAY ĐỔI) Ẩn cài đặt nâng cao ***
                    max_subj_fit = 5_000
                    plot_subj = 80
                    run = st.button("▶️ Run", use_container_width=True, key="rep_run")

                if run:
                    d0 = DF[[y_col, id_col, cond_col]].dropna().copy()
                    count = d0.groupby([id_col, cond_col]).size().unstack(cond_col).dropna()
                    subj_keep = count.index
                    d = d0[d0[id_col].isin(subj_keep)]

                    uniq_ids = d[id_col].unique()
                    if len(uniq_ids) > max_subj_fit:
                        keep = pd.Index(uniq_ids).sample(max_subj_fit, random_state=42)
                        d = d[d[id_col].isin(keep)]

                    pivot = d.pivot_table(index=id_col, columns=cond_col, values=y_col, aggfunc="mean")
                    pivot = pivot.dropna(axis=0)
                    levels = list(pivot.columns)
                    m = len(levels); n = pivot.shape[0]

                    if m == 2:
                        a = pivot[levels[0]].values
                        b = pivot[levels[1]].values
                        res = stats.wilcoxon(a, b, zero_method="wilcox", correction=False, alternative="two-sided", mode="auto")
                        p = float(res.pvalue); Wstat = float(res.statistic)
                        z = float(stats.norm.isf(p/2.0)) if p > 0 else np.inf
                        r_eff = z / np.sqrt(n) if n > 0 and np.isfinite(z) else np.nan
                        st.markdown(f"**Wilcoxon signed-rank**: W = {Wstat:.3f}, p = {p:.4g}, r ≈ {r_eff:.3f} (effect size)")

                        means = pivot.mean().reset_index()
                        means.columns = ["cond","mean"]
                        fig = px.line(means, x="cond", y="mean", markers=True)
                        
                        if plot_subj > 0:
                            samp = pivot.sample(min(plot_subj, pivot.shape[0]), random_state=42)
                            for _, row in samp.iterrows():
                                fig.add_trace(go.Scatter(x=levels, y=row.values, mode="lines", opacity=0.25, showlegend=False))
                        st.plotly_chart(fig, use_container_width=True)

                        level = ("mạnh" if (not np.isnan(r_eff) and r_eff >= 0.5)
                                 else "vừa" if (not np.isnan(r_eff) and r_eff >= 0.3) else "yếu")
                        st.success(f"**Kết luận:** Khác biệt **{level}** (r≈{r_eff:.2f}).")

                    elif m > 2:
                        fr = stats.friedmanchisquare(*[pivot[c].values for c in levels])
                        chi2 = float(fr.statistic); p = float(fr.pvalue)
                        W = chi2 / (n * m * (m + 1) / 12.0) if n > 0 else np.nan
                        st.markdown(f"**Friedman**: χ² = {chi2:.3f}, p = {p:.4g}, W = {W:.3f} (Kendall's W effect size)")

                        means = pivot.mean().reset_index()
                        means.columns = ["cond","mean"]
                        fig = px.line(means, x="cond", y="mean", markers=True)

                        if plot_subj > 0:
                            samp = pivot.sample(min(plot_subj, pivot.shape[0]), random_state=42)
                            for _, row in samp.iterrows():
                                fig.add_trace(go.Scatter(x=levels, y=row.values, mode="lines", opacity=0.25, showlegend=False))
                        st.plotly_chart(fig, use_container_width=True)

                        # Post-hoc: pairwise Wilcoxon + Holm
                        pvals, labs = [], []
                        for i in range(m):
                            for j in range(i+1, m):
                                wi = stats.wilcoxon(pivot[levels[i]], pivot[levels[j]],
                                                    zero_method="wilcox", correction=False,
                                                    alternative="two-sided", mode="auto")
                                pvals.append(float(wi.pvalue))
                                labs.append(f"{levels[i]} vs {levels[j]}")
                        if pvals:
                            adj = holm_bonferroni(np.array(pvals), np.array(labs))
                            st.dataframe(adj.head(50), use_container_width=True, hide_index=True)
                            st.caption("Pairwise Wilcoxon (Holm-adjusted).")

                        strength = ("yếu" if (np.isnan(W) or W < 0.1) else ("vừa" if W < 0.3 else "mạnh"))
                        best = str(means.sort_values("mean", ascending=False).iloc[0]["cond"])
                        st.success(f"**Kết luận:** Khác biệt **{strength}** (W={W:.2f}). Điều kiện cao nhất: **{best}**.")
                    else:
                        st.warning("Cần ít nhất 2 điều kiện để so sánh.")
## ============================== TAB 6 : REGRESSION (Predictive & Audit) ==============================
with TAB6:
    from sklearn.model_selection import train_test_split
    from sklearn.linear_model import LinearRegression, LogisticRegression
    from sklearn.metrics import mean_squared_error, r2_score, accuracy_score, confusion_matrix, classification_report, roc_auc_score, roc_curve
    from sklearn.preprocessing import StandardScaler
    import plotly.express as px
    import plotly.graph_objects as go
    import pandas as pd
    import numpy as np

    st.subheader("🔮 Regression Analysis (Dự báo & Phát hiện Bất thường)")

    # --- 0. Data Source ---
    df_root = SS.get('df')
    
    if df_root is None or df_root.empty:
        st.info("Chưa có dữ liệu."); st.stop()

    # --- 1. Local Drill-down Filter (ĐÃ SỬA: Flexible Selection) ---
    def _local_drilldown(df_in):
        with st.expander("🔎 Bộ lọc dữ liệu (Drill-down Filter)", expanded=False):
            st.caption("Chọn cột và giá trị để khoanh vùng dữ liệu trước khi chạy mô hình.")
            
            # Lấy danh sách tất cả các cột để người dùng tự chọn
            all_cols = ["—"] + list(df_in.columns)
            mask = pd.Series(True, index=df_in.index)

            # Hàng 1: Checkbox kích hoạt bộ lọc
            c1, c2, c3, c4, c5 = st.columns(5)
            use_1 = c1.checkbox("Filter 1 (Region/Loc)", key="reg_chk_1")
            use_2 = c2.checkbox("Filter 2 (Channel)", key="reg_chk_2")
            use_3 = c3.checkbox("Filter 3 (Product)", key="reg_chk_3")
            use_4 = c4.checkbox("Filter 4 (Customer)", key="reg_chk_4")
            use_t = c5.checkbox("Time Filter", key="reg_chk_t", value=True)

            # Layout cho phần chọn cột và giá trị
            r1, r2 = st.columns([1.5, 2.5])

            # Hàm helper để render cặp Selectbox (Cột) + Multiselect (Giá trị)
            def _render_selector(label, use_flag, keyword, key_suffix):
                col_name = None
                if use_flag:
                    with r1:
                        # Tự động tìm index mặc định nếu tên cột khớp từ khóa
                        def_idx = 0
                        if keyword:
                            for i, c in enumerate(all_cols):
                                if keyword.lower() in str(c).lower():
                                    def_idx = i; break
                        
                        col_name = st.selectbox(f"Chọn Cột ({label})", all_cols, index=def_idx, key=f"reg_col_{key_suffix}")
                    
                    if col_name and col_name != "—":
                        with r2:
                            # Lấy Top 200 giá trị để tránh lag UI
                            top_vals = df_in[col_name].astype(str).value_counts().head(200).index.tolist()
                            vals = st.multiselect(f"Giá trị ({col_name})", top_vals, key=f"reg_val_{key_suffix}")
                            return col_name, vals
                return None, []

            # Render 4 bộ lọc Categorical
            c1_n, v1 = _render_selector("Vị trí", use_1, "region", "1")
            if c1_n and v1: mask &= df_in[c1_n].astype(str).isin(v1)

            c2_n, v2 = _render_selector("Kênh", use_2, "channel", "2")
            if c2_n and v2: mask &= df_in[c2_n].astype(str).isin(v2)

            c3_n, v3 = _render_selector("Sản phẩm", use_3, "prod", "3")
            if c3_n and v3: mask &= df_in[c3_n].astype(str).isin(v3)

            c4_n, v4 = _render_selector("Khách hàng", use_4, "cust", "4")
            if c4_n and v4: mask &= df_in[c4_n].astype(str).isin(v4)

            # Xử lý riêng cho Time Filter
            if use_t:
                with r1:
                    # Gợi ý cột datetime
                    dt_cands = [c for c in df_in.columns if 'date' in str(c).lower() or 'time' in str(c).lower()]
                    dt_opts = ["—"] + dt_cands + [c for c in df_in.columns if c not in dt_cands]
                    time_col = st.selectbox("Chọn Cột Thời gian", dt_opts, key="reg_col_time")
                
                if time_col and time_col != "—":
                    with r2:
                        try:
                            ts = pd.to_datetime(df_in[time_col], errors='coerce')
                            # Gom nhóm theo Tháng (Month) để lọc cho gọn
                            periods = sorted(ts.dt.to_period("M").astype(str).dropna().unique())
                            def_sel = periods[-3:] if len(periods) > 3 else periods
                            sel_t = st.multiselect(f"Chọn Kỳ (Tháng)", periods, default=def_sel, key="reg_val_time")
                            
                            if sel_t:
                                mask &= ts.dt.to_period("M").astype(str).isin(sel_t)
                        except:
                            st.warning(f"Cột '{time_col}' không chuẩn định dạng ngày tháng.")

            return df_in.loc[mask].copy()

    # Áp dụng lọc
    df_reg = _local_drilldown(df_root)
    
    # Chỉ hiện số dòng nếu dữ liệu thay đổi hoặc người dùng quan tâm
    if len(df_reg) < len(df_root):
        st.caption(f"⚡ Dữ liệu phân tích: **{len(df_reg):,}** dòng (đã lọc từ {len(df_root):,}).")

    if df_reg.empty: st.warning("Dữ liệu rỗng sau khi lọc."); st.stop()

    # --- 2. Cấu hình Mô hình ---
    nums = df_reg.select_dtypes(include=[np.number]).columns.tolist()
    
    c1, c2, c3 = st.columns([1.2, 2, 1])
    target_col = c1.selectbox("🎯 Biến mục tiêu (Y)", ["—"] + nums, key="reg_y")
    
    # Khởi tạo biến Session State để lưu mô hình (FIX LỖI RELOAD)
    if 'reg_result' not in SS: SS['reg_result'] = None

    if target_col and target_col != "—":
        candidate_x = [c for c in nums if c != target_col]
        feat_cols = c2.multiselect("Biến độc lập (X)", candidate_x, default=candidate_x[:3], key="reg_x")
        
        # Xác định loại bài toán
        is_binary = (df_reg[target_col].nunique() == 2)
        model_type = "Logistic Regression" if is_binary else "Linear Regression"
        
        # Nút chạy mô hình
        if feat_cols:
            if c3.button("🚀 Chạy Mô hình", type="primary", use_container_width=True):
                try:
                    # A. Chuẩn bị dữ liệu
                    XY = df_reg[[target_col] + feat_cols].dropna()
                    if XY.empty: st.error("Dữ liệu rỗng."); st.stop()
                    
                    X = XY[feat_cols]
                    y = XY[target_col]

                    # B. Chia & Scale (Fix Data Leakage)
                    X_train, X_test, y_train, y_test = train_test_split(
                        X, y, test_size=0.2, random_state=42, stratify=(y if is_binary else None)
                    )
                    scaler = StandardScaler()
                    X_train_scaled = scaler.fit_transform(X_train)
                    X_test_scaled = scaler.transform(X_test)
                    
                    X_train_df = pd.DataFrame(X_train_scaled, columns=feat_cols, index=X_train.index)
                    X_test_df = pd.DataFrame(X_test_scaled, columns=feat_cols, index=X_test.index)

                    # C. Huấn luyện
                    if not is_binary:
                        model = LinearRegression()
                        model.fit(X_train_df, y_train)
                        y_pred = model.predict(X_test_df)
                        metrics = {
                            "r2": r2_score(y_test, y_pred),
                            "rmse": np.sqrt(mean_squared_error(y_test, y_pred)),
                            "mae": np.mean(np.abs(y_test - y_pred))
                        }
                    else:
                        model = LogisticRegression(class_weight='balanced', max_iter=1000)
                        model.fit(X_train_df, y_train)
                        y_pred = model.predict(X_test_df)
                        y_prob = model.predict_proba(X_test_df)[:, 1]
                        metrics = {
                            "acc": accuracy_score(y_test, y_pred),
                            "roc": roc_auc_score(y_test, y_prob) if len(np.unique(y_test)) > 1 else 0.5,
                            "y_prob": y_prob
                        }

                    # D. LƯU VÀO SESSION STATE (QUAN TRỌNG ĐỂ KHÔNG BỊ MẤT KHI BẤM WHAT-IF)
                    SS['reg_result'] = {
                        "model": model,
                        "scaler": scaler,
                        "is_binary": is_binary,
                        "features": feat_cols,
                        "target": target_col,
                        "metrics": metrics,
                        "y_test": y_test,
                        "y_pred": y_pred,
                        "model_type": model_type,
                        "XY_mean": XY.mean() # Dùng cho gợi ý What-if
                    }
                    if is_binary: SS['reg_result']['y_prob'] = y_prob

                except Exception as e:
                    st.error(f"Lỗi khi chạy mô hình: {str(e)}")

    # --- 4. HIỂN THỊ KẾT QUẢ (Lấy từ Session State) ---
    res = SS.get('reg_result')
    
    # Chỉ hiển thị nếu đã có kết quả (và người dùng chưa đổi biến mục tiêu khác)
    if res and res['target'] == target_col:
        st.divider()
        st.markdown(f"### 📊 Kết quả: {res['model_type']}")
        
        # 4.1 Metrics
        m = res['metrics']
        c1, c2, c3 = st.columns(3)
        if not res['is_binary']:
            c1.metric("R² (Độ chính xác)", f"{m['r2']:.3f}", help="Càng gần 1 càng tốt")
            c2.metric("RMSE (Sai số chuẩn)", f"{m['rmse']:,.2f}")
            c3.metric("MAE (Sai số tuyệt đối)", f"{m['mae']:,.2f}")
        else:
            c1.metric("Accuracy", f"{m['acc']:.1%}")
            c2.metric("ROC-AUC", f"{m['roc']:.3f}")

        # 4.2 Coefficients & Equation
        model = res['model']
        feats = res['features']
        
        if not res['is_binary']:
            st.markdown("#### 📐 Phân tích Tác động")
            coef_df = pd.DataFrame({
                "Biến số": feats,
                "Hệ số (Beta)": model.coef_,
                "Độ lớn tác động": [abs(c) for c in model.coef_]
            }).sort_values("Độ lớn tác động", ascending=False)
            
            # Thêm Intercept
            intercept_df = pd.DataFrame([{"Biến số": "Hằng số (Intercept)", "Hệ số (Beta)": model.intercept_, "Độ lớn tác động": 0}])
            full_df = pd.concat([intercept_df, coef_df], ignore_index=True)
            
            c_tbl, c_txt = st.columns([3, 2])
            with c_tbl:
                st.dataframe(full_df.drop(columns=["Độ lớn tác động"]).style.format({"Hệ số (Beta)": "{:,.4f}"}), use_container_width=True, hide_index=True)
                # Phương trình (chuyển xuống dưới bảng)
                eq_str = f"$Y = {model.intercept_:.2f} " + " ".join([f"{'+' if c>=0 else '-'} {abs(c):.2f} \\cdot X_{i+1}" for i,c in enumerate(model.coef_)]) + "$"
                st.info(f"**Phương trình:**\n\n{eq_str}")
            
            with c_txt:
                # Business Interpretation (giữ nguyên)
                interpret_text = f"**💡 Diễn giải Nghiệp vụ (Business Interpretation):**\n\n"
                interpret_text += f"Để dự báo **{target_col}**, mô hình bắt đầu từ mức cơ bản là **{model.intercept_:,.2f}**. Sau đó, dựa trên dữ liệu quá khứ, ta thấy các yếu tố tác động mạnh nhất là:\n"
                
                # Lấy Top 3 biến quan trọng nhất để diễn giải
                top_drivers = coef_df.head(3)
                for _, row in top_drivers.iterrows():
                    direction = "tăng thêm" if row['Hệ số (Beta)'] > 0 else "giảm đi"
                    # Lưu ý về Standardized Unit
                    interpret_text += f"- **{row['Biến số']}**: Khi biến này tăng (1 độ lệch chuẩn), **{target_col}** sẽ **{direction} khoảng {abs(row['Hệ số (Beta)']):,.2f}** đơn vị.\n"
                
                interpret_text += "\n*(Lưu ý: Các hệ số được tính trên dữ liệu đã chuẩn hóa để so sánh công bằng giữa các đơn vị khác nhau)*"
                
                st.info(interpret_text)

            # 4.3 Residuals Audit
            st.markdown("---")
            st.markdown(f"#### 🕵️ Phân tích Bất thường (Residuals Audit)")
            y_test, y_pred = res['y_test'], res['y_pred']
            residuals = y_test - y_pred
            
            res_df = pd.DataFrame({
                f"Thực tế ({res['target']})": y_test,
                f"Dự báo ({res['target']})": y_pred,
                "Độ lệch (Error)": residuals,
                "Sai số %": (residuals / (y_test+0.001) * 100).round(1)
            })
            top_outliers = res_df.reindex(res_df["Độ lệch (Error)"].abs().sort_values(ascending=False).index).head(50)
            
            c_g1, c_g2 = st.columns([2, 1])
            with c_g1:
                fig = px.scatter(x=y_test, y=y_pred, labels={'x': 'Thực tế', 'y': 'Dự báo'}, title="Biểu đồ Thực tế vs Dự báo")
                fig.add_shape(type="line", x0=y_test.min(), y0=y_test.min(), x1=y_test.max(), y1=y_test.max(), line=dict(color="Red", dash="dash"))
                st.plotly_chart(fig, use_container_width=True)
                st.caption("Điểm nằm càng xa đường đỏ nét đứt là các giao dịch có rủi ro cao (Mô hình không giải thích được).")
            
            with c_g2:
                st.markdown("**Top giao dịch lệch nhiều nhất:**")
                st.dataframe(top_outliers.style.format("{:,.2f}"), use_container_width=True, height=400)
            # THÊM VÀO CUỐI PHẦN 4.3 Residuals Audit (dưới c_g2):

            # Tính toán các chỉ số rủi ro cần thiết
            res_mean = residuals.mean()
            res_std = residuals.std()
            
            # Đếm số lượng giao dịch vượt ngưỡng 2 Sigma
            # 2 Sigma là ngưỡng phổ biến cho bất thường trong kiểm toán
            two_sigma = 2 * res_std
            outlier_mask = (residuals.abs() >= two_sigma)
            n_high_risk = outlier_mask.sum()
            
            st.markdown("---")
            st.markdown("#### 🚨 Nhận định Rủi ro về Độ lệch (Residual Risk Assessment)")
            
            with st.container(border=True):
                st.markdown(f"""
                * **Độ chính xác Trung bình (MAE):** {res['metrics']['mae']:,.2f}
                * **Sai số Chuẩn (Std Dev of Residuals):** {res_std:,.2f}
                * **Ngưỡng Rủi ro Cao (±2σ):** ±{two_sigma:,.2f} (Giá trị lệch so với dự báo lớn hơn ngưỡng này)
                """)
                
                if n_high_risk > 0:
                    st.error(f"**🚨 Cảnh báo Đỏ:** Phát hiện **{n_high_risk:,}** giao dịch có độ lệch vượt quá **±2 Std Dev** ({n_high_risk / len(y_test) * 100:.2f}% mẫu kiểm tra).")
                    st.markdown(f"👉 **Hành động Audit:** Các giao dịch này đại diện cho **rủi ro cao nhất** vì mô hình không thể giải thích được hành vi của chúng. Trích xuất **Top 50** trong bảng bên cạnh để kiểm tra chi tiết.")
                else:
                    st.success("🟢 Độ lệch Residuals đang ở mức kiểm soát. Không có ngoại lai vượt ngưỡng 2σ rõ rệt.")
        else:
            # Binary charts (ROC, Confusion Matrix) - Giữ nguyên code logic cũ nếu cần
            st.markdown("#### 📐 Hệ số Log-Odds")
            coef_df = pd.DataFrame({
                "Feature": feats,
                "Log-Odds Coeff": model.coef_[0]
            }).sort_values("Log-Odds Coeff", key=abs, ascending=False)
            st.dataframe(coef_df.style.format("{:,.4f}"), use_container_width=True)

            st.markdown("#### Confusion Matrix & ROC")
            c_conf, c_roc = st.columns(2)
            with c_conf:
                cm = confusion_matrix(y_test, y_pred)
                fig_cm = px.imshow(cm, text_auto=True, labels=dict(x="Dự báo", y="Thực tế"), x=['False', 'True'], y=['False', 'True'], color_continuous_scale='Blues')
                st.plotly_chart(fig_cm, use_container_width=True)
            with c_roc:
                fpr, tpr, _ = roc_curve(y_test, y_prob)
                fig_roc = px.area(x=fpr, y=tpr, title=f'ROC Curve (AUC={m["roc"]:.2f})', labels=dict(x='False Positive Rate', y='True Positive Rate'))
                fig_roc.add_shape(type='line', line=dict(dash='dash'), x0=0, x1=1, y0=0, y1=1)
                st.plotly_chart(fig_roc, use_container_width=True)

        # 4.4 WHAT-IF ANALYSIS (ĐÃ KHẮC PHỤC LỖI RESET)
        st.markdown("---")
        st.subheader("🧮 Giả lập Kịch bản (What-if Simulator)")
        
        with st.form("whatif_form"):
            st.write("Điều chỉnh các thông số đầu vào để dự báo kết quả:")
            cols = st.columns(3)
            input_vals = []
            
            # Tạo input field cho từng biến
            means = res['XY_mean']
            for i, col in enumerate(res['features']):
                default_val = float(means[col]) if col in means else 0.0
                val = cols[i % 3].number_input(f"{col}", value=default_val)
                input_vals.append(val)
            
            submit = st.form_submit_button("🔮 Dự báo ngay")
        
        if submit:
            # Lấy scaler và model từ session state để dự báo
            scaler_saved = res['scaler']
            model_saved = res['model']
            
            # Transform input y hệt như lúc train
            input_scaled = scaler_saved.transform([input_vals])
            
            if not res['is_binary']:
                pred_val = model_saved.predict(input_scaled)[0]
                st.success(f"💰 Giá trị dự báo **{res['target']}**: **{pred_val:,.2f}**")
                
                # So sánh với trung bình
                avg_target = float(means[res['target']])
                diff = pred_val - avg_target
                pct = (diff / avg_target * 100) if avg_target != 0 else 0
                st.caption(f"So với mức trung bình ({avg_target:,.0f}): {'Tăng' if diff>0 else 'Giảm'} **{abs(diff):,.0f}** ({pct:+.1f}%)")
            else:
                pred_prob = model_saved.predict_proba(input_scaled)[0, 1]
                st.success(f"Khả năng thuộc lớp Positive: **{pred_prob:.1%}**")

        else:
            st.info("👈 Vui lòng chọn ít nhất 1 biến độc lập (X) để chạy mô hình.")
# ============================== TAB 7 : PARETO & CONCENTRATION (ABC Analysis) ==============================
with TAB7:
    import numpy as np
    import pandas as pd
    import plotly.express as px
    import plotly.graph_objects as go
    import streamlit as st
    
    # --- Định nghĩa lại hàm _top_values (cần cho Drill-down) ---
    def _top_values(df_local, col, k=200):
        if not col or col not in df_local.columns:
            return pd.Series([], dtype='object')
        return df_local[col].astype(str).value_counts(dropna=False).head(k)
        
    st.subheader("⚖️ Pareto Principle (80/20 Rule) & Concentration Risk")
    
    # --- 1. Data Source & Filter (Giữ nguyên Drill-down) ---
    df = SS.get('df')
    if df is None or df.empty: st.info("Hãy nạp dữ liệu trước."); st.stop()
        
    def _render_filter_inline(df_in, key_prefix="par"):
        # Hàm lọc cục bộ: giữ nguyên logic lọc phức hợp của bạn
        with st.expander("🔎 Drill-down Filter (Bộ lọc dữ liệu)", expanded=False):
            st.caption("Chọn cột và giá trị để khoanh vùng dữ liệu trước khi phân tích.")
            all_cols_in = ["—"] + list(df_in.columns)
            mask = pd.Series(True, index=df_in.index)
            c1, c2, c3, c4, c5 = st.columns(5)
            use_1 = c1.checkbox("Filter 1 (Region)", key=f"{key_prefix}_chk_1")
            use_2 = c2.checkbox("Filter 2 (Channel)", key=f"{key_prefix}_chk_2")
            use_3 = c3.checkbox("Filter 3 (Product)", key=f"{key_prefix}_chk_3")
            use_4 = c4.checkbox("Filter 4 (Customer)", key=f"{key_prefix}_chk_4")
            use_t = c5.checkbox("Time Filter", key=f"{key_prefix}_chk_t", value=False) 
            r1, r2 = st.columns([1.5, 2.5])
            def _render_sel(label, use_flag, keyword, suffix):
                col_name = None
                if use_flag:
                    with r1:
                        def_idx = 0
                        if keyword: 
                            for i, c in enumerate(all_cols_in):
                                if keyword.lower() in str(c).lower(): def_idx = i; break
                        col_name = st.selectbox(f"Chọn Cột ({label})", all_cols_in, index=def_idx, key=f"{key_prefix}_col_{suffix}")
                    if col_name and col_name != "—":
                        with r2:
                            top_vals = _top_values(df_in, col_name).index.tolist()
                            vals = st.multiselect(f"Giá trị ({col_name})", top_vals, key=f"{key_prefix}_val_{suffix}")
                            return col_name, vals
                return None, []
            c1_n, v1 = _render_sel("Vị trí", use_1, "region", "1")
            if c1_n and v1: mask &= df_in[c1_n].astype(str).isin(v1)
            c2_n, v2 = _render_sel("Kênh", use_2, "channel", "2")
            if c2_n and v2: mask &= df_in[c2_n].astype(str).isin(v2)
            c3_n, v3 = _render_sel("Sản phẩm", use_3, "prod", "3")
            if c3_n and v3: mask &= df_in[c3_n].astype(str).isin(v3)
            c4_n, v4 = _render_sel("Khách hàng", use_4, "cust", "4")
            if c4_n and v4: mask &= df_in[c4_n].astype(str).isin(v4)
            if use_t:
                with r1:
                    dt_cands = [c for c in df_in.columns if 'date' in str(c).lower() or 'time' in str(c).lower()]
                    dt_opts = ["—"] + dt_cands + [c for c in df_in.columns if c not in dt_cands]
                    time_col = st.selectbox("Cột Thời gian", dt_opts, key=f"{key_prefix}_col_time")
                if time_col and time_col != "—":
                    with r2:
                        try:
                            ts = pd.to_datetime(df_in[time_col], errors='coerce')
                            years = sorted(ts.dt.year.dropna().unique().astype(int).tolist())
                            sel_y = st.multiselect(f"Chọn Năm", years, default=years, key=f"{key_prefix}_val_year")
                            if sel_y: mask &= ts.dt.year.isin(sel_y)
                        except: st.warning("Lỗi định dạng thời gian.")
            return df_in.loc[mask]

    # --- ÁP DỤNG BỘ LỌC VÀ CẬP NHẬT SOURCE DATA ---
    dfx = _render_filter_inline(df, "par")
    all_cols = list(dfx.columns)
    num_cols = list(dfx.select_dtypes(include=[np.number]).columns)

    # --- 2. Cấu hình & Tính toán ---
    st.markdown("### ⚙️ 1. Cấu hình Phân tích")
    with st.container(border=True):
        c1, c2 = st.columns(2)
        dim_col = c1.selectbox("🔍 Phân tích theo (Dimension)", ["—"] + all_cols, index=0, key="par_dim")
        met_col = c2.selectbox("💰 Giá trị đo lường (Metric)", ["—"] + num_cols, index=0, key="par_met")
        
        c3, c4 = st.columns(2)
        threshold_A = c3.slider("Ngưỡng Nhóm A (Cumulative %)", 50, 90, 80, step=5, key="par_th_a")
        threshold_B = c4.slider("Ngưỡng Nhóm B (Cumulative %)", threshold_A, 99, 95, step=1, key="par_th_b")

    if dim_col != "—" and met_col != "—":
        # --- Calculations ---
        df_agg = dfx.groupby(dim_col)[met_col].sum().reset_index()
        df_agg = df_agg[df_agg[met_col] > 0].copy()
        if df_agg.empty: st.warning("Không có dữ liệu > 0 để phân tích."); st.stop()
        df_agg = df_agg.sort_values(by=met_col, ascending=False).reset_index(drop=True)
        
        total_val = df_agg[met_col].sum()
        df_agg["Share"] = df_agg[met_col] / total_val
        df_agg["CumPct"] = df_agg["Share"].cumsum() * 100.0
        
        def classify_abc(cum_pct):
            if cum_pct <= threshold_A: return "A"
            elif cum_pct <= threshold_B: return "B"
            return "C"
        
        df_agg["Class"] = df_agg["CumPct"].apply(classify_abc)
        if df_agg.loc[0, "CumPct"] > threshold_A: df_agg.loc[0, "Class"] = "A" 

        summary = df_agg.groupby("Class").agg(Count=(dim_col, "count"), Value=(met_col, "sum")).reindex(["A", "B", "C"]).fillna(0)
        summary["Count %"] = summary["Count"] / len(df_agg) * 100
        summary["Value %"] = summary["Value"] / total_val * 100

        n = len(df_agg)
        cum_y = df_agg["CumPct"].values / 100.0
        cum_x = np.arange(1, n + 1) / n
        area_under_curve = np.trapz(cum_y, cum_x)
        gini = 1 - 2 * area_under_curve
        
        df_agg["Share_Sq"] = df_agg["Share"] ** 2
        hhi = float(df_agg["Share_Sq"].sum() * 10000.0)

        if "A" in summary.index:
            cnt_A = int(summary.loc["A", "Count"])
            val_A_pct = summary.loc["A", "Value %"]
        else:
            cnt_A = 0
            val_A_pct = 0.0
        # --- 3. Metrics & Insight ---
        st.markdown("### 📊 2. Kết quả Metrics & Phân tích ABC")
        
        # 3.1. Hiển thị KPIs (HHI/Gini và Group A, B, C)
        # SỬA LỖI TRÙNG LẶP: Đã bỏ Bảng Chi tiết ABC, dồn vào Metrics Card
        
        share_A = summary.loc["A", "Value %"] if "A" in summary.index else 0
        share_B = summary.loc["B", "Value %"] if "B" in summary.index else 0
        share_C = summary.loc["C", "Value %"] if "C" in summary.index else 0
        cnt_A = int(summary.loc["A", "Count"])
        
        k1, k2, k3, k4, k5 = st.columns(5)

        k1.metric("Nhóm A (Vital Few)", f"{cnt_A} items", f"Chiếm {share_A:.1f}% Giá trị")
        k2.metric("Nhóm B", f"{int(summary.loc['B','Count'])} items", f"Chiếm {share_B:.1f}% Giá trị")
        k3.metric("Nhóm C", f"{int(summary.loc['C','Count'])} items", f"Chiếm {share_C:.1f}% Giá trị")
        k4.metric("Hệ số Gini", f"{gini:.3f}", help="Gini > 0.6 là rủi ro tập trung cao.")
        k5.metric("Chỉ số HHI", f"{hhi:,.0f}", help="HHI > 1800 là Mức độ Tập trung Cao.")
        
        
        # --- 4. Visualization (Bố cục gọn, tránh chồng lấn) ---
        st.markdown("### 📈 3. Trực quan hóa (Pareto & Lorenz)")
        
        # 4.1. Pareto Bar và Cumulative % Line (Lorenz Line)
        st.markdown("#### Pareto Bar (Theo Class ABC) & Cumulative % Line")
        MAX_SHOW = 100
        plot_df = df_agg.head(MAX_SHOW).copy() if n > MAX_SHOW else df_agg.copy()

        fig = go.Figure()

        # Cumulative Percentage Line (Lorenz Line) - TRỤC PHẢI (Y2)
        fig.add_trace(go.Scatter(
            x=plot_df[dim_col].astype(str),
            y=plot_df["CumPct"], 
            name="Cumulative Share % (Line)",
            yaxis="y2",
            mode="lines+markers",
            line=dict(color='#4d16b9', width=3),
            # Hovertemplate gọn gàng:
            hovertemplate="**Cumulative Share:** %{y:.1f}%<extra></extra>"
        ))
        
        # Bar chart (Giá trị) - TRỤC TRÁI (Y1)
        fig.add_trace(go.Bar(
            x=plot_df[dim_col].astype(str), y=plot_df[met_col], name=met_col,
            marker_color=plot_df["Class"].map({"A": "#ff7675", "B": "#ffeaa7", "C": "#74b9ff"}),
            text=plot_df["Class"], yaxis="y1",
            hovertemplate="**Value:** %{y:,.0f}<br>**Class:** %{text}<extra></extra>"
        ))
        
        # Thêm đường 80%
        fig.add_hline(y=threshold_A, line_dash="dot", line_color="red", annotation_text=f"Cut-off {threshold_A}% (Group A)", yref="y2")
        fig.update_layout(
            height=500,
            xaxis=dict(title=dim_col, type='category'),
            yaxis=dict(title=met_col, side='left', showgrid=False),
            # Tinh chỉnh Y2: Range 0-110 và suffix %. Giảm khoảng cách titles để tránh chồng
            yaxis2=dict(title="Cumulative %", side='right', overlaying='y', range=[0, 110], ticksuffix="%", title_standoff=0),
            # Legend đặt ở trên cùng, trung tâm:
            legend=dict(x=0.5, y=1.08, xanchor="center", orientation="h"),
            margin=dict(l=20, r=20, t=50, b=20),
            hovermode="x unified"
        )
        st.plotly_chart(fig, use_container_width=True)

        # 4.3. Audit Insight
        risk_hhi = "CAO (HHI > 1800)" if hhi >= 1800 else ("TRUNG BÌNH (HHI > 1000)" if hhi >= 1000 else "THẤP")
        
        st.info(f"""
        **💡 Audit Insight:**
        - **Rủi ro Tập trung (HHI & Gini):** Hệ số HHI là **{hhi:,.0f}** và Gini Index là **{gini:.3f}**. Mức rủi ro tập trung được đánh giá là **{risk_hhi}**. Rủi ro này cần được kiểm tra chi tiết trong nhóm A.
        - **Nhóm A (Vital Few):** Gồm **{cnt_A}** {dim_col} ({(cnt_A/n*100):.1f}% số lượng) nhưng đóng góp **{val_A_pct:.1f}%** tổng {met_col}.
          👉 **Hành động Audit:** Kiểm tra 100% các giao dịch trong nhóm A để đảm bảo tuân thủ chính sách giá/chiết khấu.
        - **Nhóm C (Trivial Many):** Gồm **{int(summary.loc['C', 'Count'])}** {dim_col} nhưng chỉ đóng góp **{summary.loc['C', 'Value %']:.1f}%** giá trị.
          👉 **Hành động Audit:** Tập trung vào kiểm tra tính **gian lận hệ thống** (ví dụ: các giao dịch nhỏ lặp lại) thay vì kiểm tra giá trị giao dịch đơn lẻ.
        """)
        
        # 4.4. Detail Table (Bọc trong Expander)
        with st.expander("📄 Chi tiết phân loại ABC (Danh sách đầy đủ)", expanded=False):
            fil_c = st.radio("Lọc theo nhóm:", ["All", "A (Quan trọng)", "B (Trung bình)", "C (Ít quan trọng)"], horizontal=True, key='fil_abc')
            
            view_df = df_agg.copy()
            if fil_c == "A (Quan trọng)": view_df = view_df[view_df["Class"]=="A"]
            elif fil_c == "B (Trung bình)": view_df = view_df[view_df["Class"]=="B"]
            elif fil_c == "C (Ít quan trọng)": view_df = view_df[view_df["Class"]=="C"]
            
            view_df_show = view_df.copy()
            view_df_show[met_col] = view_df_show[met_col].map(lambda x: f"{x:,.0f}")
            view_df_show["Share"] = view_df_show["Share"].map(lambda x: f"{x*100:.2f}%")
            view_df_show["CumPct"] = view_df_show["CumPct"].map(lambda x: f"{x:.2f}%")
            st.dataframe(view_df_show, use_container_width=True, hide_index=True)

    else:
        st.info("👈 Vui lòng chọn Dimension và Metric ở trên để bắt đầu phân tích.")
